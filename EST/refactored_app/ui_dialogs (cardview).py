"""
This module contains all QDialog-based windows for the application,
such as the settings, search, database manager, and rule manager dialogs.
"""

import sqlite3
import json
import re
import openpyxl

from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLineEdit, QListWidget, QPushButton,
                             QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem, QHBoxLayout, 
                             QFileDialog, QMessageBox, QGroupBox, QFormLayout, QComboBox, 
                             QSpinBox, QHeaderView, QInputDialog, QWidget, QSplitter, 
                             QTreeWidget, QTreeWidgetItem, QLabel, QScrollArea)
from PyQt6.QtCore import Qt

# Import shared constants
from constants import PROPERTY_DATA, FORMULA_VARS

class SearchDialog(QDialog):
    """
    A dialog for searching the materials or labor database and adding
    an item to the estimate.
    """
    def __init__(self, db_type, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Search & Add {db_type}")
        self.setFixedSize(600, 400)
        self.layout = QVBoxLayout(self)
        
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Type to search official items...")
        self.layout.addWidget(self.search_box)
        
        self.list_widget = QListWidget()
        self.layout.addWidget(self.list_widget)
        self.search_box.textChanged.connect(self.filter_list)
        
        self.add_btn = QPushButton("Add Selected to Estimate")
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        self.add_btn.clicked.connect(self.accept)
        self.layout.addWidget(self.add_btn)

        self.items_data = {}
        self.load_data(db_type)

    def load_data(self, db_type):
        """Loads data from the specified database table."""
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        if db_type == "Material":
            cursor.execute("SELECT item_code, item_name, unit, rate FROM materials")
        else:
            cursor.execute("SELECT labor_code, task_name, unit, rate FROM labor")
        
        for row in cursor.fetchall():
            display_text = f"{row[1]} ({row[2]}) - Rs. {row[3]}"
            self.items_data[display_text] = {"code": row[0], "name": row[1], "unit": row[2], "rate": row[3], "type": db_type}
            self.list_widget.addItem(display_text)
        conn.close()

    def filter_list(self, text):
        """Filters the list widget based on the search box text."""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def get_selected(self):
        """Returns the data for the currently selected item."""
        selected = self.list_widget.currentItem()
        if selected:
            return self.items_data[selected.text()]
        return None


class SettingsDialog(QDialog):
    """
    A dialog for accessing advanced application settings like the rule
    engine toggle and the database/ruleset managers.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("Advanced Settings")
        self.setFixedSize(300, 150)
        
        layout = QVBoxLayout(self)
        
        db_sync_btn = QPushButton("🗃️ Master DB (Excel Sync)")
        db_sync_btn.clicked.connect(self.parent_app.open_db_manager)
        layout.addWidget(db_sync_btn)

        rules_btn = QPushButton("🧠 Ruleset Manager")
        rules_btn.clicked.connect(self.parent_app.open_rule_manager)
        layout.addWidget(rules_btn)

        layout.addStretch()


class DatabaseManagerDialog(QDialog):
    """
    A dialog for managing the master database, allowing import from and
    export to Excel files.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Master DB Sync (Excel)")
        self.setGeometry(100, 100, 800, 600)
        
        layout = QVBoxLayout(self)
        
        button_layout = QHBoxLayout()
        import_btn = QPushButton("📥 Import from Excel")
        import_btn.clicked.connect(self.import_from_excel)
        export_btn = QPushButton("📤 Export to Excel")
        export_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(import_btn)
        button_layout.addWidget(export_btn)
        layout.addLayout(button_layout)
        
        self.tabs = QTabWidget()
        self.materials_table = QTableWidget()
        self.labor_table = QTableWidget()
        self.tabs.addTab(self.materials_table, "Materials")
        self.tabs.addTab(self.labor_table, "Labor")
        layout.addWidget(self.tabs)
        
        self.load_table_data()

    def populate_table(self, table_widget, data, headers):
        """Fills a table widget with data."""
        table_widget.setRowCount(0)
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)
        for row_num, row_data in enumerate(data):
            table_widget.insertRow(row_num)
            for col_num, col_data in enumerate(row_data):
                table_widget.setItem(row_num, col_num, QTableWidgetItem(str(col_data)))
        table_widget.resizeColumnsToContents()

    def load_table_data(self):
        """Loads data from the database and populates the tables."""
        self.materials_table.clear()
        self.labor_table.clear()
        
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM materials")
        self.populate_table(self.materials_table, cursor.fetchall(), ["Item Code", "Item Name", "Rate", "Unit"])
        
        cursor.execute("SELECT * FROM labor")
        self.populate_table(self.labor_table, cursor.fetchall(), ["Labor Code", "Task Name", "Rate", "Unit"])
        
        conn.close()

    def export_to_excel(self):
        """Exports the database tables to an Excel file."""
        filename, _ = QFileDialog.getSaveFileName(self, "Export DB to Excel", "master_database.xlsx", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                ws = wb.create_sheet(title=table_name)
                cursor.execute(f"PRAGMA table_info({table_name})")
                headers = [info[1] for info in cursor.fetchall()]
                ws.append(headers)

                cursor.execute(f"SELECT * FROM {table_name}")
                for row in cursor.fetchall():
                    ws.append(row)
            
            conn.close()
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Database exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export database: {e}")

    def import_from_excel(self):
        """Imports data from an Excel file into the database."""
        filename, _ = QFileDialog.getOpenFileName(self, "Import DB from Excel", "", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.load_workbook(filename)
            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                if table_name in wb.sheetnames:
                    ws = wb[table_name]
                    cursor.execute(f"DELETE FROM {table_name}")
                    headers = [cell.value for cell in ws[1]]
                    placeholders = ', '.join(['?'] * len(headers))
                    query = f"INSERT OR REPLACE INTO {table_name} ({', '.join(headers)}) VALUES ({placeholders})"
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            cursor.execute(query, row)
            conn.commit()
            conn.close()
            self.load_table_data()
            QMessageBox.information(self, "Success", "Database imported successfully. Please restart or refresh to see updates in 'Add Custom' lists.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import database: {e}")


class RulesetManagerDialog(QDialog):
    """
    Three-panel Rule Manager:
      Left   — Object type tabs (SmartPole, SmartSpan, SmartHome)
      Centre — Flat card list of all rules for the selected object type
      Right  — Editor panel: structured condition rows + formula + item picker
    Rules are stored as a flat list in rules.json; every rule belongs to
    exactly one object type, so there is no tree-placement ambiguity.
    """

    # Human-readable labels for each object type tab
    OBJECT_LABELS = {
        "SmartPole": "🔵 Pole",
        "SmartSpan": "📏 Span",
        "SmartHome":  "🏠 Home",
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Rule Manager")
        self.setGeometry(100, 100, 1350, 820)

        self.rules = []
        self.selected_rule_index = -1   # index into self.rules
        self.selected_result_item = None
        self.condition_widgets = []
        self.current_object_type = "SmartPole"
        self.property_data = PROPERTY_DATA

        # ── Root layout: three columns ────────────────────────────────────────
        root = QHBoxLayout(self)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)

        # ── LEFT: object-type sidebar ─────────────────────────────────────────
        self.sidebar = QWidget()
        self.sidebar.setFixedWidth(160)
        self.sidebar.setStyleSheet("background: #f5f5f5; border-right: 1px solid #ddd;")
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 8, 0, 8)
        sidebar_layout.setSpacing(2)

        sidebar_title = QLabel("Object types")
        sidebar_title.setStyleSheet("font-size: 10px; color: #888; padding: 4px 14px 8px 14px; font-weight: bold;")
        sidebar_layout.addWidget(sidebar_title)

        self.tab_buttons = {}
        for obj_type, label in self.OBJECT_LABELS.items():
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setStyleSheet(self._tab_style(False))
            btn.clicked.connect(lambda checked, t=obj_type: self.switch_object_type(t))
            sidebar_layout.addWidget(btn)
            self.tab_buttons[obj_type] = btn

        sidebar_layout.addStretch()
        root.addWidget(self.sidebar)

        # ── CENTRE: rule card list ────────────────────────────────────────────
        centre = QWidget()
        centre.setMinimumWidth(380)
        centre_layout = QVBoxLayout(centre)
        centre_layout.setContentsMargins(0, 0, 0, 0)
        centre_layout.setSpacing(0)

        # Topbar
        topbar = QWidget()
        topbar.setStyleSheet("border-bottom: 1px solid #ddd; background: white;")
        topbar_layout = QHBoxLayout(topbar)
        topbar_layout.setContentsMargins(12, 8, 12, 8)
        self.centre_title = QLabel("")
        self.centre_title.setStyleSheet("font-weight: bold; font-size: 13px;")
        new_btn = QPushButton("+ New rule")
        new_btn.setStyleSheet(
            "background: #185FA5; color: white; border: none; padding: 5px 14px;"
            "border-radius: 4px; font-size: 12px;"
        )
        new_btn.clicked.connect(self.create_new_rule)
        topbar_layout.addWidget(self.centre_title)
        topbar_layout.addStretch()
        topbar_layout.addWidget(new_btn)
        centre_layout.addWidget(topbar)

        # Scrollable card area
        self.card_scroll = QWidget()
        self.card_scroll.setStyleSheet("background: #fafafa;")
        self.card_list_layout = QVBoxLayout(self.card_scroll)
        self.card_list_layout.setContentsMargins(10, 10, 10, 10)
        self.card_list_layout.setSpacing(6)
        self.card_list_layout.addStretch()

        scroll_area = QScrollArea()
        scroll_area.setWidget(self.card_scroll)
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QScrollArea.Shape.NoFrame)
        centre_layout.addWidget(scroll_area)
        root.addWidget(centre)

        # ── RIGHT: editor panel ───────────────────────────────────────────────
        self.editor_outer = QWidget()
        self.editor_outer.setFixedWidth(420)
        self.editor_outer.setStyleSheet("border-left: 1px solid #ddd; background: white;")
        self.editor_outer_layout = QVBoxLayout(self.editor_outer)
        self.editor_outer_layout.setContentsMargins(0, 0, 0, 0)
        self.editor_outer_layout.setSpacing(0)

        self.editor_header = QLabel("Select a rule to edit")
        self.editor_header.setStyleSheet(
            "font-weight: bold; font-size: 13px; padding: 10px 16px;"
            "border-bottom: 1px solid #ddd; background: white;"
        )
        self.editor_outer_layout.addWidget(self.editor_header)

        # Scroll area for editor body
        self.editor_body_widget = QWidget()
        self.editor_body_layout = QVBoxLayout(self.editor_body_widget)
        self.editor_body_layout.setContentsMargins(16, 12, 16, 12)
        self.editor_body_layout.setSpacing(12)
        self.editor_body_layout.addStretch()

        editor_scroll = QScrollArea()
        editor_scroll.setWidget(self.editor_body_widget)
        editor_scroll.setWidgetResizable(True)
        editor_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        self.editor_outer_layout.addWidget(editor_scroll)

        # Fixed footer with Save / Delete
        editor_footer = QWidget()
        editor_footer.setStyleSheet("border-top: 1px solid #ddd; background: white;")
        footer_layout = QHBoxLayout(editor_footer)
        footer_layout.setContentsMargins(12, 8, 12, 8)
        self.delete_btn = QPushButton("🗑 Delete rule")
        self.delete_btn.setStyleSheet(
            "color: #c0392b; border: 1px solid #c0392b; padding: 5px 12px;"
            "border-radius: 4px; background: white;"
        )
        self.delete_btn.clicked.connect(self.delete_selected_rule)
        self.delete_btn.setEnabled(False)
        self.save_btn = QPushButton("💾 Save rule")
        self.save_btn.setStyleSheet(
            "background: #27ae60; color: white; border: none; padding: 5px 14px;"
            "border-radius: 4px; font-weight: bold;"
        )
        self.save_btn.clicked.connect(self.save_rule_changes)
        self.save_btn.setEnabled(False)
        footer_layout.addWidget(self.delete_btn)
        footer_layout.addStretch()
        footer_layout.addWidget(self.save_btn)
        self.editor_outer_layout.addWidget(editor_footer)

        root.addWidget(self.editor_outer)

        self.load_rules()
        self.switch_object_type("SmartPole")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _tab_style(self, active: bool) -> str:
        if active:
            return (
                "text-align: left; padding: 9px 14px; border: none;"
                "border-left: 3px solid #185FA5; background: white;"
                "font-weight: bold; font-size: 12px;"
            )
        return (
            "text-align: left; padding: 9px 14px; border: none;"
            "border-left: 3px solid transparent; background: transparent;"
            "font-size: 12px;"
        )

    def _section_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "font-size: 10px; font-weight: bold; color: #888; text-transform: uppercase;"
            "padding-bottom: 4px; border-bottom: 1px solid #eee; letter-spacing: 0.05em;"
        )
        return lbl

    def _clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self._clear_layout(child.layout())

    def _rules_for_type(self, obj_type: str):
        """Returns list of (original_index, rule) for the given object type."""
        return [(i, r) for i, r in enumerate(self.rules) if r.get("object") == obj_type]

    # ── Left sidebar: switching tabs ──────────────────────────────────────────

    def switch_object_type(self, obj_type: str):
        self.current_object_type = obj_type
        self.selected_rule_index = -1
        for t, btn in self.tab_buttons.items():
            btn.setChecked(t == obj_type)
            btn.setStyleSheet(self._tab_style(t == obj_type))
        self.refresh_card_list()
        self.clear_editor()

    # ── Centre: card list ─────────────────────────────────────────────────────

    def refresh_card_list(self):
        """Rebuild the rule card list for the currently selected object type."""
        # Remove all cards (everything except the trailing stretch)
        while self.card_list_layout.count() > 1:
            item = self.card_list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        rules = self._rules_for_type(self.current_object_type)
        label = self.OBJECT_LABELS.get(self.current_object_type, self.current_object_type)
        self.centre_title.setText(f"{label}  —  {len(rules)} rule{'s' if len(rules) != 1 else ''}")

        for orig_idx, rule in rules:
            card = self._make_rule_card(orig_idx, rule)
            self.card_list_layout.insertWidget(self.card_list_layout.count() - 1, card)

    def _make_rule_card(self, rule_index: int, rule: dict) -> QWidget:
        """Creates a single clickable rule card widget."""
        card = QWidget()
        card.setCursor(Qt.CursorShape.PointingHandCursor)
        card.setObjectName(f"card_{rule_index}")

        is_selected = (rule_index == self.selected_rule_index)
        card.setStyleSheet(self._card_style(is_selected))

        layout = QHBoxLayout(card)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(10)

        # Type badge (M / L)
        r_type = rule.get("type", "Material")
        badge = QLabel("M" if r_type == "Material" else "L")
        badge.setFixedSize(26, 26)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge_style = (
            "border-radius: 4px; font-size: 11px; font-weight: bold;"
        )
        if r_type == "Material":
            badge.setStyleSheet(badge_style + "background: #ddeeff; color: #185FA5;")
        else:
            badge.setStyleSheet(badge_style + "background: #fff3e0; color: #854F0B;")
        layout.addWidget(badge)

        # Text body
        body = QWidget()
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(2)

        name_lbl = QLabel(rule.get("item_name", "Unnamed rule"))
        name_lbl.setStyleSheet("font-size: 12px; font-weight: bold;")
        name_lbl.setWordWrap(False)

        cond_text = rule.get("condition", "") or "(no condition)"
        cond_lbl = QLabel(cond_text)
        cond_lbl.setStyleSheet("font-size: 11px; color: #555; font-family: monospace;")
        cond_lbl.setWordWrap(False)

        formula_lbl = QLabel(f"qty = {rule.get('formula', '1')}")
        formula_lbl.setStyleSheet("font-size: 10px; color: #999;")

        body_layout.addWidget(name_lbl)
        body_layout.addWidget(cond_lbl)
        body_layout.addWidget(formula_lbl)
        layout.addWidget(body, 1)

        # Click handler — capture rule_index in closure
        def on_click(event, idx=rule_index):
            self.on_card_clicked(idx)

        card.mousePressEvent = on_click
        return card

    def _card_style(self, selected: bool) -> str:
        if selected:
            return (
                "background: white; border: 1.5px solid #185FA5;"
                "border-radius: 6px;"
            )
        return (
            "background: white; border: 0.5px solid #ddd;"
            "border-radius: 6px;"
        )

    def on_card_clicked(self, rule_index: int):
        self.selected_rule_index = rule_index
        self.refresh_card_list()          # re-render to update selection highlight
        self.build_editor(self.rules[rule_index])

    # ── Right: editor ─────────────────────────────────────────────────────────

    def clear_editor(self):
        self._clear_layout(self.editor_body_layout)
        self.editor_body_layout.addStretch()
        self.editor_header.setText("Select a rule to edit")
        self.save_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
        self.condition_widgets = []
        self.selected_result_item = None

    def build_editor(self, rule: dict):
        """Populates the right editor panel for the given rule."""
        self._clear_layout(self.editor_body_layout)
        self.condition_widgets = []
        self.selected_result_item = None

        self.editor_header.setText(f"Editing: {rule.get('item_name', '')}")
        self.save_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)

        # ── Section: Item ─────────────────────────────────────────────────────
        self.editor_body_layout.addWidget(self._section_label("Item"))

        # Type (Material / Labor)
        type_row = QHBoxLayout()
        type_lbl = QLabel("Type:")
        type_lbl.setFixedWidth(90)
        type_lbl.setStyleSheet("font-size: 12px;")
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Material", "Labor"])
        self.type_combo.setCurrentText(rule.get("type", "Material"))
        type_row.addWidget(type_lbl)
        type_row.addWidget(self.type_combo, 1)
        self.editor_body_layout.addLayout(type_row)

        # Item name (read-only) + change button
        item_row = QHBoxLayout()
        item_label_lbl = QLabel("Item:")
        item_label_lbl.setFixedWidth(90)
        item_label_lbl.setStyleSheet("font-size: 12px;")
        self.item_display = QLineEdit(rule.get("item_name", ""))
        self.item_display.setReadOnly(True)
        self.item_display.setStyleSheet("background: #f5f5f5; font-size: 12px;")
        change_btn = QPushButton("🔍 Change…")
        change_btn.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        change_btn.clicked.connect(self.search_database_for_item)
        item_row.addWidget(item_label_lbl)
        item_row.addWidget(self.item_display, 1)
        item_row.addWidget(change_btn)
        self.editor_body_layout.addLayout(item_row)

        # Item code (read-only)
        code_row = QHBoxLayout()
        code_lbl = QLabel("Code:")
        code_lbl.setFixedWidth(90)
        code_lbl.setStyleSheet("font-size: 12px;")
        self.code_display = QLineEdit(rule.get("item_code", ""))
        self.code_display.setReadOnly(True)
        self.code_display.setStyleSheet("background: #f5f5f5; font-size: 11px; color: #666;")
        code_row.addWidget(code_lbl)
        code_row.addWidget(self.code_display, 1)
        self.editor_body_layout.addLayout(code_row)

        # ── Section: Conditions ───────────────────────────────────────────────
        self.editor_body_layout.addWidget(self._section_label("Conditions (all rows must match)"))

        self.cond_container = QWidget()
        self.cond_rows_layout = QVBoxLayout(self.cond_container)
        self.cond_rows_layout.setContentsMargins(0, 0, 0, 0)
        self.cond_rows_layout.setSpacing(4)
        self.editor_body_layout.addWidget(self.cond_container)

        add_cond_btn = QPushButton("+ add condition row")
        add_cond_btn.setStyleSheet("color: #185FA5; border: none; background: none; font-size: 11px; text-align: left;")
        add_cond_btn.clicked.connect(self.add_condition_row)
        self.editor_body_layout.addWidget(add_cond_btn)

        # Condition preview (read-only generated string)
        self.editor_body_layout.addWidget(self._section_label("Condition preview"))
        self.condition_preview = QLabel("")
        self.condition_preview.setWordWrap(True)
        self.condition_preview.setStyleSheet(
            "background: #f0f0f0; border-radius: 4px; padding: 6px 8px;"
            "font-family: monospace; font-size: 11px; color: #333;"
        )
        self.editor_body_layout.addWidget(self.condition_preview)

        # ── Section: Formula ──────────────────────────────────────────────────
        self.editor_body_layout.addWidget(self._section_label("Quantity formula"))

        avail_vars = FORMULA_VARS.get(rule.get("object", ""), [])
        hint = QLabel(f"Available variables: {', '.join(avail_vars) if avail_vars else 'none'}")
        hint.setStyleSheet("font-size: 10px; color: #999;")
        self.editor_body_layout.addWidget(hint)

        self.formula_input = QLineEdit(rule.get("formula", "1"))
        self.formula_input.setStyleSheet("font-family: monospace; font-size: 12px;")
        self.editor_body_layout.addWidget(self.formula_input)

        self.editor_body_layout.addStretch()

        # Parse existing conditions and populate rows
        self._parse_and_display_conditions(rule)
        self._update_condition_preview()

    # ── Condition rows ────────────────────────────────────────────────────────

    def _parse_and_display_conditions(self, rule: dict):
        """Splits the saved condition string back into structured rows."""
        condition_str = rule.get("condition", "")
        if not condition_str or condition_str.strip() == "True":
            self.add_condition_row()
            return

        # Split on top-level ' and ' / ' or ' only
        tokens = re.split(r'\s+(and|or)\s+', condition_str, flags=re.IGNORECASE)
        self.add_condition_row(expression=tokens[0])
        for i in range(1, len(tokens), 2):
            logical_op = tokens[i].upper()
            expr = tokens[i + 1] if i + 1 < len(tokens) else ""
            self.add_condition_row(logical_op=logical_op, expression=expr)

    def add_condition_row(self, logical_op: str = None, expression: str = None):
        """Adds one condition row to the editor."""
        obj_name = self.current_object_type
        prop_names = list(self.property_data.get(obj_name, {}).keys())

        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(4)

        # AND / OR connector (hidden for first row)
        logic_combo = QComboBox()
        logic_combo.addItems(["AND", "OR"])
        logic_combo.setFixedWidth(54)
        logic_combo.setVisible(len(self.condition_widgets) > 0)
        if logical_op:
            logic_combo.setCurrentText(logical_op)
        logic_combo.currentTextChanged.connect(self._update_condition_preview)

        # Property selector
        prop_combo = QComboBox()
        prop_combo.addItems(prop_names)

        # Operator selector
        op_combo = QComboBox()
        op_combo.addItems(["==", "!=", ">", "<", ">=", "<="])
        op_combo.setFixedWidth(52)

        # Value widget (will be replaced dynamically)
        value_widget = QLineEdit()

        # Remove button
        rem_btn = QPushButton("✕")
        rem_btn.setFixedWidth(24)
        rem_btn.setStyleSheet("color: #999; border: none; background: none; font-size: 11px;")

        row_layout.addWidget(logic_combo)
        row_layout.addWidget(prop_combo, 2)
        row_layout.addWidget(op_combo)
        row_layout.addWidget(value_widget, 2)
        row_layout.addWidget(rem_btn)

        widget_map = {
            "widget": row_widget,
            "logic": logic_combo,
            "prop": prop_combo,
            "op": op_combo,
            "value": value_widget,
        }
        self.condition_widgets.append(widget_map)
        self.cond_rows_layout.addWidget(row_widget)

        # Wire up signals
        prop_combo.currentTextChanged.connect(
            lambda text, wm=widget_map: self._on_property_change(text, wm)
        )
        op_combo.currentTextChanged.connect(self._update_condition_preview)
        rem_btn.clicked.connect(lambda _, w=row_widget: self._remove_condition_row(w))

        # Populate value widget for the initial property
        self._on_property_change(prop_combo.currentText(), widget_map)

        # Now restore saved values from the expression string
        if expression:
            self._restore_expression(expression.strip(), widget_map, op_combo)

        self._update_condition_preview()

    def _restore_expression(self, expression: str, widget_map: dict, op_combo: QComboBox):
        """Parses a single condition token and restores it into the row widgets."""
        # Handle 'not <prop>' style (boolean False)
        not_match = re.match(r"^not\s+(\w+)$", expression)
        if not_match:
            prop = not_match.group(1)
            widget_map["prop"].setCurrentText(prop)
            self._on_property_change(prop, widget_map)
            op_combo.setCurrentText("==")
            val_w = widget_map["value"]
            if isinstance(val_w, QComboBox):
                val_w.setCurrentText("False")
            else:
                val_w.setText("False")
            return

        # Handle plain '<prop>' style (boolean True)
        plain_match = re.match(r"^(\w+)$", expression)
        if plain_match:
            prop = plain_match.group(1)
            widget_map["prop"].setCurrentText(prop)
            self._on_property_change(prop, widget_map)
            op_combo.setCurrentText("==")
            val_w = widget_map["value"]
            if isinstance(val_w, QComboBox):
                val_w.setCurrentText("True")
            else:
                val_w.setText("True")
            return

        # Standard: <prop> <op> <value>
        match = re.match(r"(\w+)\s*([<>=!]+)\s*(.*)", expression)
        if not match:
            return
        prop, op, val = match.groups()
        prop = prop.strip()
        op = op.strip()
        val = val.strip().strip("'\"")

        widget_map["prop"].setCurrentText(prop)
        self._on_property_change(prop, widget_map)
        op_combo.setCurrentText(op)
        val_w = widget_map["value"]
        if isinstance(val_w, QComboBox):
            val_w.setCurrentText(val)
        elif isinstance(val_w, QSpinBox):
            try:
                val_w.setValue(int(float(val)))
            except ValueError:
                pass
        else:
            val_w.setText(val)

    def _on_property_change(self, prop_name: str, widget_map: dict):
        """Swaps the value widget to QComboBox/QSpinBox/QLineEdit based on prop type."""
        obj_name = self.current_object_type
        prop_info = self.property_data.get(obj_name, {}).get(prop_name)

        current_widget = widget_map["value"]
        if isinstance(prop_info, list):
            target_cls = QComboBox
        elif prop_info == "int":
            target_cls = QSpinBox
        else:
            target_cls = QLineEdit

        if not isinstance(current_widget, target_cls):
            new_widget = target_cls()
            if isinstance(new_widget, QSpinBox):
                new_widget.setRange(-100000, 100000)
                new_widget.valueChanged.connect(self._update_condition_preview)
            elif isinstance(new_widget, QComboBox):
                new_widget.currentTextChanged.connect(self._update_condition_preview)
            else:
                new_widget.textChanged.connect(self._update_condition_preview)

            layout = widget_map["widget"].layout()
            layout.replaceWidget(current_widget, new_widget)
            current_widget.deleteLater()
            widget_map["value"] = new_widget
            current_widget = new_widget

        if isinstance(current_widget, QComboBox) and isinstance(prop_info, list):
            current_widget.blockSignals(True)
            current_widget.clear()
            current_widget.addItems([str(p) for p in prop_info])
            current_widget.blockSignals(False)

        self._update_condition_preview()

    def _remove_condition_row(self, row_widget: QWidget):
        if len(self.condition_widgets) <= 1:
            return
        self.condition_widgets = [wm for wm in self.condition_widgets if wm["widget"] is not row_widget]
        row_widget.deleteLater()
        # First row never shows a logic connector
        if self.condition_widgets:
            self.condition_widgets[0]["logic"].setVisible(False)
        self._update_condition_preview()

    def _update_condition_preview(self):
        """Regenerates the condition string and shows it in the preview label."""
        parts = self._build_condition_parts()
        preview = " ".join(parts) if parts else "(no conditions)"
        if hasattr(self, "condition_preview"):
            self.condition_preview.setText(preview)

    def _build_condition_parts(self) -> list:
        """Returns the list of tokens that form the full condition string."""
        parts = []
        for i, wm in enumerate(self.condition_widgets):
            prop = wm["prop"].currentText()
            op = wm["op"].currentText()
            val_w = wm["value"]

            if isinstance(val_w, QSpinBox):
                val = str(val_w.value())
            elif isinstance(val_w, QComboBox):
                val = val_w.currentText()
            else:
                val = val_w.text().strip()

            if not prop:
                continue

            # Add AND/OR connector for rows after the first
            if i > 0:
                parts.append(wm["logic"].currentText().lower())

            # Format the value: quote strings, leave numbers and booleans bare
            if val.lower() in ("true", "false") or re.match(r"^-?\d+(\.\d+)?$", val):
                parts.append(f"{prop} {op} {val}")
            else:
                parts.append(f"{prop} {op} '{val}'")

        return parts

    # ── Editor actions ────────────────────────────────────────────────────────

    def search_database_for_item(self):
        """Opens the SearchDialog so the user can pick a material or labor item."""
        db_type, ok = QInputDialog.getItem(
            self, "Select type", "Which database?", ["Material", "Labor"], 0, False
        )
        if not (ok and db_type):
            return
        dialog = SearchDialog(db_type, self)
        if dialog.exec():
            item = dialog.get_selected()
            if item:
                self.selected_result_item = item
                self.item_display.setText(item["name"])
                self.code_display.setText(item.get("code", ""))
                self.type_combo.setCurrentText(item["type"])

    def save_rule_changes(self):
        if self.selected_rule_index == -1:
            return
        rule = self.rules[self.selected_rule_index]

        # Build condition string from rows
        parts = self._build_condition_parts()
        rule["condition"] = " ".join(parts)

        # Update item fields if user picked a new one
        if self.selected_result_item:
            rule["type"] = self.selected_result_item["type"]
            rule["item_code"] = self.selected_result_item.get("code", "")
            rule["item_name"] = (
                self.selected_result_item.get("name")
                or self.selected_result_item.get("item_name", "")
            )
        else:
            rule["type"] = self.type_combo.currentText()

        rule["formula"] = self.formula_input.text().strip() or "1"

        self.save_rules()
        self.refresh_card_list()
        self.editor_header.setText(f"Editing: {rule.get('item_name', '')}")
        QMessageBox.information(self, "Saved", "Rule saved successfully.")

    def create_new_rule(self):
        """Appends a blank rule for the current object type and selects it."""
        new_rule = {
            "object": self.current_object_type,
            "item_name": "New Rule — edit me",
            "condition": "",
            "type": "Material",
            "item_code": "N/A",
            "formula": "1",
        }
        self.rules.append(new_rule)
        self.save_rules()
        self.selected_rule_index = len(self.rules) - 1
        self.refresh_card_list()
        self.build_editor(new_rule)

    def delete_selected_rule(self):
        if self.selected_rule_index == -1:
            return
        rule = self.rules[self.selected_rule_index]
        reply = QMessageBox.question(
            self, "Delete rule",
            f"Delete rule:\n'{rule.get('item_name')}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            del self.rules[self.selected_rule_index]
            self.selected_rule_index = -1
            self.save_rules()
            self.refresh_card_list()
            self.clear_editor()

    # ── Persistence ───────────────────────────────────────────────────────────

    def load_rules(self):
        try:
            with open("rules.json", "r") as f:
                self.rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.rules = []

    def save_rules(self):
        try:
            with open("rules.json", "w") as f:
                json.dump(self.rules, f, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save rules.json:\n{e}")
