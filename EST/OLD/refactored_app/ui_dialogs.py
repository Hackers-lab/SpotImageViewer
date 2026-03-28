"""
This module contains all QDialog-based windows for the application,
such as the settings, search, database manager, and rule manager dialogs.
"""

import sqlite3
import json
import re
import openpyxl

from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLineEdit, QListWidget, QPushButton,
                             QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem, QHBoxLayout, 
                             QFileDialog, QMessageBox, QGroupBox, QFormLayout, QComboBox, 
                             QSpinBox, QHeaderView, QInputDialog, QWidget, QSplitter, 
                             QTreeWidget, QTreeWidgetItem, QLabel, QScrollArea)
from PyQt6.QtCore import Qt

# Import shared constants
from constants import PROPERTY_DATA, FORMULA_VARS

class SearchDialog(QDialog):
    """
    A dialog for searching the materials or labor database and adding
    an item to the estimate.
    """
    def __init__(self, db_type, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Search & Add {db_type}")
        self.setFixedSize(600, 400)
        self.layout = QVBoxLayout(self)
        
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Type to search official items...")
        self.layout.addWidget(self.search_box)
        
        self.list_widget = QListWidget()
        self.layout.addWidget(self.list_widget)
        self.search_box.textChanged.connect(self.filter_list)
        
        self.add_btn = QPushButton("Add Selected to Estimate")
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        self.add_btn.clicked.connect(self.accept)
        self.layout.addWidget(self.add_btn)

        self.items_data = {}
        self.load_data(db_type)

    def load_data(self, db_type):
        """Loads data from the specified database table."""
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        if db_type == "Material":
            cursor.execute("SELECT item_code, item_name, unit, rate FROM materials")
        else:
            cursor.execute("SELECT labor_code, task_name, unit, rate FROM labor")
        
        for row in cursor.fetchall():
            display_text = f"{row[1]} ({row[2]}) - Rs. {row[3]}"
            self.items_data[display_text] = {"code": row[0], "name": row[1], "unit": row[2], "rate": row[3], "type": db_type}
            self.list_widget.addItem(display_text)
        conn.close()

    def filter_list(self, text):
        """Filters the list widget based on the search box text."""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def get_selected(self):
        """Returns the data for the currently selected item."""
        selected = self.list_widget.currentItem()
        if selected:
            return self.items_data[selected.text()]
        return None


class SettingsDialog(QDialog):
    """
    A dialog for accessing advanced application settings like the rule
    engine toggle and the database/ruleset managers.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("Advanced Settings")
        self.setFixedSize(300, 150)
        
        layout = QVBoxLayout(self)
        
        db_sync_btn = QPushButton("🗃️ Master DB (Excel Sync)")
        db_sync_btn.clicked.connect(self.parent_app.open_db_manager)
        layout.addWidget(db_sync_btn)

        rules_btn = QPushButton("🧠 Ruleset Manager")
        rules_btn.clicked.connect(self.parent_app.open_rule_manager)
        layout.addWidget(rules_btn)

        layout.addStretch()


class DatabaseManagerDialog(QDialog):
    """
    A dialog for managing the master database, allowing import from and
    export to Excel files.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Master DB Sync (Excel)")
        self.setGeometry(100, 100, 800, 600)
        
        layout = QVBoxLayout(self)
        
        button_layout = QHBoxLayout()
        import_btn = QPushButton("📥 Import from Excel")
        import_btn.clicked.connect(self.import_from_excel)
        export_btn = QPushButton("📤 Export to Excel")
        export_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(import_btn)
        button_layout.addWidget(export_btn)
        layout.addLayout(button_layout)
        
        self.tabs = QTabWidget()
        self.materials_table = QTableWidget()
        self.labor_table = QTableWidget()
        self.tabs.addTab(self.materials_table, "Materials")
        self.tabs.addTab(self.labor_table, "Labor")
        layout.addWidget(self.tabs)
        
        self.load_table_data()

    def populate_table(self, table_widget, data, headers):
        """Fills a table widget with data."""
        table_widget.setRowCount(0)
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)
        for row_num, row_data in enumerate(data):
            table_widget.insertRow(row_num)
            for col_num, col_data in enumerate(row_data):
                table_widget.setItem(row_num, col_num, QTableWidgetItem(str(col_data)))
        table_widget.resizeColumnsToContents()

    def load_table_data(self):
        """Loads data from the database and populates the tables."""
        self.materials_table.clear()
        self.labor_table.clear()
        
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM materials")
        self.populate_table(self.materials_table, cursor.fetchall(), ["Item Code", "Item Name", "Rate", "Unit"])
        
        cursor.execute("SELECT * FROM labor")
        self.populate_table(self.labor_table, cursor.fetchall(), ["Labor Code", "Task Name", "Rate", "Unit"])
        
        conn.close()

    def export_to_excel(self):
        """Exports the database tables to an Excel file."""
        filename, _ = QFileDialog.getSaveFileName(self, "Export DB to Excel", "master_database.xlsx", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                ws = wb.create_sheet(title=table_name)
                cursor.execute(f"PRAGMA table_info({table_name})")
                headers = [info[1] for info in cursor.fetchall()]
                ws.append(headers)

                cursor.execute(f"SELECT * FROM {table_name}")
                for row in cursor.fetchall():
                    ws.append(row)
            
            conn.close()
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Database exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export database: {e}")

    def import_from_excel(self):
        """Imports data from an Excel file into the database."""
        filename, _ = QFileDialog.getOpenFileName(self, "Import DB from Excel", "", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.load_workbook(filename)
            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                if table_name in wb.sheetnames:
                    ws = wb[table_name]
                    cursor.execute(f"DELETE FROM {table_name}")
                    headers = [cell.value for cell in ws[1]]
                    placeholders = ', '.join(['?'] * len(headers))
                    query = f"INSERT OR REPLACE INTO {table_name} ({', '.join(headers)}) VALUES ({placeholders})"
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            cursor.execute(query, row)
            conn.commit()
            conn.close()
            self.load_table_data()
            QMessageBox.information(self, "Success", "Database imported successfully. Please restart or refresh to see updates in 'Add Custom' lists.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import database: {e}")



class RulesetManagerDialog(QDialog):
    """
    Rule Manager — v3 (Tree + Checkbox Filter + Simulator)
    -------------------------------------------------------
    Left   : Two/three-level QTreeWidget
                 SmartPole → LT / HT / DTR → DTR sub-sizes
                 SmartSpan → conductor types
                 SmartHome
             + search box above the tree
    Centre : Flat rule-card list (filtered by tree selection + checkboxes)
             Top bar : search box + AND/OR toggle + "New rule" button
             Filter chips : context-aware checkboxes for key properties
             Cards : M/L badge, item name, condition, formula
             Bottom : collapsible Simulator strip
                         Set property values → Run → cards highlight,
                         BOM preview table appears
    Right  : Editor panel
                 Item (type, name, code, Change button)
                 Condition rows (prop / op / value dropdowns)
                 Live condition preview
                 Formula + available-vars hint
             Fixed footer : Delete / Save
    """

    # ── Tree structure definition ─────────────────────────────────────────────
    # Each entry: (display_label, obj_type, filter_dict, children)
    # filter_dict keys match rule context keys; children are sub-nodes
    TREE_DEF = [
        ("SmartPole", "SmartPole", {}, [
            ("LT Pole",  "SmartPole", {"pole_type": "LT"}, [
                ("Height 8m", "SmartPole", {"pole_type": "LT", "height": 8},  []),
                ("Height 9m", "SmartPole", {"pole_type": "LT", "height": 9},  []),
            ]),
            ("HT Pole",  "SmartPole", {"pole_type": "HT"}, [
                ("Height 8m", "SmartPole", {"pole_type": "HT", "height": 8},  []),
                ("Height 9m", "SmartPole", {"pole_type": "HT", "height": 9},  []),
            ]),
            ("DTR / DP", "SmartPole", {"pole_type": "DTR"}, [
                ("16 KVA",  "SmartPole", {"pole_type": "DTR", "dtr_size": "16 KVA"},  []),
                ("25 KVA",  "SmartPole", {"pole_type": "DTR", "dtr_size": "25KVA"},   []),
                ("63 KVA",  "SmartPole", {"pole_type": "DTR", "dtr_size": "63KVA"},   []),
                ("100 KVA", "SmartPole", {"pole_type": "DTR", "dtr_size": "100KVA"},  []),
                ("160 KVA", "SmartPole", {"pole_type": "DTR", "dtr_size": "160KVA"},  []),
            ]),
            ("Existing Pole", "SmartPole", {"is_existing": True}, []),
        ]),
        ("SmartSpan", "SmartSpan", {}, [
            ("AB Cable",     "SmartSpan", {"conductor": "AB Cable"},     [
                ("New",            "SmartSpan", {"conductor": "AB Cable", "aug_type": "New"},           []),
                ("Replace 2W→4W",  "SmartSpan", {"conductor": "AB Cable", "aug_type": "Replace 2W->4W"},[]),
                ("Add-on 2W",      "SmartSpan", {"conductor": "AB Cable", "aug_type": "Add-on 2W"},     []),
            ]),
            ("ACSR",         "SmartSpan", {"conductor": "ACSR"},         [
                ("3 Wire", "SmartSpan", {"conductor": "ACSR", "wire_count": "3"}, []),
                ("4 Wire", "SmartSpan", {"conductor": "ACSR", "wire_count": "4"}, []),
            ]),
            ("PVC Cable",    "SmartSpan", {"conductor": "PVC Cable"},    [
                ("10 SQMM", "SmartSpan", {"conductor": "PVC Cable", "cable_size": "10 SQMM"}, []),
                ("16 SQMM", "SmartSpan", {"conductor": "PVC Cable", "cable_size": "16 SQMM"}, []),
                ("25 SQMM", "SmartSpan", {"conductor": "PVC Cable", "cable_size": "25 SQMM"}, []),
            ]),
            ("Service Drop", "SmartSpan", {"conductor": "Service Drop"}, [
                ("1 Phase", "SmartSpan", {"conductor": "Service Drop", "phase": "1 Phase"}, []),
                ("3 Phase", "SmartSpan", {"conductor": "Service Drop", "phase": "3 Phase"}, []),
            ]),
        ]),
        ("SmartHome", "SmartHome", {}, []),
    ]

    # Context-aware filter chips per object type
    FILTER_CHIPS = {
        "SmartPole": [
            ("Not Existing",  "is_existing",    False),
            ("Is Existing",   "is_existing",    True),
            ("Has Extension", "has_extension",  True),
            ("Has CG",        "has_cg",         True),
            ("With Earth",    "earth_count_gt", 0),
            ("With Stay",     "stay_count_gt",  0),
        ],
        "SmartSpan": [
            ("Not Existing Span", "is_existing_span", False),
            ("Service Drop",      "is_service_drop",  True),
            ("Has CG",            "has_cg",           True),
            ("New Work",          "aug_type",          "New"),
        ],
        "SmartHome": [],
    }

    # Simulator default property values per object type
    SIM_DEFAULTS = {
        "SmartPole": {
            "pole_type":     ("combo", ["LT", "HT", "DTR"],          "LT"),
            "is_existing":   ("combo", ["False", "True"],             "False"),
            "height":        ("combo", ["8", "9"],                    "8"),
            "dtr_size":      ("combo", ["None","16 KVA","25KVA","63KVA","100KVA","160KVA"], "None"),
            "earth_count":   ("spin",  (0, 10),                       1),
            "stay_count":    ("spin",  (0, 10),                       0),
            "has_extension": ("combo", ["False", "True"],             "False"),
            "has_cg":        ("combo", ["False", "True"],             "False"),
            "ht_spans_count":("spin",  (0, 10),                       0),
            "use_uh":        ("combo", ["False", "True"],             "False"),
        },
        "SmartSpan": {
            "conductor":        ("combo", ["AB Cable","ACSR","PVC Cable","Service Drop"], "AB Cable"),
            "is_existing_span": ("combo", ["False", "True"],  "False"),
            "is_service_drop":  ("combo", ["False", "True"],  "False"),
            "length":           ("spin",  (1, 1000),           40),
            "wire_count":       ("combo", ["2","3","4"],       "3"),
            "wire_size":        ("combo", ["30SQMM","50SQMM"],"50SQMM"),
            "cable_size":       ("combo", ["10 SQMM","16 SQMM","25 SQMM"], "10 SQMM"),
            "phase":            ("combo", ["1 Phase","3 Phase"], "3 Phase"),
            "has_cg":           ("combo", ["False","True"],    "False"),
            "aug_type":         ("combo", ["New","Replace 2W->4W","Add-on 2W"], "New"),
            "is_lt_span":       ("combo", ["True","False"],    "True"),
            "consider_cable":   ("combo", ["False","True"],    "False"),
            "use_uh":           ("combo", ["False","True"],    "False"),
        },
        "SmartHome": {},
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Rule Manager")
        self.setGeometry(80, 80, 1420, 860)

        self.rules = []
        self.selected_rule_index = -1
        self.selected_result_item = None
        self.condition_widgets = []
        self.active_tree_filter = {}          # filter dict from selected tree node
        self.active_obj_type = "SmartPole"
        self.filter_logic = "AND"             # "AND" or "OR"
        self.active_chips = set()             # set of chip keys currently checked
        self.sim_visible = False
        self.sim_widgets = {}                 # prop_name -> widget
        self.property_data = PROPERTY_DATA

        self._build_ui()
        self.load_rules()
        self._select_tree_root("SmartPole")

    # ═════════════════════════════════════════════════════════════════════════
    # UI CONSTRUCTION
    # ═════════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        root = QHBoxLayout(self)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)
        root.addWidget(self._build_left_panel())
        root.addWidget(self._build_centre_panel())
        root.addWidget(self._build_right_panel())

    # ── LEFT: tree ────────────────────────────────────────────────────────────

    def _build_left_panel(self):
        panel = QWidget()
        panel.setFixedWidth(220)
        panel.setStyleSheet("background:#f5f5f5; border-right:1px solid #ddd;")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Search box
        self.tree_search = QLineEdit()
        self.tree_search.setPlaceholderText("Search tree…")
        self.tree_search.setStyleSheet(
            "margin:6px; padding:4px 8px; border:0.5px solid #ccc;"
            "border-radius:4px; font-size:12px;"
        )
        self.tree_search.textChanged.connect(self._filter_tree)
        lay.addWidget(self.tree_search)

        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.setStyleSheet("""
            QTreeWidget { border:none; background:#f5f5f5; font-size:12px; }
            QTreeWidget::item { padding:4px 6px; }
            QTreeWidget::item:selected { background:#ddeeff; color:#0C447C; }
            QTreeWidget::item:hover:!selected { background:#ebebeb; }
        """)
        self.tree.itemClicked.connect(self._on_tree_click)
        lay.addWidget(self.tree)

        self._populate_tree()
        return panel

    def _populate_tree(self):
        self.tree.clear()
        self._tree_items = []   # list of (QTreeWidgetItem, obj_type, filter_dict)

        def add_node(parent, label, obj_type, fdict, children):
            item = QTreeWidgetItem(parent if parent else self.tree, [label])
            item.setData(0, Qt.ItemDataRole.UserRole, (obj_type, fdict))
            self._tree_items.append((item, obj_type, fdict))
            for ch in children:
                add_node(item, ch[0], ch[1], ch[2], ch[3])
            return item

        for entry in self.TREE_DEF:
            add_node(None, entry[0], entry[1], entry[2], entry[3])

        self._update_tree_counts()
        self.tree.expandToDepth(1)

    def _update_tree_counts(self):
        """Append rule counts as suffix to each tree item label."""
        for item, obj_type, fdict in self._tree_items:
            base = item.text(0).split("  ")[0]   # strip old count
            count = len(self._get_matching_rules(obj_type, fdict, set()))
            item.setText(0, f"{base}   ({count})" if count else base)

    def _filter_tree(self, text):
        """Show/hide tree items by search text."""
        text = text.lower()
        for item, obj_type, fdict in self._tree_items:
            label = item.text(0).lower()
            hide = text != "" and text not in label
            item.setHidden(hide)

    def _select_tree_root(self, obj_type: str):
        """Programmatically select the root node for an object type."""
        for item, ot, fd in self._tree_items:
            if ot == obj_type and not fd:
                self.tree.setCurrentItem(item)
                self._on_tree_click(item, 0)
                return

    def _on_tree_click(self, item, _col):
        obj_type, fdict = item.data(0, Qt.ItemDataRole.UserRole)
        self.active_obj_type = obj_type
        self.active_tree_filter = fdict
        self.active_chips.clear()
        self.selected_rule_index = -1
        self._rebuild_filter_chips()
        self._refresh_card_list()
        self._clear_editor()
        self._update_centre_title()

    # ── CENTRE: cards ─────────────────────────────────────────────────────────

    def _build_centre_panel(self):
        self._centre = QWidget()
        lay = QVBoxLayout(self._centre)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Top bar
        topbar = QWidget()
        topbar.setStyleSheet("background:white; border-bottom:1px solid #ddd;")
        tl = QHBoxLayout(topbar)
        tl.setContentsMargins(10, 7, 10, 7)
        tl.setSpacing(8)
        self.centre_title = QLabel("")
        self.centre_title.setStyleSheet("font-weight:bold; font-size:13px;")
        self.card_search = QLineEdit()
        self.card_search.setPlaceholderText("Search rules…")
        self.card_search.setStyleSheet(
            "padding:4px 8px; border:0.5px solid #ccc; border-radius:4px; font-size:12px; max-width:200px;"
        )
        self.card_search.textChanged.connect(self._refresh_card_list)

        self.logic_toggle = QPushButton("AND")
        self.logic_toggle.setCheckable(True)
        self.logic_toggle.setChecked(True)
        self.logic_toggle.setFixedWidth(46)
        self.logic_toggle.setStyleSheet(
            "QPushButton{padding:4px; border:1px solid #ccc; border-radius:4px; font-size:11px; font-weight:bold;}"
            "QPushButton:checked{background:#185FA5; color:white; border-color:#185FA5;}"
        )
        self.logic_toggle.clicked.connect(self._toggle_logic)

        new_btn = QPushButton("+ New rule")
        new_btn.setStyleSheet(
            "background:#185FA5; color:white; border:none; padding:5px 12px;"
            "border-radius:4px; font-size:12px;"
        )
        new_btn.clicked.connect(self.create_new_rule)
        tl.addWidget(self.centre_title)
        tl.addStretch()
        tl.addWidget(self.card_search)
        tl.addWidget(self.logic_toggle)
        tl.addWidget(new_btn)
        lay.addWidget(topbar)

        # Filter chip bar
        self.chip_bar = QWidget()
        self.chip_bar.setStyleSheet("background:#fafafa; border-bottom:1px solid #eee;")
        self.chip_bar_layout = QHBoxLayout(self.chip_bar)
        self.chip_bar_layout.setContentsMargins(10, 5, 10, 5)
        self.chip_bar_layout.setSpacing(6)
        lay.addWidget(self.chip_bar)

        # Card scroll
        self.card_container = QWidget()
        self.card_container.setStyleSheet("background:#f8f8f8;")
        self.card_layout = QVBoxLayout(self.card_container)
        self.card_layout.setContentsMargins(8, 8, 8, 8)
        self.card_layout.setSpacing(5)
        self.card_layout.addStretch()

        scroll = QScrollArea()
        scroll.setWidget(self.card_container)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        lay.addWidget(scroll, 1)

        # Simulator panel (collapsed by default)
        lay.addWidget(self._build_sim_panel())
        return self._centre

    def _toggle_logic(self):
        self.filter_logic = "AND" if self.logic_toggle.isChecked() else "OR"
        self.logic_toggle.setText(self.filter_logic)
        self._refresh_card_list()

    def _rebuild_filter_chips(self):
        """Clear and rebuild chip bar for the current object type."""
        while self.chip_bar_layout.count():
            item = self.chip_bar_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        chips = self.FILTER_CHIPS.get(self.active_obj_type, [])
        if not chips:
            self.chip_bar.setVisible(False)
            return
        self.chip_bar.setVisible(True)

        lbl = QLabel("Filter:")
        lbl.setStyleSheet("font-size:11px; color:#888;")
        self.chip_bar_layout.addWidget(lbl)

        self._chip_checkboxes = {}
        for label, key, val in chips:
            cb = QCheckBox(label)
            cb.setStyleSheet(
                "QCheckBox{font-size:11px; padding:2px 6px; border:0.5px solid #ccc;"
                "border-radius:10px; background:white;}"
                "QCheckBox:checked{background:#ddeeff; border-color:#378ADD; color:#0C447C;}"
            )
            chip_key = (key, str(val))
            cb.stateChanged.connect(
                lambda state, k=chip_key: self._on_chip_toggle(k, state)
            )
            self.chip_bar_layout.addWidget(cb)
            self._chip_checkboxes[chip_key] = cb

        self.chip_bar_layout.addStretch()

    def _on_chip_toggle(self, chip_key, state):
        if state:
            self.active_chips.add(chip_key)
        else:
            self.active_chips.discard(chip_key)
        self._refresh_card_list()

    def _update_centre_title(self):
        visible = self._visible_rule_indices()
        label = self.active_tree_filter or {"object": self.active_obj_type}
        self.centre_title.setText(
            f"{self.active_obj_type.replace('Smart','')}  —  {len(visible)} rule(s)"
        )

    # ── Card list ─────────────────────────────────────────────────────────────

    def _get_matching_rules(self, obj_type, fdict, chips):
        """
        Returns list of (original_index, rule) whose conditions are consistent
        with fdict AND (if chips non-empty) the chip filters.
        Chip filter uses self.filter_logic (AND/OR).
        """
        result = []
        for i, rule in enumerate(self.rules):
            if rule.get("object") != obj_type:
                continue
            cond = rule.get("condition", "")
            # Tree filter: all keys in fdict must appear somewhere in condition
            # (or fdict is empty = show all for this object type)
            if fdict:
                ok = True
                for prop, val in fdict.items():
                    # Check if condition string references this value
                    val_str = str(val)
                    # Match  prop == 'val'  or  prop == val  or  prop != ...
                    patterns = [
                        f"{prop} == '{val_str}'",
                        f"{prop} == \"{val_str}\"",
                        f"{prop} == {val_str}",
                        f"== {val_str}",
                    ]
                    if not any(p in cond for p in patterns):
                        ok = False
                        break
                if not ok:
                    continue
            # Chip filters
            if chips:
                chip_results = []
                for (key, val_str) in chips:
                    val = val_str
                    if key.endswith("_gt"):
                        # e.g. earth_count_gt → earth_count > 0
                        real_key = key[:-3]
                        match = (f"{real_key} >" in cond)
                    elif val_str.lower() == "false":
                        match = (f"not {key}" in cond or f"{key} == False" in cond)
                    elif val_str.lower() == "true":
                        match = (
                            f"{key}" in cond
                            and f"not {key}" not in cond
                        ) or f"{key} == True" in cond
                    else:
                        match = (
                            f"{key} == '{val}'" in cond
                            or f"{key} == \"{val}\"" in cond
                            or f"{key} == {val}" in cond
                        )
                    chip_results.append(match)

                if self.filter_logic == "AND" and not all(chip_results):
                    continue
                if self.filter_logic == "OR" and not any(chip_results):
                    continue

            result.append((i, rule))
        return result

    def _visible_rule_indices(self):
        search = self.card_search.text().lower() if hasattr(self, "card_search") else ""
        matched = self._get_matching_rules(
            self.active_obj_type, self.active_tree_filter, self.active_chips
        )
        if search:
            matched = [
                (i, r) for i, r in matched
                if search in r.get("item_name", "").lower()
                or search in r.get("condition", "").lower()
            ]
        return matched

    def _refresh_card_list(self):
        # Remove all cards (keep trailing stretch)
        while self.card_layout.count() > 1:
            item = self.card_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        matched = self._visible_rule_indices()
        sim_hits = self._sim_matching_indices() if self.sim_visible else set()

        for orig_idx, rule in matched:
            card = self._make_card(orig_idx, rule, orig_idx in sim_hits)
            self.card_layout.insertWidget(self.card_layout.count() - 1, card)

        self._update_centre_title()
        self._update_tree_counts()

    def _make_card(self, rule_index, rule, sim_highlight=False):
        card = QWidget()
        card.setCursor(Qt.CursorShape.PointingHandCursor)
        selected = (rule_index == self.selected_rule_index)

        border_color = "#378ADD" if selected else ("#5DCAA5" if sim_highlight else "#ddd")
        border_w = "1.5px" if (selected or sim_highlight) else "0.5px"
        bg = "#eaf8f4" if sim_highlight else "white"
        card.setStyleSheet(
            f"background:{bg}; border:{border_w} solid {border_color};"
            "border-radius:6px;"
        )

        lay = QHBoxLayout(card)
        lay.setContentsMargins(9, 7, 9, 7)
        lay.setSpacing(8)

        r_type = rule.get("type", "Material")
        badge = QLabel("M" if r_type == "Material" else "L")
        badge.setFixedSize(24, 24)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge.setStyleSheet(
            "border-radius:4px; font-size:10px; font-weight:bold; "
            + ("background:#ddeeff; color:#185FA5;" if r_type == "Material"
               else "background:#fff3e0; color:#854F0B;")
        )
        lay.addWidget(badge)

        body = QWidget()
        bl = QVBoxLayout(body)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(1)

        name = QLabel(rule.get("item_name", "Unnamed"))
        name.setStyleSheet("font-size:12px; font-weight:bold;")
        cond = QLabel(rule.get("condition", "") or "(no condition)")
        cond.setStyleSheet("font-size:11px; color:#555; font-family:monospace;")
        formula = QLabel(f"qty = {rule.get('formula','1')}")
        formula.setStyleSheet("font-size:10px; color:#999;")
        bl.addWidget(name)
        bl.addWidget(cond)
        bl.addWidget(formula)
        lay.addWidget(body, 1)

        tag = QLabel(r_type)
        tag.setStyleSheet(
            "font-size:10px; padding:2px 7px; border-radius:10px; "
            + ("background:#ddeeff; color:#185FA5;" if r_type == "Material"
               else "background:#fff3e0; color:#854F0B;")
        )
        lay.addWidget(tag)

        card.mousePressEvent = lambda e, idx=rule_index: self._on_card_click(idx)
        return card

    def _on_card_click(self, rule_index):
        self.selected_rule_index = rule_index
        self._refresh_card_list()
        self._build_editor(self.rules[rule_index])

    # ── SIMULATOR ─────────────────────────────────────────────────────────────

    def _build_sim_panel(self):
        self._sim_outer = QWidget()
        self._sim_outer.setStyleSheet("border-top:1px solid #ddd; background:#f5f5f5;")
        sim_layout = QVBoxLayout(self._sim_outer)
        sim_layout.setContentsMargins(0, 0, 0, 0)
        sim_layout.setSpacing(0)

        # Collapsible header
        self._sim_header_btn = QPushButton("▲  Simulator — set values and see which rules fire")
        self._sim_header_btn.setStyleSheet(
            "text-align:left; padding:6px 12px; border:none; background:#f0f0f0;"
            "font-size:12px; font-weight:bold; color:#333;"
        )
        self._sim_header_btn.clicked.connect(self._toggle_sim)
        sim_layout.addWidget(self._sim_header_btn)

        # Body (hidden by default)
        self._sim_body = QWidget()
        self._sim_body.setVisible(False)
        sb_lay = QVBoxLayout(self._sim_body)
        sb_lay.setContentsMargins(10, 8, 10, 8)
        sb_lay.setSpacing(6)

        # Property input row
        self._sim_inputs_widget = QWidget()
        self._sim_inputs_layout = QHBoxLayout(self._sim_inputs_widget)
        self._sim_inputs_layout.setContentsMargins(0, 0, 0, 0)
        self._sim_inputs_layout.setSpacing(8)
        sb_lay.addWidget(self._sim_inputs_widget)

        # Run button + match count
        run_row = QHBoxLayout()
        run_btn = QPushButton("▶  Run simulation")
        run_btn.setStyleSheet(
            "background:#185FA5; color:white; border:none; padding:5px 16px;"
            "border-radius:4px; font-size:12px;"
        )
        run_btn.clicked.connect(self._run_sim)
        self._sim_count_lbl = QLabel("")
        self._sim_count_lbl.setStyleSheet("font-size:12px; color:#0F6E56; font-weight:bold;")
        run_row.addWidget(run_btn)
        run_row.addWidget(self._sim_count_lbl)
        run_row.addStretch()
        sb_lay.addLayout(run_row)

        # BOM preview table
        self._sim_table = QTableWidget(0, 4)
        self._sim_table.setHorizontalHeaderLabels(["Type", "Item", "Qty", "Formula"])
        self._sim_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._sim_table.setMaximumHeight(160)
        self._sim_table.setStyleSheet("font-size:11px;")
        self._sim_table.setVisible(False)
        sb_lay.addWidget(self._sim_table)

        sim_layout.addWidget(self._sim_body)
        return self._sim_outer

    def _toggle_sim(self):
        self.sim_visible = not self.sim_visible
        self._sim_body.setVisible(self.sim_visible)
        arrow = "▼" if self.sim_visible else "▲"
        self._sim_header_btn.setText(
            f"{arrow}  Simulator — set values and see which rules fire"
        )
        if self.sim_visible:
            self._rebuild_sim_inputs()

    def _rebuild_sim_inputs(self):
        """Rebuild simulator property input widgets for current object type."""
        while self._sim_inputs_layout.count():
            item = self._sim_inputs_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.sim_widgets = {}

        defaults = self.SIM_DEFAULTS.get(self.active_obj_type, {})
        for prop, (wtype, options, default) in defaults.items():
            col = QWidget()
            cl = QVBoxLayout(col)
            cl.setContentsMargins(0, 0, 0, 0)
            cl.setSpacing(2)
            lbl = QLabel(prop)
            lbl.setStyleSheet("font-size:10px; color:#666;")
            cl.addWidget(lbl)
            if wtype == "combo":
                w = QComboBox()
                w.addItems([str(o) for o in options])
                w.setCurrentText(str(default))
                w.setStyleSheet("font-size:11px; padding:3px;")
            else:
                w = QSpinBox()
                w.setRange(options[0], options[1])
                w.setValue(default)
                w.setStyleSheet("font-size:11px; padding:3px;")
                w.setFixedWidth(64)
            cl.addWidget(w)
            self.sim_widgets[prop] = w
            self._sim_inputs_layout.addWidget(col)
        self._sim_inputs_layout.addStretch()

    def _get_sim_context(self):
        """Build eval context dict from simulator widget values."""
        ctx = {"use_uh": False, "object_type": self.active_obj_type}
        for prop, w in self.sim_widgets.items():
            if isinstance(w, QSpinBox):
                ctx[prop] = w.value()
            else:
                val = w.currentText()
                if val == "True":
                    ctx[prop] = True
                elif val == "False":
                    ctx[prop] = False
                else:
                    try:
                        ctx[prop] = int(val)
                    except ValueError:
                        ctx[prop] = val
        return ctx

    def _sim_matching_indices(self):
        """Return set of rule indices that fire given simulator context."""
        if not self.sim_widgets:
            return set()
        ctx = self._get_sim_context()
        hits = set()
        for i, rule in enumerate(self.rules):
            if rule.get("object") != self.active_obj_type:
                continue
            cond = rule.get("condition", "True") or "True"
            try:
                if eval(cond, {"__builtins__": {}, "math": __import__("math")}, ctx):
                    hits.add(i)
            except Exception:
                pass
        return hits

    def _run_sim(self):
        """Run simulator: highlight cards, populate BOM preview table."""
        ctx = self._get_sim_context()
        hits = self._sim_matching_indices()
        self._refresh_card_list()   # re-render with highlights

        # Populate BOM table
        self._sim_table.setRowCount(0)
        self._sim_table.setVisible(bool(hits))
        mat_count = lab_count = 0
        for i in sorted(hits):
            rule = self.rules[i]
            r_type = rule.get("type", "")
            formula = rule.get("formula", "1")
            try:
                qty = eval(formula, {"__builtins__": {}, "math": __import__("math")}, ctx)
                qty_str = f"{qty:.3f}".rstrip("0").rstrip(".")
            except Exception:
                qty_str = formula

            row = self._sim_table.rowCount()
            self._sim_table.insertRow(row)
            self._sim_table.setItem(row, 0, QTableWidgetItem(r_type))
            self._sim_table.setItem(row, 1, QTableWidgetItem(rule.get("item_name", "")))
            self._sim_table.setItem(row, 2, QTableWidgetItem(qty_str))
            self._sim_table.setItem(row, 3, QTableWidgetItem(formula))
            if r_type == "Material":
                mat_count += 1
            else:
                lab_count += 1

        total = len(hits)
        self._sim_count_lbl.setText(
            f"{total} rule(s) fire  |  {mat_count} material, {lab_count} labor"
            if total else "No rules matched."
        )

    # ── RIGHT: editor ─────────────────────────────────────────────────────────

    def _build_right_panel(self):
        self._editor_outer = QWidget()
        self._editor_outer.setFixedWidth(400)
        self._editor_outer.setStyleSheet("border-left:1px solid #ddd; background:white;")
        lay = QVBoxLayout(self._editor_outer)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._editor_header = QLabel("Select a rule to edit")
        self._editor_header.setStyleSheet(
            "font-weight:bold; font-size:13px; padding:10px 14px;"
            "border-bottom:1px solid #ddd;"
        )
        lay.addWidget(self._editor_header)

        self._editor_body = QWidget()
        self._editor_body_layout = QVBoxLayout(self._editor_body)
        self._editor_body_layout.setContentsMargins(14, 12, 14, 12)
        self._editor_body_layout.setSpacing(10)
        self._editor_body_layout.addStretch()

        scroll = QScrollArea()
        scroll.setWidget(self._editor_body)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        lay.addWidget(scroll, 1)

        # Footer
        footer = QWidget()
        footer.setStyleSheet("border-top:1px solid #ddd; background:white;")
        fl = QHBoxLayout(footer)
        fl.setContentsMargins(12, 8, 12, 8)
        self._del_btn = QPushButton("🗑 Delete")
        self._del_btn.setStyleSheet(
            "color:#c0392b; border:1px solid #c0392b; padding:5px 10px;"
            "border-radius:4px; background:white; font-size:12px;"
        )
        self._del_btn.clicked.connect(self.delete_selected_rule)
        self._del_btn.setEnabled(False)
        self._save_btn = QPushButton("💾 Save rule")
        self._save_btn.setStyleSheet(
            "background:#27ae60; color:white; border:none; padding:5px 14px;"
            "border-radius:4px; font-weight:bold; font-size:12px;"
        )
        self._save_btn.clicked.connect(self.save_rule_changes)
        self._save_btn.setEnabled(False)
        fl.addWidget(self._del_btn)
        fl.addStretch()
        fl.addWidget(self._save_btn)
        lay.addWidget(footer)
        return self._editor_outer

    def _sec_lbl(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "font-size:10px; font-weight:bold; color:#888; text-transform:uppercase;"
            "padding-bottom:4px; border-bottom:1px solid #eee; letter-spacing:.05em;"
        )
        return lbl

    def _field_row(self, label_text, widget):
        row = QWidget()
        rl = QHBoxLayout(row)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)
        lbl = QLabel(label_text)
        lbl.setStyleSheet("font-size:11px; color:#666;")
        lbl.setFixedWidth(54)
        rl.addWidget(lbl)
        rl.addWidget(widget, 1)
        return row

    def _clear_editor(self):
        self._clear_layout(self._editor_body_layout)
        self._editor_body_layout.addStretch()
        self._editor_header.setText("Select a rule to edit")
        self._save_btn.setEnabled(False)
        self._del_btn.setEnabled(False)
        self.condition_widgets = []
        self.selected_result_item = None

    def _build_editor(self, rule):
        self._clear_layout(self._editor_body_layout)
        self.condition_widgets = []
        self.selected_result_item = None

        self._editor_header.setText(f"Editing: {rule.get('item_name','')}")
        self._save_btn.setEnabled(True)
        self._del_btn.setEnabled(True)

        # Item section
        self._editor_body_layout.addWidget(self._sec_lbl("Item"))

        self._type_combo = QComboBox()
        self._type_combo.addItems(["Material", "Labor"])
        self._type_combo.setCurrentText(rule.get("type", "Material"))
        self._editor_body_layout.addWidget(self._field_row("Type", self._type_combo))

        self._item_display = QLineEdit(rule.get("item_name", ""))
        self._item_display.setReadOnly(True)
        self._item_display.setStyleSheet("background:#f5f5f5; font-size:12px;")
        change_btn = QPushButton("Change…")
        change_btn.setStyleSheet("font-size:11px; padding:3px 8px;")
        change_btn.clicked.connect(self.search_database_for_item)
        item_row = QWidget()
        ir = QHBoxLayout(item_row)
        ir.setContentsMargins(0, 0, 0, 0)
        ir.setSpacing(4)
        lbl = QLabel("Item")
        lbl.setStyleSheet("font-size:11px; color:#666;")
        lbl.setFixedWidth(54)
        ir.addWidget(lbl)
        ir.addWidget(self._item_display, 1)
        ir.addWidget(change_btn)
        self._editor_body_layout.addWidget(item_row)

        self._code_display = QLineEdit(rule.get("item_code", ""))
        self._code_display.setReadOnly(True)
        self._code_display.setStyleSheet("background:#f5f5f5; font-size:11px; color:#888;")
        self._editor_body_layout.addWidget(self._field_row("Code", self._code_display))

        # Conditions section
        self._editor_body_layout.addWidget(self._sec_lbl("Conditions"))

        self._cond_container = QWidget()
        self._cond_rows_layout = QVBoxLayout(self._cond_container)
        self._cond_rows_layout.setContentsMargins(0, 0, 0, 0)
        self._cond_rows_layout.setSpacing(3)
        self._editor_body_layout.addWidget(self._cond_container)

        add_btn = QPushButton("+ add condition row")
        add_btn.setStyleSheet(
            "color:#185FA5; background:none; border:none; font-size:11px; text-align:left;"
        )
        add_btn.clicked.connect(self.add_condition_row)
        self._editor_body_layout.addWidget(add_btn)

        # Preview
        self._editor_body_layout.addWidget(self._sec_lbl("Condition preview"))
        self._preview_lbl = QLabel("")
        self._preview_lbl.setWordWrap(True)
        self._preview_lbl.setStyleSheet(
            "background:#f0f0f0; border-radius:4px; padding:5px 8px;"
            "font-family:monospace; font-size:11px; color:#333;"
        )
        self._editor_body_layout.addWidget(self._preview_lbl)

        # Formula
        self._editor_body_layout.addWidget(self._sec_lbl("Quantity formula"))
        avail = FORMULA_VARS.get(rule.get("object", ""), [])
        hint = QLabel(f"vars: {', '.join(avail)}" if avail else "no formula vars")
        hint.setStyleSheet("font-size:10px; color:#aaa;")
        self._editor_body_layout.addWidget(hint)
        self._formula_input = QLineEdit(rule.get("formula", "1"))
        self._formula_input.setStyleSheet("font-family:monospace; font-size:12px;")
        self._editor_body_layout.addWidget(self._formula_input)

        self._editor_body_layout.addStretch()
        self._parse_conditions(rule)
        self._update_preview()

    # ── Condition rows ────────────────────────────────────────────────────────

    def _parse_conditions(self, rule):
        cond = rule.get("condition", "")
        if not cond or cond.strip() == "True":
            self.add_condition_row()
            return
        tokens = re.split(r"\s+(and|or)\s+", cond, flags=re.IGNORECASE)
        self.add_condition_row(expression=tokens[0])
        for i in range(1, len(tokens), 2):
            logic = tokens[i].upper()
            expr = tokens[i + 1] if i + 1 < len(tokens) else ""
            self.add_condition_row(logical_op=logic, expression=expr)

    def add_condition_row(self, logical_op=None, expression=None):
        obj = self.active_obj_type
        props = list(self.property_data.get(obj, {}).keys())

        row_w = QWidget()
        rl = QHBoxLayout(row_w)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(3)

        logic_cb = QComboBox()
        logic_cb.addItems(["AND", "OR"])
        logic_cb.setFixedWidth(52)
        logic_cb.setVisible(len(self.condition_widgets) > 0)
        if logical_op:
            logic_cb.setCurrentText(logical_op)
        logic_cb.currentTextChanged.connect(self._update_preview)

        prop_cb = QComboBox()
        prop_cb.addItems(props)

        op_cb = QComboBox()
        op_cb.addItems(["==", "!=", ">", "<", ">=", "<="])
        op_cb.setFixedWidth(50)
        op_cb.currentTextChanged.connect(self._update_preview)

        val_w = QLineEdit()
        val_w.textChanged.connect(self._update_preview)

        rem_btn = QPushButton("✕")
        rem_btn.setFixedWidth(22)
        rem_btn.setStyleSheet("color:#aaa; border:none; background:none; font-size:11px;")

        rl.addWidget(logic_cb)
        rl.addWidget(prop_cb, 2)
        rl.addWidget(op_cb)
        rl.addWidget(val_w, 2)
        rl.addWidget(rem_btn)

        wm = {"widget": row_w, "logic": logic_cb, "prop": prop_cb,
              "op": op_cb, "value": val_w}
        self.condition_widgets.append(wm)
        self._cond_rows_layout.addWidget(row_w)

        prop_cb.currentTextChanged.connect(
            lambda t, w=wm: self._on_prop_change(t, w)
        )
        rem_btn.clicked.connect(lambda _, w=row_w: self._remove_cond_row(w))

        self._on_prop_change(prop_cb.currentText(), wm)

        if expression:
            self._restore_expr(expression.strip(), wm, op_cb)

        self._update_preview()

    def _on_prop_change(self, prop, wm):
        obj = self.active_obj_type
        prop_info = self.property_data.get(obj, {}).get(prop)
        cur = wm["value"]

        if isinstance(prop_info, list):
            cls = QComboBox
        elif prop_info == "int":
            cls = QSpinBox
        else:
            cls = QLineEdit

        if not isinstance(cur, cls):
            new_w = cls()
            if isinstance(new_w, QSpinBox):
                new_w.setRange(-100000, 100000)
                new_w.valueChanged.connect(self._update_preview)
            elif isinstance(new_w, QComboBox):
                new_w.currentTextChanged.connect(self._update_preview)
            else:
                new_w.textChanged.connect(self._update_preview)
            wm["widget"].layout().replaceWidget(cur, new_w)
            cur.deleteLater()
            wm["value"] = new_w
            cur = new_w

        if isinstance(cur, QComboBox) and isinstance(prop_info, list):
            cur.blockSignals(True)
            cur.clear()
            cur.addItems([str(p) for p in prop_info])
            cur.blockSignals(False)

        self._update_preview()

    def _restore_expr(self, expr, wm, op_cb):
        not_m = re.match(r"^not\s+(\w+)$", expr)
        if not_m:
            wm["prop"].setCurrentText(not_m.group(1))
            self._on_prop_change(not_m.group(1), wm)
            op_cb.setCurrentText("==")
            v = wm["value"]
            (v.setCurrentText if isinstance(v, QComboBox) else v.setText)("False")
            return
        plain_m = re.match(r"^(\w+)$", expr)
        if plain_m:
            wm["prop"].setCurrentText(plain_m.group(1))
            self._on_prop_change(plain_m.group(1), wm)
            op_cb.setCurrentText("==")
            v = wm["value"]
            (v.setCurrentText if isinstance(v, QComboBox) else v.setText)("True")
            return
        m = re.match(r"(\w+)\s*([<>=!]+)\s*(.*)", expr)
        if not m:
            return
        prop, op, val = m.group(1).strip(), m.group(2).strip(), m.group(3).strip().strip("'\"")
        wm["prop"].setCurrentText(prop)
        self._on_prop_change(prop, wm)
        op_cb.setCurrentText(op)
        v = wm["value"]
        if isinstance(v, QComboBox):
            v.setCurrentText(val)
        elif isinstance(v, QSpinBox):
            try:
                v.setValue(int(float(val)))
            except ValueError:
                pass
        else:
            v.setText(val)

    def _remove_cond_row(self, widget):
        if len(self.condition_widgets) <= 1:
            return
        self.condition_widgets = [w for w in self.condition_widgets if w["widget"] is not widget]
        widget.deleteLater()
        if self.condition_widgets:
            self.condition_widgets[0]["logic"].setVisible(False)
        self._update_preview()

    def _build_condition_parts(self):
        parts = []
        for i, wm in enumerate(self.condition_widgets):
            prop = wm["prop"].currentText()
            op = wm["op"].currentText()
            v = wm["value"]
            if isinstance(v, QSpinBox):
                val = str(v.value())
            elif isinstance(v, QComboBox):
                val = v.currentText()
            else:
                val = v.text().strip()

            if not prop:
                continue
            if i > 0:
                parts.append(wm["logic"].currentText().lower())

            if val.lower() in ("true", "false") or re.match(r"^-?\d+(\.\d+)?$", val):
                parts.append(f"{prop} {op} {val}")
            else:
                parts.append(f"{prop} {op} '{val}'")
        return parts

    def _update_preview(self):
        parts = self._build_condition_parts()
        text = " ".join(parts) if parts else "(no conditions)"
        if hasattr(self, "_preview_lbl"):
            self._preview_lbl.setText(text)

    # ── Editor actions ────────────────────────────────────────────────────────

    def search_database_for_item(self):
        db_type, ok = QInputDialog.getItem(
            self, "Select type", "Which database?", ["Material", "Labor"], 0, False
        )
        if not (ok and db_type):
            return
        dlg = SearchDialog(db_type, self)
        if dlg.exec():
            item = dlg.get_selected()
            if item:
                self.selected_result_item = item
                self._item_display.setText(item["name"])
                self._code_display.setText(item.get("code", ""))
                self._type_combo.setCurrentText(item["type"])

    def save_rule_changes(self):
        if self.selected_rule_index == -1:
            return
        rule = self.rules[self.selected_rule_index]
        rule["condition"] = " ".join(self._build_condition_parts())
        if self.selected_result_item:
            rule["type"] = self.selected_result_item["type"]
            rule["item_code"] = self.selected_result_item.get("code", "")
            rule["item_name"] = (
                self.selected_result_item.get("name")
                or self.selected_result_item.get("item_name", "")
            )
        else:
            rule["type"] = self._type_combo.currentText()
        rule["formula"] = self._formula_input.text().strip() or "1"
        self.save_rules()
        self._update_tree_counts()
        self._refresh_card_list()
        self._editor_header.setText(f"Editing: {rule.get('item_name','')}")
        QMessageBox.information(self, "Saved", "Rule saved.")

    def create_new_rule(self):
        new_rule = {
            "object": self.active_obj_type,
            "item_name": "New Rule — edit me",
            "condition": "",
            "type": "Material",
            "item_code": "N/A",
            "formula": "1",
        }
        self.rules.append(new_rule)
        self.save_rules()
        self.selected_rule_index = len(self.rules) - 1
        self._update_tree_counts()
        self._refresh_card_list()
        self._build_editor(new_rule)

    def delete_selected_rule(self):
        if self.selected_rule_index == -1:
            return
        rule = self.rules[self.selected_rule_index]
        reply = QMessageBox.question(
            self, "Delete rule",
            f"Delete:\n'{rule.get('item_name')}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            del self.rules[self.selected_rule_index]
            self.selected_rule_index = -1
            self.save_rules()
            self._update_tree_counts()
            self._refresh_card_list()
            self._clear_editor()

    # ── Persistence ───────────────────────────────────────────────────────────

    def load_rules(self):
        try:
            with open("rules.json", "r") as f:
                self.rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.rules = []

    def save_rules(self):
        try:
            with open("rules.json", "w") as f:
                json.dump(self.rules, f, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save rules.json:\n{e}")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _clear_layout(self, layout):
        if not layout:
            return
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self._clear_layout(child.layout())
