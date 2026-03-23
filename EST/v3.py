import sys
import math
import json
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QComboBox, QGraphicsView, QGraphicsScene, 
                             QGraphicsTextItem, QFormLayout, QGroupBox, QSpinBox, QGraphicsPathItem, 
                             QLineEdit, QFileDialog, QMessageBox, QCheckBox, QTableWidget, QTableWidgetItem,
                             QHeaderView, QSplitter, QDialog, QListWidget)
from PyQt6.QtGui import QPen, QBrush, QColor, QPainterPath, QTextOption, QPainter, QPageLayout, QFont
from PyQt6.QtCore import Qt, QTimer, QRectF
from PyQt6.QtPrintSupport import QPrinter

# --- 1. EXACT OFFICIAL DATABASE ENGINE ---
def setup_database():
    conn = sqlite3.connect('erp_master.db')
    cursor = conn.cursor()
    cursor.execute('''DROP TABLE IF EXISTS materials''')
    cursor.execute('''DROP TABLE IF EXISTS labor''')
    
    cursor.execute('''CREATE TABLE materials (item_code TEXT, item_name TEXT PRIMARY KEY, rate REAL, unit TEXT)''')
    cursor.execute('''CREATE TABLE labor (task_name TEXT PRIMARY KEY, rate REAL, unit TEXT)''')
    
    materials = [
        ("110030141", "P C C POLE:8 Mtrs.Long", 5363.44, "NOS"), ("110030241", "P C C POLE:9 Mtrs.Long", 10198.28, "NOS"),
        ("301018141", "Dist. Transformer 25KVA", 103528.11, "NOS"), ("301018241", "Dist. Transformer 63KVA", 107589.53, "NOS"),
        ("102010611", "M.S Channel 75X40 mm", 110043.09, "MT"), ("101011311", "M.S Angle 65X65X6mm", 108667.24, "MT"),
        ("103011511", "M.S Flat 65X6 mm", 117493.74, "MT"), ("505030641", "Suspension Clamp Suitable for 35sq.mm. Messenger Conductor", 367.62, "NOS"),
        ("505034141", "Dead end clamp LT ABC", 389.66, "NOS"), ("508040441", "Shakle Insulator", 23.34, "NOS"), 
        ("508030541", "11 KV Polymer Disc Insulator 45KN", 183.15, "NOS"), ("508011141", "11 KV Polymer Pin Insulator 45KN", 243.79, "NOS"),
        ("504010132", "Hardware fittings 11KV", 327.83, "SET"), ("504070441", "LT Spacer 3 PHASE 4 WIRE", 77.62, "NOS"),
        ("502010921", "ACSR Conductor 50SQMM", 62290.12, "KM"), ("501030521", "LT AB CABLE 1.1KV 3CX50+1CX16+1CX35sqmm", 315558.99, "KM"), 
        ("504110541", "G.I. Earth Spike 6*3.25ft", 367.98, "NOS"), ("504130432", "H.T. Stay Set Complete", 795.83, "SET"),
        ("504130332", "LT Stay set", 462.17, "SET"), ("508040841", "H.T. Guy Insulator 11KV", 52.24, "NOS"),
        ("508040741", "LT Guy Insulator", 21.11, "NOS"), ("503050711", "G.I. Stay Wire 7/3..15MM 10 SWG(HT)", 142310.93, "MT"),
        ("503050611", "GI STAY WIRE 7/12 SWG", 145404.60, "MT"), ("503010711", "G.I. 8 SWG Wire (4mm)", 137360.98, "MT"),
        ("503010811", "G.I. 6 SWG Wire (5mm)", 136865.98, "MT"), ("910010241", "Caution Board-11KVA", 160.18, "NOS"),
        ("195021741", "UH-LT BKT 4 WAY", 500.00, "NOS"), ("597011541", "UH-CLAMP FOR 8 MTR PCC POLE", 150.00, "NOS"),
        ("597011741", "UH-Diron Clump", 40.00, "NOS"), ("304010532", "T.P.G.O. Isolator (200Amps) 11KV", 10384.98, "SET"),
        ("309010841", "Lightning Arrestor 12 KV", 524.23, "NOS"), ("912011441", "G.I. Turn Buckle", 238.70, "NOS"),
        ("407010641", "LT Distribution KIOSK FOR 25 KVA DTR", 8878.41, "NOS"), ("501017821", "PVC Cable 4 Core 25SQMM", 229429.95, "KM"),
        ("504060941", "LT Distribution Box along with steel Strap & Buckle for 3ph connection in ABC system", 1489.9, "NOS"),
        ("501017421", "CABLE (PVC 1.1 KV GRADE) 4Core X10 sq mm", 125852.36, "KM"), ("501017721", "CABLE (PVC 1.1 KV GRADE) 4CX16 sq mm", 119154.63, "KM"),
        ("501017821", "CABLE (PVC 1.1 KV GRADE) 4CX25 sq mm", 229429.95, "KM")
    ]
    labor = [
        ("Erection of . 8mtr  PCC Pole ( LT)", 1501.00, "NOS"), ("Erection of . 8mtr  PCC Pole (HT)", 1680.00, "NOS"),
        ("Erection of . 9mtr  PCC Pole (HT)", 2620.00, "NOS"), ("Erection of . 9mtr  PCC Pole (HT) Without Painted", 2620.00, "NOS"),
        ("Erection of S/S D.P. Structure  (8 mtr without Painted)", 9875.00, "NOS"), ("Sub-Stationn Str with 9 Mtr PCC pole DP", 13169.00, "NOS"),
        ("Sub-Stationn Str with 9 Mtr PCC pole DP Without Painted", 13169.00, "NOS"),
        ("Erection of 8 mtr D.P structure (HT)", 5654.00, "NOS"), ("Erection of 9 MTR Long PCC pole D/P for HTOH line", 9438.00, "NOS"), 
        ("Erection of 25 KVA Transformer", 1925.00, "NOS"),
        ("Stringing & Sagging with 50 sq.mm A.C.S.R. 3 Wire", 8289.00, "KM"), ("Strining& saging with ACSR 50sqmm 4wire", 9715.00, "KM"),
        ("Stringing & Sagging of LT AB Cable", 46000.00, "KM"), ("H.T. Stay Set Complete Labor", 641.00, "SET"),
        ("LT Stay set complete", 555.00, "SET"), ("Earthing Complete", 313.00, "NOS"),
        ("Fabrication & Fixing  of Cattle Guard Bracket (SP)", 237.00, "NOS"), ("Fabrication & Fixing  of Cattle Guard Bracket (DP)", 369.00, "NOS"),
        ("Extension of 8 mtr PCC Pole (Without Painted)HT", 1506.00, "NOS"), ("Fixing Cross lacing", 15.00, "NOS"), 
        ("Lead Wire above above 60 Mtrs (2 Wire)", 506.00, "NOS"), ("Fixing of Caution Board", 24.00, "NOS"), 
        ("Fixing of LT Bracket(Without Painted)", 596.00, "NOS"), ("Pole GIS survey", 31.00, "NOS"),
        ("Fixing of 11 KV Pin Insulator", 63.00, "NOS"), ("Fixing of 11 KV Disc Insulator", 65.00, "NOS"), 
        ("Fixing of LT Shackle Insulator (with N/B)", 52.00, "NOS"), ("Fixing of LT spacer", 56.00, "NOS"),
        ("Fixing of  11 KV TGPO Isolator on S/Stn Structure", 1193.00, "SET"), ("Fixing of neutral earthing of DTR WITH G", 3816.00, "NOS"),
        ("Fixing of 11 KV Lightning Arrestor", 339.00, "SET"), ("FIXING OF LT Distribution KIOSK FOR 25 KVA DTR", 2155.00, "NOS"),
        ("Erection of Anchoring/Suspension Clamp", 154.00, "NOS"), ("Survey for H.T.O.H Line", 2761.00, "KM"), ("Survey for L.T.O.H Line", 1714.00, "KM"),
        ("Fixing of Solid Tee-off Bracket on S.P", 1495.00, "NOS"), ("Fixing of Solid Tee-off Bracket on D.P", 1483.00, "NOS"),
        ("DTR Code Painting", 65.00, "NOS"), 
        ("Fixing of 3ph Service Connection (Cable provided)", 570.00, "NOS"), ("Fixing of 1ph Service Connection (Cable provided)", 270.00, "NOS"),
        ("Fixing of 3ph Service Connection (No Cable)", 6117.00, "NOS"), ("Fixing of 1ph Service Connection (No Cable)", 1578.00, "NOS"),
        ("Erection of distribution box", 507.00, "NOS"), ("Laying & Dressing of 1.1 KV PVC/XLPE 2x10,4x10/16, 3.5/4x25 Sqmm Cable", 15000.00, "KM")
    ]
    cursor.executemany('INSERT INTO materials VALUES (?,?,?,?)', materials)
    cursor.executemany('INSERT INTO labor VALUES (?,?,?)', labor)
    conn.commit(); conn.close()

# --- 2. SEARCH DIALOG & CUSTOM COMPONENTS ---
class SearchDialog(QDialog):
    def __init__(self, db_type, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Search & Add {db_type}")
        self.setFixedSize(600, 400); self.layout = QVBoxLayout(self)
        self.search_box = QLineEdit(); self.search_box.setPlaceholderText("Type to search official items...")
        self.layout.addWidget(self.search_box)
        
        self.list_widget = QListWidget()
        self.layout.addWidget(self.list_widget)
        self.search_box.textChanged.connect(self.filter_list)
        
        self.add_btn = QPushButton("Add Selected to Estimate")
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        self.add_btn.clicked.connect(self.accept)
        self.layout.addWidget(self.add_btn)

        self.items_data = {}; self.load_data(db_type)

    def load_data(self, db_type):
        conn = sqlite3.connect('erp_master.db'); cursor = conn.cursor()
        if db_type == "Material": cursor.execute("SELECT item_code, item_name, unit, rate FROM materials")
        else: cursor.execute("SELECT '' as item_code, task_name, unit, rate FROM labor")
        for row in cursor.fetchall():
            display_text = f"{row[1]} ({row[2]}) - Rs. {row[3]}"
            self.items_data[display_text] = {"code": row[0], "name": row[1], "unit": row[2], "rate": row[3], "type": db_type}
            self.list_widget.addItem(display_text)
        conn.close()

    def filter_list(self, text):
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i); item.setHidden(text.lower() not in item.text().lower())

    def get_selected(self):
        selected = self.list_widget.currentItem()
        if selected: return self.items_data[selected.text()]
        return None

class InteractiveView(QGraphicsView):
    def __init__(self, scene, parent_app):
        super().__init__(scene)
        self.parent_app = parent_app
        self.setRenderHints(self.renderHints() | self.renderHints().Antialiasing)
        self.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)

    def wheelEvent(self, event):
        zoom = 1.15 if event.angleDelta().y() > 0 else 1.0 / 1.15
        self.scale(zoom, zoom)
        self.parent_app.update_view_drag_mode()

    def mousePressEvent(self, event):
        self.parent_app.handle_canvas_click(event, self)
        super().mousePressEvent(event)

class DraggableLabel(QGraphicsTextItem):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFlag(QGraphicsTextItem.GraphicsItemFlag.ItemIsMovable)
        self.setFlag(QGraphicsTextItem.GraphicsItemFlag.ItemIsSelectable)
        self.setTextInteractionFlags(Qt.TextInteractionFlag.NoTextInteraction)
        self.document().setDefaultTextOption(QTextOption(Qt.AlignmentFlag.AlignCenter)); self.setZValue(20)
        self.setFont(QFont("Arial", 7))

    def mouseDoubleClickEvent(self, event):
        self.setTextInteractionFlags(Qt.TextInteractionFlag.TextEditorInteraction)
        self.setFocus(Qt.FocusReason.MouseFocusReason)
        super().mouseDoubleClickEvent(event)

    def focusOutEvent(self, event):
        self.setTextInteractionFlags(Qt.TextInteractionFlag.NoTextInteraction)
        super().focusOutEvent(event)

    def paint(self, painter, option, widget):
        painter.setBrush(QColor(255, 255, 255)) # White, 100% opacity
        painter.setPen(Qt.PenStyle.NoPen)
        
        doc = self.document()
        layout = doc.documentLayout()
        
        for i in range(doc.blockCount()):
            block = doc.findBlockByNumber(i)
            if not block.isValid():
                continue
            
            text_layout = block.layout()
            if not text_layout:
                continue
                
            block_rect = layout.blockBoundingRect(block)
            line = text_layout.lineAt(0)
            if line.isValid():
                used_width = line.naturalTextWidth()
                offset_x = (block_rect.width() - used_width) / 2
                
                highlighter_rect = QRectF(
                    block_rect.left() + offset_x, 
                    block_rect.top(), 
                    used_width, 
                    block_rect.height()
                )
                painter.drawRect(highlighter_rect.adjusted(-3, 0, 3, 0))

        super().paint(painter, option, widget)

class SmartPole(QGraphicsPathItem):
    def __init__(self, x, y, pole_type="LT", is_existing=False):
        super().__init__(); self.setPos(x, y); self.setZValue(10) 
        self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemIsSelectable); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemIsMovable); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemSendsGeometryChanges)
        self.pole_type = pole_type; self.is_existing = is_existing; self.height = "8MTR" if pole_type == "LT" else "9MTR"
        self.has_extension = False

        if self.is_existing: self.dtr_size = "None"; self.earth_count = 0; self.stay_count = 0
        elif self.pole_type == "DTR": self.dtr_size = "None"; self.earth_count = 2; self.stay_count = 4  
        else: self.dtr_size = "None"; self.earth_count = 1; self.stay_count = 0

        self.connected_spans = []
        self.label = DraggableLabel(self); self.label.setTextWidth(80)
        self.ext_label = DraggableLabel(self); self.ext_label.setFont(QFont("Arial", 6, QFont.Weight.Bold)); self.ext_label.setPlainText("Ext"); self.ext_label.setPos(-10, -30)
        self.update_visuals()

    def update_visuals(self):
        path = QPainterPath()
        if self.pole_type == "DTR":
            path.addEllipse(-8, -20, 16, 16); path.addEllipse(-8, 4, 16, 16); path.moveTo(0, -4); path.lineTo(0, 4); self.label.setPos(-40, 20) 
        else: path.addEllipse(-10, -10, 20, 20); self.label.setPos(-40, 12) 
        self.setPath(path)

        self.ext_label.setVisible(self.has_extension)

        if self.is_existing:
            self.setBrush(QBrush(QColor("#dddddd"))) # Lighter gray
            self.setPen(QPen(Qt.GlobalColor.black, 1, Qt.PenStyle.SolidLine))
            if self.pole_type == 'DTR':
                lbl_text = "Existing DP/DTR"
            else:
                lbl_text = f"Existing {self.pole_type} Pole"
        else:
            self.setPen(QPen(Qt.GlobalColor.black, 1))
            if self.pole_type == "LT": self.setBrush(QBrush(Qt.GlobalColor.blue)); lbl_text = f"LT Pole ({self.height[:-2]})"
            elif self.pole_type == "HT": self.setBrush(QBrush(Qt.GlobalColor.red)); lbl_text = f"HT Pole ({self.height[:-2]})"
            elif self.pole_type == "DTR": self.setBrush(QBrush(Qt.GlobalColor.green)); lbl_text = "DP Structure" if self.dtr_size == "None" else f"DP Structure\n{self.dtr_size} DTR"
        
        if self.earth_count > 0: lbl_text += f"\n+ {self.earth_count} Earth"
        if self.stay_count > 0: lbl_text += f"\n+ {self.stay_count} Stay"
        self.label.setPlainText(lbl_text)

    def itemChange(self, change, value):
        if change == QGraphicsPathItem.GraphicsItemChange.ItemPositionHasChanged:
            for span in self.connected_spans: span.update_position()
        return super().itemChange(change, value)

class SmartHome(QGraphicsPathItem):
    def __init__(self, x, y):
        super().__init__(); self.setPos(x, y); self.setZValue(10); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemIsSelectable); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemIsMovable); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemSendsGeometryChanges)
        self.connected_spans = []; path = QPainterPath(); path.addRect(-10, 0, 20, 20); path.moveTo(-15, 0); path.lineTo(0, -15); path.lineTo(15, 0); path.closeSubpath()
        self.setPath(path); self.setBrush(QBrush(Qt.GlobalColor.yellow)); self.setPen(QPen(Qt.GlobalColor.black, 1)); self.label = DraggableLabel(self); self.label.setTextWidth(60); self.label.setPos(-30, 22); self.label.setPlainText("Consumer\nHome")
    def itemChange(self, change, value):
        if change == QGraphicsPathItem.GraphicsItemChange.ItemPositionHasChanged:
            for span in self.connected_spans: span.update_position()
        return super().itemChange(change, value)

class SmartSpan(QGraphicsPathItem):
    def __init__(self, pole1, pole2):
        super().__init__(); self.p1 = pole1; self.p2 = pole2; self.setZValue(0); self.setFlag(QGraphicsPathItem.GraphicsItemFlag.ItemIsSelectable)
        self.is_existing_span = getattr(self.p1, 'is_existing', False) and getattr(self.p2, 'is_existing', False)

        self.is_service_drop = isinstance(self.p1, SmartHome) or isinstance(self.p2, SmartHome)
        if self.is_service_drop:
            self.conductor = "Service Drop"; self.length = 20; self.consider_cable = False; self.cable_size = "10 SQMM"; self.phase = "3 Phase"; self.has_cg = False
            self.aug_type = "New"; self.wire_count = "3"; self.wire_size = "50SQMM"
        else:
            is_ht = getattr(self.p1, 'pole_type', 'LT') != 'LT' or getattr(self.p2, 'pole_type', 'LT') != 'LT'
            self.conductor = "ACSR" if is_ht else "AB Cable"
            self.length = 40; self.aug_type = "New"
            self.wire_count = "3"; self.wire_size = "50SQMM"; self.cable_size = "25 SQMM"; self.has_cg = False 

        self.label = DraggableLabel(); self.label.setTextWidth(80); self.update_position(); self.update_visuals()

    def update_position(self):
        path = QPainterPath()
        path.moveTo(self.p1.x(), self.p1.y())
        dx = self.p2.x() - self.p1.x()
        dy = self.p2.y() - self.p1.y()
        px_length = math.hypot(dx, dy)

        # Wavy line for ABC, PVC, and Service Drops
        if self.conductor in ["AB Cable", "PVC Cable", "Service Drop"] and px_length > 0:
            steps = max(20, int(px_length / 2)) # More steps for a smoother curve
            nx = -dy / px_length  # Normalized normal vector x
            ny = dx / px_length   # Normalized normal vector y
            
            wave_wavelength_pixels = 15
            frequency = px_length / wave_wavelength_pixels
            amplitude = 4

            for i in range(1, steps + 1):
                t = i / float(steps)
                cx = self.p1.x() + dx * t
                cy = self.p1.y() + dy * t
                sine_offset = math.sin(t * frequency * 2 * math.pi) * amplitude
                path.lineTo(cx + nx * sine_offset, cy + ny * sine_offset)
        else: # Straight line for ACSR
            path.lineTo(self.p2.x(), self.p2.y())

        self.setPath(path)
        nx_norm = -dy / (px_length if px_length > 0 else 1)
        ny_norm = dx / (px_length if px_length > 0 else 1)
        mid_x = (self.p1.x() + self.p2.x()) / 2
        mid_y = (self.p1.y() + self.p2.y()) / 2
        self.label.setPos(mid_x + (nx_norm * 15) - 40, mid_y + (ny_norm * 15) - 10)

    def update_visuals(self):
        self.update_position()
        pen = QPen(Qt.GlobalColor.black, 1.5)

        if self.is_existing_span:
            pen.setColor(Qt.GlobalColor.black) # 3. make existing line color to black instead of grey.
            # 1. ACSR is hard line, ABC/PVC is wave (handled in update_position)
            pen.setStyle(Qt.PenStyle.SolidLine)
        else: # New line
            # Only new ACSR is dashed. New AB Cable is solid (and wavy).
            if self.conductor == "ACSR":
                 pen.setStyle(Qt.PenStyle.DashLine)
            
            if self.conductor == "PVC Cable":
                pen.setColor(QColor("#107C41"))
        
        self.setPen(pen)
        
        # 4. show Ex. ABC, or Ex. ACSR or Ex. PVC, in place of existing span.
        if self.is_existing_span:
            lbl_text = f"Ex. {self.conductor}"
        elif self.is_service_drop:
            lbl_text = f"Service Cable {self.length}m\n{self.phase}"
            if self.consider_cable: lbl_text += f"\n({self.cable_size} PVC)"
        else:
            if self.conductor == "ACSR": lbl_text = f"{self.length}m\n{self.wire_count}W ACSR"
            elif self.conductor == "PVC Cable": lbl_text = f"{self.length}m\n{self.cable_size} PVC"
            else: lbl_text = f"{self.length}m\nABC"
            if self.aug_type != "New": lbl_text += f"\n({self.aug_type})"
            if self.has_cg: lbl_text += f"\n[+CG]"
            
        self.label.setPlainText(lbl_text)
        if not self.label.scene() and self.scene(): self.scene().addItem(self.label)

# --- 4. THE MASTER APPLICATION ---
class EstimateAppV9(QMainWindow):
    def __init__(self):
        super().__init__()
        setup_database()
        self.setWindowTitle("ERP Estimate Generator - Version 3.0")
        self.setGeometry(50, 50, 1600, 900)
        self.current_tool = "SELECT"; self.span_start_pole = None; self.autosave_file = "autosave_erp.json"
        
        self.bom_overrides = {}
        self.live_bom_data = [] 
        self.escalations = [] 

        main_widget = QWidget(); main_layout = QHBoxLayout(main_widget); self.setCentralWidget(main_widget)
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self.splitter)

        left_panel = QWidget(); left_layout = QVBoxLayout(left_panel); self.splitter.addWidget(left_panel)

        file_toolbar = QHBoxLayout()
        for txt, cmd in [("📄 New", self.new_drawing), ("📂 Open", self.load_from_file), ("💾 Save", self.save_to_file)]:
            btn = QPushButton(txt); btn.clicked.connect(cmd); btn.setStyleSheet("padding: 5px; font-weight: bold;"); file_toolbar.addWidget(btn)
        
        file_toolbar.addStretch()
        pdf_btn = QPushButton("🗺️ Export PDF Drawing"); pdf_btn.clicked.connect(self.export_pdf)
        pdf_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #d32f2f; color: white;")
        xl_btn = QPushButton("📊 Generate ERP Excel"); xl_btn.clicked.connect(self.generate_excel)
        xl_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #107C41; color: white;")
        file_toolbar.addWidget(pdf_btn); file_toolbar.addWidget(xl_btn)
        left_layout.addLayout(file_toolbar)

        draw_toolbar = QHBoxLayout(); self.tools_btns = {}
        self.tools = { "SELECT": "🖱 Select / Edit", "ADD_LT": "🔵 LT Pole", "ADD_HT": "🔴 HT Pole", "ADD_DTR": "🟩 DP/DTR", "ADD_EXISTING": "⚪ Ex Pole", "ADD_HOME": "🏠 Home", "ADD_SPAN": "📏 Span"}
        for key, txt in self.tools.items():
            btn = QPushButton(txt); btn.clicked.connect(lambda checked, t=key: self.set_tool(t))
            btn.setStyleSheet("padding: 8px; font-weight: bold; background-color: lightgray;"); 
            draw_toolbar.addWidget(btn); self.tools_btns[key] = btn
        left_layout.addLayout(draw_toolbar)

        self.scene = QGraphicsScene(); self.scene.selectionChanged.connect(self.on_selection_changed)
        self.view = InteractiveView(self.scene, self); left_layout.addWidget(self.view)

        right_splitter = QSplitter(Qt.Orientation.Vertical)
        self.splitter.addWidget(right_splitter)
        self.splitter.setSizes([950, 650]) # Fixed width allocation
        
        editor_widget = QWidget(); editor_widget_layout = QVBoxLayout(editor_widget)
        self.subject_input = QLineEdit(); self.subject_input.setPlaceholderText("Enter Project Name / Subject...")
        editor_widget_layout.addWidget(QLabel("<b>Project Subject:</b>")); editor_widget_layout.addWidget(self.subject_input)
        
        lat_long_layout = QHBoxLayout()
        self.lat_input = QLineEdit(); self.lat_input.setPlaceholderText("Latitude...")
        self.long_input = QLineEdit(); self.long_input.setPlaceholderText("Longitude...")
        lat_long_layout.addWidget(QLabel("<b>Lat:</b>")); lat_long_layout.addWidget(self.lat_input)
        lat_long_layout.addWidget(QLabel("<b>Long:</b>")); lat_long_layout.addWidget(self.long_input)
        editor_widget_layout.addLayout(lat_long_layout)
        
        self.uh_checkbox = QCheckBox("Use UH (Readymade) Materials instead of Raw Steel"); self.uh_checkbox.setStyleSheet("font-weight: bold; color: #107C41;")
        self.uh_checkbox.stateChanged.connect(self.refresh_live_estimate); editor_widget_layout.addWidget(self.uh_checkbox)
        
        # UI FIX: Removed ScrollArea to make Property Box perfectly visible always
        self.editor_group = QGroupBox("1. Object Properties"); self.editor_layout = QFormLayout(); self.editor_group.setLayout(self.editor_layout)
        editor_widget_layout.addWidget(self.editor_group)
        editor_widget_layout.addStretch() # Pushes properties to top cleanly
        right_splitter.addWidget(editor_widget)

        table_widget = QWidget(); table_layout = QVBoxLayout(table_widget)
        table_layout.addWidget(QLabel("<b>2. Live Estimate (Double-Click Qty to Edit)</b>"))
        
        self.live_table = QTableWidget(0, 6)
        self.live_table.setHorizontalHeaderLabels(["Type", "Code", "Name", "Qty", "Unit", "Total (Rs)"])
        self.live_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.live_table.setColumnWidth(0, 60); self.live_table.setColumnWidth(1, 80); self.live_table.setColumnWidth(3, 60)
        self.live_table.itemChanged.connect(self.on_table_edit)
        table_layout.addWidget(self.live_table)
        
        search_layout = QHBoxLayout()
        add_mat_btn = QPushButton("+ Add Custom Material"); add_mat_btn.clicked.connect(lambda: self.open_search("Material"))
        add_lab_btn = QPushButton("+ Add Custom Labor"); add_lab_btn.clicked.connect(lambda: self.open_search("Labor"))
        add_mat_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; padding: 5px;")
        add_lab_btn.setStyleSheet("background-color: #e67e22; color: white; font-weight: bold; padding: 5px;")
        search_layout.addWidget(add_mat_btn); search_layout.addWidget(add_lab_btn)
        table_layout.addLayout(search_layout)

        self.grand_total_label = QLabel("<b>Grand Total: Rs. 0.00</b>")
        self.grand_total_label.setStyleSheet("font-size: 16px; color: #d32f2f; margin-top: 5px;")
        table_layout.addWidget(self.grand_total_label)
        right_splitter.addWidget(table_widget)
        right_splitter.setSizes([300, 700]) # Optimal height for properties vs table

        self.set_tool("SELECT"); self.load_autosave() 

    # --- UI & CANVAS LOGIC ---
    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace): self.delete_selected_items()
        super().keyPressEvent(event)
        
    def delete_selected_items(self):
        items = self.scene.selectedItems()
        for item in items:
            if isinstance(item, SmartSpan): self.delete_item(item)
        for item in items:
            if isinstance(item, (SmartPole, SmartHome)): self.delete_item(item)

    def set_tool(self, tool_name):
        self.current_tool = tool_name
        if self.span_start_pole: self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1))
        self.span_start_pole = None
        for key, btn in self.tools_btns.items():
            btn.setStyleSheet("padding: 8px; font-weight: bold; background-color: " + ("lightblue" if key == tool_name else "lightgray"))
        self.update_view_drag_mode()

    def handle_canvas_click(self, event, view):
        if event.button() == Qt.MouseButton.RightButton: self.set_tool("SELECT"); return
        if self.current_tool == "SELECT": return

        pos = view.mapToScene(event.pos()); item_clicked = self.scene.itemAt(pos, view.transform())

        if self.current_tool in ["ADD_LT", "ADD_HT", "ADD_DTR", "ADD_EXISTING"]:
            p_type = "LT" if self.current_tool == "ADD_EXISTING" else self.current_tool.split("_")[1]
            self.scene.addItem(SmartPole(pos.x(), pos.y(), p_type, self.current_tool == "ADD_EXISTING"))
            self.refresh_live_estimate()
        elif self.current_tool == "ADD_HOME":
            self.scene.addItem(SmartHome(pos.x(), pos.y())); self.refresh_live_estimate()
        elif self.current_tool == "ADD_SPAN" and isinstance(item_clicked, (SmartPole, SmartHome)):
            if not self.span_start_pole:
                self.span_start_pole = item_clicked; item_clicked.setPen(QPen(Qt.GlobalColor.yellow, 3)) 
            elif self.span_start_pole != item_clicked: 
                span = SmartSpan(self.span_start_pole, item_clicked)
                self.span_start_pole.connected_spans.append(span); item_clicked.connected_spans.append(span)
                self.scene.addItem(span); self.scene.addItem(span.label) 
                self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1)); self.span_start_pole = None
                self.refresh_live_estimate()

    def on_selection_changed(self):
        while self.editor_layout.count():
            child = self.editor_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()

        sel = self.scene.selectedItems()
        if not sel: self.editor_group.setTitle("Select an item to edit"); return
        if len(sel) > 1: self.editor_group.setTitle(f"{len(sel)} Items Selected"); return
        item = sel[0]
        
        if isinstance(item, DraggableLabel): self.editor_group.setTitle("Text Label Selected"); return
        
        if isinstance(item, SmartHome):
            self.editor_group.setTitle("Editing Consumer Home")
            self.editor_layout.addRow(QLabel("<i>Service cable connection point. Select the actual Line (Span) connecting to this home to edit PVC sizes and Phase details.</i>"))
            del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
            del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)
            return

        if isinstance(item, SmartPole):
            if item.is_existing:
                self.editor_group.setTitle("Editing Existing Pole")
                pole_type_cb = QComboBox(); pole_type_cb.addItems(["LT", "HT", "DTR"]); pole_type_cb.setCurrentText(item.pole_type)
                pole_type_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "pole_type", t)); self.editor_layout.addRow("Pole Type:", pole_type_cb)
                stay_type_cb = QComboBox(); stay_type_cb.addItems(["HT", "LT"]); stay_type_cb.setCurrentText(getattr(item, 'stay_type', 'HT'))
                stay_type_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "stay_type", t)); self.editor_layout.addRow("Stay Type:", stay_type_cb)
                stay_spin = QSpinBox(); stay_spin.setRange(0, 10); stay_spin.setValue(item.stay_count)
                stay_spin.valueChanged.connect(lambda v: self.update_pole(item, "stay_count", v)); self.editor_layout.addRow("Stay Sets:", stay_spin)
                del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
                del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)
                return

            self.editor_group.setTitle(f"Editing {item.pole_type} Structure")
            height_cb = QComboBox(); height_cb.addItems(["8MTR", "9MTR"]); height_cb.setCurrentText(item.height)
            height_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "height", t)); self.editor_layout.addRow("Height:", height_cb)

            if item.height == "8MTR" and item.pole_type in ["HT", "DTR"]:
                ext_check = QCheckBox("Add 8MTR Extension")
                ext_check.setChecked(item.has_extension)
                ext_check.stateChanged.connect(lambda v, i=item: self.update_pole_extension(i, v == 2))
                self.editor_layout.addRow(ext_check)

            if item.pole_type == "DTR":
                dtr_cb = QComboBox(); dtr_cb.addItems(["None", "16 KVA", "25KVA", "63KVA", "100KVA", "160KVA"])
                dtr_cb.setCurrentText(item.dtr_size); dtr_cb.currentTextChanged.connect(lambda t: self.update_dtr_logic(item, t)); self.editor_layout.addRow("DTR Size:", dtr_cb)
            
            earth_spin = QSpinBox(); earth_spin.setRange(0, 10); earth_spin.setValue(item.earth_count)
            earth_spin.valueChanged.connect(lambda v: self.update_pole(item, "earth_count", v)); self.editor_layout.addRow("Earthing Sets:", earth_spin)
            stay_spin = QSpinBox(); stay_spin.setRange(0, 10); stay_spin.setValue(item.stay_count)
            stay_spin.valueChanged.connect(lambda v: self.update_pole(item, "stay_count", v)); self.editor_layout.addRow("Stay Sets:", stay_spin)

        elif isinstance(item, SmartSpan):
            length_spin = QSpinBox(); length_spin.setRange(1, 150); length_spin.setValue(int(item.length))
            length_spin.valueChanged.connect(lambda v: self.update_span(item, "length", v))
            
            if item.is_service_drop:
                self.editor_group.setTitle("Editing Service Connection")
                self.editor_layout.addRow("Length (Meters):", length_spin)
                phase_cb = QComboBox(); phase_cb.addItems(["1 Phase", "3 Phase"]); phase_cb.setCurrentText(item.phase)
                phase_cb.currentTextChanged.connect(lambda t: self.update_span(item, "phase", t)); self.editor_layout.addRow("Phase:", phase_cb)
                cons_cb = QCheckBox("Consider Cable in Estimate?"); cons_cb.setChecked(item.consider_cable)
                cons_cb.stateChanged.connect(lambda v: self.update_span(item, "consider_cable", v == 2)); self.editor_layout.addRow(cons_cb)
                sz_cb = QComboBox(); sz_cb.addItems(["10 SQMM", "16 SQMM", "25 SQMM"]); sz_cb.setCurrentText(item.cable_size)
                sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "cable_size", t)); self.editor_layout.addRow("PVC Size:", sz_cb)
            else:
                self.editor_group.setTitle("Editing Span")
                self.editor_layout.addRow("Length (Meters):", length_spin)
                cg_chk = QCheckBox(); cg_chk.setChecked(item.has_cg)
                cg_chk.stateChanged.connect(lambda v: self.update_span(item, "has_cg", v == 2)); self.editor_layout.addRow("Cattle Guard:", cg_chk)
                
                cond_cb = QComboBox(); cond_cb.addItems(["ACSR", "AB Cable", "PVC Cable"]); cond_cb.setCurrentText(item.conductor)
                cond_cb.currentTextChanged.connect(lambda t: self.update_conductor_logic(item, t)); self.editor_layout.addRow("Conductor:", cond_cb)
                
                if item.conductor == "ACSR":
                    wire_cnt_cb = QComboBox(); wire_cnt_cb.addItems(["2", "3", "4"]); wire_cnt_cb.setCurrentText(item.wire_count)
                    wire_cnt_cb.currentTextChanged.connect(lambda t: self.update_span(item, "wire_count", t)); self.editor_layout.addRow("Wire Count:", wire_cnt_cb)
                    wire_sz_cb = QComboBox(); wire_sz_cb.addItems(["30SQMM", "50SQMM"]); wire_sz_cb.setCurrentText(item.wire_size)
                    wire_sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "wire_size", t)); self.editor_layout.addRow("Wire Size:", wire_sz_cb)
                elif item.conductor == "PVC Cable":
                    sz_cb = QComboBox(); sz_cb.addItems(["10 SQMM", "16 SQMM", "25 SQMM"]); sz_cb.setCurrentText(item.cable_size)
                    sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "cable_size", t)); self.editor_layout.addRow("PVC Size:", sz_cb)
                    
                aug_cb = QComboBox(); aug_cb.addItems(["New", "Replace 2W->4W", "Add-on 2W"]); aug_cb.setCurrentText(item.aug_type)
                aug_cb.currentTextChanged.connect(lambda t: self.update_span(item, "aug_type", t)); self.editor_layout.addRow("Work Nature:", aug_cb)

        del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
        del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)

    def update_pole(self, item, prop, value): 
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed) 

    def update_pole_extension(self, item, value):
        item.has_extension = value
        item.update_visuals()
        self.refresh_live_estimate()
        
    def update_span(self, item, prop, value): setattr(item, prop, value); item.update_visuals(); self.refresh_live_estimate()
    def update_dtr_logic(self, item, size): item.dtr_size = size; item.earth_count = 5 if size != "None" else 2; item.update_visuals(); self.refresh_live_estimate()
    def update_conductor_logic(self, item, conductor): item.conductor = conductor; item.update_visuals(); QTimer.singleShot(50, self.on_selection_changed); self.refresh_live_estimate()
    def delete_item(self, item):
        if not item or not item.scene(): return
        if hasattr(item, 'connected_spans'):
            for span in list(item.connected_spans):
                if span.label.scene(): self.scene.removeItem(span.label)
                if span.scene(): self.scene.removeItem(span)
                if span in getattr(span.p1, 'connected_spans', []): span.p1.connected_spans.remove(span)
                if span in getattr(span.p2, 'connected_spans', []): span.p2.connected_spans.remove(span)
        if isinstance(item, SmartSpan) and item.label.scene(): self.scene.removeItem(item.label)
        if item.scene(): self.scene.removeItem(item)
        self.refresh_live_estimate()

    # --- 5. EXACT MATH ENGINE ---
    def open_search(self, db_type):
        dialog = SearchDialog(db_type, self)
        if dialog.exec():
            selected = dialog.get_selected()
            if selected:
                name = selected['name']
                current_data = self.bom_overrides.get(name, {"qty": 0, "type": db_type})
                self.bom_overrides[name] = {"qty": current_data["qty"] + 1, "type": db_type}
                self.refresh_live_estimate()

    def on_table_edit(self, item):
        if item.column() == 3: 
            try:
                new_qty = float(item.text())
                name = self.live_table.item(item.row(), 2).text()
                row_type = self.live_table.item(item.row(), 0).text()
                self.bom_overrides[name] = {"qty": new_qty, "type": row_type}
                self.live_table.itemChanged.disconnect(self.on_table_edit)
                self.refresh_live_estimate()
                self.live_table.itemChanged.connect(self.on_table_edit)
            except ValueError: pass 

    def refresh_live_estimate(self):
        raw_bom = {}; total_lab_tasks = {}; use_uh = self.uh_checkbox.isChecked()
        d_boxes_poles = set() # Track poles that receive Distribution Boxes
        
        for item in self.scene.items():
            if isinstance(item, SmartSpan):
                p1_is_existing = getattr(item.p1, 'is_existing', False)
                p2_is_existing = getattr(item.p2, 'is_existing', False)
                if p1_is_existing and p2_is_existing:
                    continue
                
                length_km = item.length / 1000.0

                if item.is_service_drop:
                    pole = item.p1 if isinstance(item.p1, SmartPole) else item.p2
                    
                    if item.consider_cable:
                        if item.phase == "3 Phase": total_lab_tasks["Fixing of 3ph Service Connection (Cable provided)"] = total_lab_tasks.get("Fixing of 3ph Service Connection (Cable provided)", 0) + 1
                        else: total_lab_tasks["Fixing of 1ph Service Connection (Cable provided)"] = total_lab_tasks.get("Fixing of 1ph Service Connection (Cable provided)", 0) + 1
                        
                        if item.cable_size == "10 SQMM": raw_bom["CABLE (PVC 1.1 KV GRADE) 4Core X10 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4Core X10 sq mm", 0) + length_km
                        elif item.cable_size == "16 SQMM": raw_bom["CABLE (PVC 1.1 KV GRADE) 4CX16 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4CX16 sq mm", 0) + length_km
                        else: raw_bom["CABLE (PVC 1.1 KV GRADE) 4CX25 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4CX25 sq mm", 0) + length_km
                        total_lab_tasks["Laying & Dressing of 1.1 KV PVC/XLPE 2x10,4x10/16, 3.5/4x25 Sqmm Cable"] = total_lab_tasks.get("Laying & Dressing of 1.1 KV PVC/XLPE 2x10,4x10/16, 3.5/4x25 Sqmm Cable", 0) + length_km
                    else:
                        if item.phase == "3 Phase": total_lab_tasks["Fixing of 3ph Service Connection (No Cable)"] = total_lab_tasks.get("Fixing of 3ph Service Connection (No Cable)", 0) + 1
                        else: total_lab_tasks["Fixing of 1ph Service Connection (No Cable)"] = total_lab_tasks.get("Fixing of 1ph Service Connection (No Cable)", 0) + 1

                    if pole not in d_boxes_poles:
                        has_ab_cable = any(getattr(s, 'conductor', '') == "AB Cable" for s in pole.connected_spans)
                        if has_ab_cable:
                            raw_bom["LT Distribution Box along with steel Strap & Buckle for 3ph connection in ABC system"] = raw_bom.get("LT Distribution Box along with steel Strap & Buckle for 3ph connection in ABC system", 0) + 1
                            total_lab_tasks["Erection of distribution box"] = total_lab_tasks.get("Erection of distribution box", 0) + 1
                            d_boxes_poles.add(pole)

                else:
                    if item.length > 60: total_lab_tasks["Lead Wire above above 60 Mtrs (2 Wire)"] = total_lab_tasks.get("Lead Wire above above 60 Mtrs (2 Wire)", 0) + 1
                    if item.has_cg:
                        raw_bom["G.I. 8 SWG Wire (4mm)"] = raw_bom.get("G.I. 8 SWG Wire (4mm)", 0) + (item.length * 0.0001)
                        total_lab_tasks["Fixing Cross lacing"] = total_lab_tasks.get("Fixing Cross lacing", 0) + int(item.length / 2)

                    is_lt_span = (getattr(item.p1, 'pole_type', '') == "LT") or (getattr(item.p2, 'pole_type', '') == "LT")

                    if item.conductor == "PVC Cable":
                        if item.cable_size == "10 SQMM": raw_bom["CABLE (PVC 1.1 KV GRADE) 4Core X10 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4Core X10 sq mm", 0) + length_km
                        elif item.cable_size == "16 SQMM": raw_bom["CABLE (PVC 1.1 KV GRADE) 4CX16 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4CX16 sq mm", 0) + length_km
                        else: raw_bom["CABLE (PVC 1.1 KV GRADE) 4CX25 sq mm"] = raw_bom.get("CABLE (PVC 1.1 KV GRADE) 4CX25 sq mm", 0) + length_km
                        total_lab_tasks["Laying & Dressing of 1.1 KV PVC/XLPE 2x10,4x10/16, 3.5/4x25 Sqmm Cable"] = total_lab_tasks.get("Laying & Dressing of 1.1 KV PVC/XLPE 2x10,4x10/16, 3.5/4x25 Sqmm Cable", 0) + length_km
                        total_lab_tasks["Survey for L.T.O.H Line"] = total_lab_tasks.get("Survey for L.T.O.H Line", 0) + length_km 
                        
                    elif item.conductor == "ACSR":
                        wire_multiplier = int(item.wire_count)
                        total_wire_km = length_km * wire_multiplier
                        if item.aug_type in ["New", "Add-on 2W", "Replace 2W->4W"]:
                            raw_bom[f"ACSR Conductor {item.wire_size}"] = raw_bom.get(f"ACSR Conductor {item.wire_size}", 0) + total_wire_km
                            total_lab_tasks[f"Stringing & Sagging with 50 sq.mm A.C.S.R. {item.wire_count} Wire"] = total_lab_tasks.get(f"Stringing & Sagging with 50 sq.mm A.C.S.R. {item.wire_count} Wire", 0) + length_km 
                        
                        if not is_lt_span: total_lab_tasks["Survey for H.T.O.H Line"] = total_lab_tasks.get("Survey for H.T.O.H Line", 0) + length_km 
                        else:
                            total_lab_tasks["Survey for L.T.O.H Line"] = total_lab_tasks.get("Survey for L.T.O.H Line", 0) + length_km 
                            raw_bom["LT Spacer 3 PHASE 4 WIRE"] = raw_bom.get("LT Spacer 3 PHASE 4 WIRE", 0) + 4
                            total_lab_tasks["Fixing of LT spacer"] = total_lab_tasks.get("Fixing of LT spacer", 0) + 4
                            if use_uh:
                                raw_bom["UH-LT BKT 4 WAY"] = raw_bom.get("UH-LT BKT 4 WAY", 0) + 1
                                raw_bom["UH-CLAMP FOR 8 MTR PCC POLE"] = raw_bom.get("UH-CLAMP FOR 8 MTR PCC POLE", 0) + 4
                                raw_bom["UH-Diron Clump"] = raw_bom.get("UH-Diron Clump", 0) + wire_multiplier
                                raw_bom["Shakle Insulator"] = raw_bom.get("Shakle Insulator", 0) + wire_multiplier
                                total_lab_tasks["Fixing of LT Bracket(Without Painted)"] = total_lab_tasks.get("Fixing of LT Bracket(Without Painted)", 0) + 1
                                total_lab_tasks["Fixing of LT Shackle Insulator (with N/B)"] = total_lab_tasks.get("Fixing of LT Shackle Insulator (with N/B)", 0) + wire_multiplier
                            else:
                                raw_bom["M.S Angle 65X65X6mm"] = raw_bom.get("M.S Angle 65X65X6mm", 0) + (1.0 * 6.5 / 1000)
                                raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (1.0 * 3.5 / 1000)
                                raw_bom["Shakle Insulator"] = raw_bom.get("Shakle Insulator", 0) + wire_multiplier
                                total_lab_tasks["Fixing of LT Shackle Insulator (with N/B)"] = total_lab_tasks.get("Fixing of LT Shackle Insulator (with N/B)", 0) + wire_multiplier
                    
                    elif item.conductor == "AB Cable":
                        if item.aug_type in ["New", "Add-on 2W", "Replace 2W->4W"]:
                            raw_bom["LT AB CABLE 1.1KV 3CX50+1CX16+1CX35sqmm"] = raw_bom.get("LT AB CABLE 1.1KV 3CX50+1CX16+1CX35sqmm", 0) + length_km
                            total_lab_tasks["Stringing & Sagging of LT AB Cable"] = total_lab_tasks.get("Stringing & Sagging of LT AB Cable", 0) + length_km 
                        total_lab_tasks["Survey for L.T.O.H Line"] = total_lab_tasks.get("Survey for L.T.O.H Line", 0) + length_km 
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (0.5 * 3.5 / 1000)
                        raw_bom["Suspension Clamp Suitable for 35sq.mm. Messenger Conductor"] = raw_bom.get("Suspension Clamp Suitable for 35sq.mm. Messenger Conductor", 0) + 1
                        total_lab_tasks["Erection of Anchoring/Suspension Clamp"] = total_lab_tasks.get("Erection of Anchoring/Suspension Clamp", 0) + 1

            elif isinstance(item, SmartPole):
                if item.is_existing:
                    if item.stay_count > 0:
                        if getattr(item, 'stay_type', 'HT') == "LT":
                            raw_bom["LT Stay set"] = raw_bom.get("LT Stay set", 0) + item.stay_count
                            raw_bom["GI STAY WIRE 7/12 SWG"] = raw_bom.get("GI STAY WIRE 7/12 SWG", 0) + (item.stay_count * 0.004)
                            raw_bom["LT Guy Insulator"] = raw_bom.get("LT Guy Insulator", 0) + item.stay_count
                            total_lab_tasks["LT Stay set complete"] = total_lab_tasks.get("LT Stay set complete", 0) + item.stay_count
                        else:
                            raw_bom["H.T. Stay Set Complete"] = raw_bom.get("H.T. Stay Set Complete", 0) + item.stay_count
                            raw_bom["G.I. Stay Wire 7/3..15MM 10 SWG(HT)"] = raw_bom.get("G.I. Stay Wire 7/3..15MM 10 SWG(HT)", 0) + (item.stay_count * 0.006)
                            raw_bom["H.T. Guy Insulator 11KV"] = raw_bom.get("H.T. Guy Insulator 11KV", 0) + item.stay_count
                            total_lab_tasks["H.T. Stay Set Complete Labor"] = total_lab_tasks.get("H.T. Stay Set Complete Labor", 0) + item.stay_count
                    has_cg = any(getattr(s, 'has_cg', False) for s in item.connected_spans)
                    if has_cg:
                        raw_bom["M.S Angle 65X65X6mm"] = raw_bom.get("M.S Angle 65X65X6mm", 0) + (1.9 * 6.5 / 1000)
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (0.5 * 3.5 / 1000)
                        total_lab_tasks["Fabrication & Fixing  of Cattle Guard Bracket (SP)"] = total_lab_tasks.get("Fabrication & Fixing  of Cattle Guard Bracket (SP)", 0) + 1
                    continue

                p_type = f"P C C POLE:{item.height[:1]} Mtrs.Long"
                pole_count = 2 if item.pole_type == "DTR" else 1
                raw_bom[p_type] = raw_bom.get(p_type, 0) + pole_count
                
                if item.pole_type != "DTR":
                    lab_pole_key = "Erection of . 8mtr  PCC Pole ( LT)" if item.pole_type=="LT" else f"Erection of . {item.height[:1]}mtr  PCC Pole (HT) Without Painted"
                    total_lab_tasks[lab_pole_key] = total_lab_tasks.get(lab_pole_key, 0) + 1
                
                total_lab_tasks["Pole GIS survey"] = total_lab_tasks.get("Pole GIS survey", 0) + pole_count
                boards = 0 if item.pole_type == "LT" else (2 if item.pole_type == "DTR" else 1)
                if boards > 0:
                    raw_bom["Caution Board-11KVA"] = raw_bom.get("Caution Board-11KVA", 0) + boards
                    total_lab_tasks["Fixing of Caution Board"] = total_lab_tasks.get("Fixing of Caution Board", 0) + boards

                if item.height == "8MTR" and item.pole_type in ["HT", "DTR"]:
                    raw_bom["M.S Channel 75X40 mm"] = raw_bom.get("M.S Channel 75X40 mm", 0) + (6.0 * 7.5 / 1000) * pole_count
                    raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (3.0 * 3.5 / 1000) * pole_count
                    total_lab_tasks["Extension of 8 mtr PCC Pole (Without Painted)HT"] = total_lab_tasks.get("Extension of 8 mtr PCC Pole (Without Painted)HT", 0) + pole_count

                if item.earth_count > 0:
                    raw_bom["G.I. Earth Spike 6*3.25ft"] = raw_bom.get("G.I. Earth Spike 6*3.25ft", 0) + item.earth_count
                    if item.pole_type == "LT": raw_bom["G.I. 8 SWG Wire (4mm)"] = raw_bom.get("G.I. 8 SWG Wire (4mm)", 0) + (item.earth_count * 0.003)
                    else: raw_bom["G.I. 6 SWG Wire (5mm)"] = raw_bom.get("G.I. 6 SWG Wire (5mm)", 0) + (item.earth_count * 0.003)
                    total_lab_tasks["Earthing Complete"] = total_lab_tasks.get("Earthing Complete", 0) + item.earth_count
                
                if item.stay_count > 0:
                    if item.pole_type == "LT":
                        raw_bom["LT Stay set"] = raw_bom.get("LT Stay set", 0) + item.stay_count
                        raw_bom["GI STAY WIRE 7/12 SWG"] = raw_bom.get("GI STAY WIRE 7/12 SWG", 0) + (item.stay_count * 0.004)
                        raw_bom["LT Guy Insulator"] = raw_bom.get("LT Guy Insulator", 0) + item.stay_count
                        total_lab_tasks["LT Stay set complete"] = total_lab_tasks.get("LT Stay set complete", 0) + item.stay_count
                    else:
                        raw_bom["H.T. Stay Set Complete"] = raw_bom.get("H.T. Stay Set Complete", 0) + item.stay_count
                        raw_bom["G.I. Stay Wire 7/3..15MM 10 SWG(HT)"] = raw_bom.get("G.I. Stay Wire 7/3..15MM 10 SWG(HT)", 0) + (item.stay_count * 0.006)
                        raw_bom["H.T. Guy Insulator 11KV"] = raw_bom.get("H.T. Guy Insulator 11KV", 0) + item.stay_count
                        total_lab_tasks["H.T. Stay Set Complete Labor"] = total_lab_tasks.get("H.T. Stay Set Complete Labor", 0) + item.stay_count

                has_cg = any(getattr(s, 'has_cg', False) for s in item.connected_spans)
                if has_cg:
                    if item.pole_type == "DTR":
                        raw_bom["M.S Angle 65X65X6mm"] = raw_bom.get("M.S Angle 65X65X6mm", 0) + (2.75 * 6.5 / 1000)
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (0.5 * 3.5 / 1000)
                        total_lab_tasks["Fabrication & Fixing  of Cattle Guard Bracket (DP)"] = total_lab_tasks.get("Fabrication & Fixing  of Cattle Guard Bracket (DP)", 0) + 1
                    else:
                        raw_bom["M.S Angle 65X65X6mm"] = raw_bom.get("M.S Angle 65X65X6mm", 0) + (1.9 * 6.5 / 1000)
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (0.5 * 3.5 / 1000)
                        total_lab_tasks["Fabrication & Fixing  of Cattle Guard Bracket (SP)"] = total_lab_tasks.get("Fabrication & Fixing  of Cattle Guard Bracket (SP)", 0) + 1

                if item.pole_type == "HT":
                    ht_spans = [s for s in item.connected_spans if s.conductor == "ACSR" and getattr(s.p1, 'pole_type', '') != "LT" and getattr(s.p2, 'pole_type', '') != "LT"]
                    if len(ht_spans) <= 2: 
                        raw_bom["11 KV Polymer Disc Insulator 45KN"] = raw_bom.get("11 KV Polymer Disc Insulator 45KN", 0) + 3 
                        raw_bom["Hardware fittings 11KV"] = raw_bom.get("Hardware fittings 11KV", 0) + 3
                        total_lab_tasks["Fixing of 11 KV Disc Insulator"] = total_lab_tasks.get("Fixing of 11 KV Disc Insulator", 0) + 3
                        raw_bom["M.S Channel 75X40 mm"] = raw_bom.get("M.S Channel 75X40 mm", 0) + (1.8 * 7.5 / 1000) 
                    elif len(ht_spans) >= 3: 
                        raw_bom["11 KV Polymer Disc Insulator 45KN"] = raw_bom.get("11 KV Polymer Disc Insulator 45KN", 0) + 3 
                        raw_bom["Hardware fittings 11KV"] = raw_bom.get("Hardware fittings 11KV", 0) + 3
                        total_lab_tasks["Fixing of 11 KV Disc Insulator"] = total_lab_tasks.get("Fixing of 11 KV Disc Insulator", 0) + 3
                        raw_bom["M.S Channel 75X40 mm"] = raw_bom.get("M.S Channel 75X40 mm", 0) + (2.0 * 7.5 / 1000) 
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (1.5 * 3.5 / 1000)
                        total_lab_tasks["Fixing of Solid Tee-off Bracket on S.P"] = total_lab_tasks.get("Fixing of Solid Tee-off Bracket on S.P", 0) + 1
                
                if item.pole_type == "DTR":
                    if item.dtr_size != "None":
                        raw_bom[f"Dist. Transformer {item.dtr_size}"] = raw_bom.get(f"Dist. Transformer {item.dtr_size}", 0) + 1
                        total_lab_tasks["Erection of 25 KVA Transformer"] = total_lab_tasks.get("Erection of 25 KVA Transformer", 0) + 1
                        total_lab_tasks["DTR Code Painting"] = total_lab_tasks.get("DTR Code Painting", 0) + 1
                        
                        if item.height == "8MTR": total_lab_tasks["Erection of S/S D.P. Structure  (8 mtr without Painted)"] = total_lab_tasks.get("Erection of S/S D.P. Structure  (8 mtr without Painted)", 0) + 1
                        else: total_lab_tasks["Sub-Stationn Str with 9 Mtr PCC pole DP Without Painted"] = total_lab_tasks.get("Sub-Stationn Str with 9 Mtr PCC pole DP Without Painted", 0) + 1

                        raw_bom["M.S Channel 75X40 mm"] = raw_bom.get("M.S Channel 75X40 mm", 0) + (14.5 * 7.5 / 1000)
                        raw_bom["M.S Angle 65X65X6mm"] = raw_bom.get("M.S Angle 65X65X6mm", 0) + (12.25 * 6.5 / 1000)
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (14.0 * 3.5 / 1000)
                        
                        raw_bom["11 KV Polymer Disc Insulator 45KN"] = raw_bom.get("11 KV Polymer Disc Insulator 45KN", 0) + 3 
                        raw_bom["Hardware fittings 11KV"] = raw_bom.get("Hardware fittings 11KV", 0) + 3
                        total_lab_tasks["Fixing of 11 KV Disc Insulator"] = total_lab_tasks.get("Fixing of 11 KV Disc Insulator", 0) + 3
                        raw_bom["11 KV Polymer Pin Insulator 45KN"] = raw_bom.get("11 KV Polymer Pin Insulator 45KN", 0) + 9 
                        total_lab_tasks["Fixing of 11 KV Pin Insulator"] = total_lab_tasks.get("Fixing of 11 KV Pin Insulator", 0) + 9
                        
                        raw_bom["ACSR Conductor 50SQMM"] = raw_bom.get("ACSR Conductor 50SQMM", 0) + 0.030 
                        raw_bom["Lightning Arrestor 12 KV"] = raw_bom.get("Lightning Arrestor 12 KV", 0) + 3
                        total_lab_tasks["Fixing of 11 KV Lightning Arrestor"] = total_lab_tasks.get("Fixing of 11 KV Lightning Arrestor", 0) + 1
                        raw_bom["T.P.G.O. Isolator (200Amps) 11KV"] = raw_bom.get("T.P.G.O. Isolator (200Amps) 11KV", 0) + 1
                        total_lab_tasks["Fixing of  11 KV TGPO Isolator on S/Stn Structure"] = total_lab_tasks.get("Fixing of  11 KV TGPO Isolator on S/Stn Structure", 0) + 1
                        raw_bom["G.I. Turn Buckle"] = raw_bom.get("G.I. Turn Buckle", 0) + 2
                        total_lab_tasks["Fixing of neutral earthing of DTR WITH G"] = total_lab_tasks.get("Fixing of neutral earthing of DTR WITH G", 0) + 1
                        raw_bom["PVC Cable 4 Core 25SQMM"] = raw_bom.get("PVC Cable 4 Core 25SQMM", 0) + 0.010
                        raw_bom["LT Distribution KIOSK FOR 25 KVA DTR"] = raw_bom.get("LT Distribution KIOSK FOR 25 KVA DTR", 0) + 1
                        total_lab_tasks["FIXING OF LT Distribution KIOSK FOR 25 KVA DTR"] = total_lab_tasks.get("FIXING OF LT Distribution KIOSK FOR 25 KVA DTR", 0) + 1
                    else:
                        if item.height == "8MTR": total_lab_tasks["Erection of 8 mtr D.P structure (HT)"] = total_lab_tasks.get("Erection of 8 mtr D.P structure (HT)", 0) + 1
                        else: total_lab_tasks["Erection of 9 MTR Long PCC pole D/P for HTOH line"] = total_lab_tasks.get("Erection of 9 MTR Long PCC pole D/P for HTOH line", 0) + 1

                        raw_bom["M.S Channel 75X40 mm"] = raw_bom.get("M.S Channel 75X40 mm", 0) + (5.0 * 7.5 / 1000)
                        raw_bom["M.S Flat 65X6 mm"] = raw_bom.get("M.S Flat 65X6 mm", 0) + (3.0 * 3.5 / 1000)
                        raw_bom["11 KV Polymer Disc Insulator 45KN"] = raw_bom.get("11 KV Polymer Disc Insulator 45KN", 0) + 3
                        raw_bom["Hardware fittings 11KV"] = raw_bom.get("Hardware fittings 11KV", 0) + 3
                        total_lab_tasks["Fixing of 11 KV Disc Insulator"] = total_lab_tasks.get("Fixing of 11 KV Disc Insulator", 0) + 3
                        raw_bom["11 KV Polymer Pin Insulator 45KN"] = raw_bom.get("11 KV Polymer Pin Insulator 45KN", 0) + 3 
                        total_lab_tasks["Fixing of 11 KV Pin Insulator"] = total_lab_tasks.get("Fixing of 11 KV Pin Insulator", 0) + 3

        conn = sqlite3.connect('erp_master.db'); cursor = conn.cursor()
        self.live_bom_data.clear()
        
        all_materials = list(set(raw_bom.keys()) | set([k for k,v in self.bom_overrides.items() if v['type'] == "Material"]))
        for name in all_materials:
            cursor.execute("SELECT item_code, unit, rate FROM materials WHERE item_name=?", (name,)); res = cursor.fetchone()
            if res:
                if name in self.bom_overrides: qty = float(self.bom_overrides[name]['qty'])
                else: 
                    qty = raw_bom.get(name, 0)
                    if res[1].upper() in ['MT', 'KM']: qty *= 1.03 
                if qty > 0: self.live_bom_data.append({"type": "Material", "code": res[0], "name": name, "qty": qty, "unit": res[1], "rate": res[2], "amt": qty * res[2]})
                
        all_labor = list(set(total_lab_tasks.keys()) | set([k for k,v in self.bom_overrides.items() if v['type'] == "Labor"]))
        for name in all_labor:
            cursor.execute("SELECT unit, rate FROM labor WHERE task_name=?", (name,)); res = cursor.fetchone()
            if res:
                if name in self.bom_overrides: qty = float(self.bom_overrides[name]['qty'])
                else: qty = total_lab_tasks.get(name, 0)
                if qty > 0: self.live_bom_data.append({"type": "Labor", "code": "", "name": name, "qty": qty, "unit": res[0], "rate": res[1], "amt": qty * res[1]})
        conn.close()

        self.live_table.itemChanged.disconnect(self.on_table_edit) 
        self.live_table.setRowCount(0)
        
        for i, item in enumerate(self.live_bom_data):
            self.live_table.insertRow(i); self.live_table.setItem(i, 0, QTableWidgetItem(item['type'])); self.live_table.setItem(i, 1, QTableWidgetItem(item['code'])); self.live_table.setItem(i, 2, QTableWidgetItem(item['name']))
            qty_item = QTableWidgetItem(f"{item['qty']:.3f}"); qty_item.setBackground(QColor("#fff3cd")); self.live_table.setItem(i, 3, qty_item)
            self.live_table.setItem(i, 4, QTableWidgetItem(item['unit'])); self.live_table.setItem(i, 5, QTableWidgetItem(f"{item['amt']:.2f}"))
            for col in [0, 1, 2, 4, 5]:
                t_item = self.live_table.item(i, col)
                t_item.setFlags(t_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self.live_table.itemChanged.connect(self.on_table_edit)
        
        mat_base = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Material'])
        current_date = datetime.now()
        current_fy_start = current_date.year if current_date.month >= 4 else current_date.year - 1
        
        self.escalations = []
        current_mat_val = mat_base
        for year in range(2024, current_fy_start + 1):
            esc_amt = current_mat_val * 0.05
            self.escalations.append((f"{str(year)[-2:]}-{str(year+1)[-2:]}", esc_amt))
            current_mat_val += esc_amt
            
        sun = current_mat_val * 0.05
        mat_sub = current_mat_val + sun
        
        lab_sub = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Labor'])
        sup = (mat_sub + lab_sub) * 0.10; gst = (lab_sub + sup) * 0.18; final_amt = (mat_sub + lab_sub + sup + gst) * 1.01
        self.grand_total_label.setText(f"<b>Estimated Cost (Inc Taxes): Rs. {final_amt:,.2f}</b>")

    def generate_excel(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Export ERP Estimate", "ERP_Estimate.xlsx", "Excel Files (*.xlsx)")
        if not filename: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estimate"
        ws.merge_cells('A1:G1'); ws['A1'] = "AUTOMATED ERP ESTIMATE"; ws['A1'].font = Font(bold=True, size=14, color="FFFFFF"); ws['A1'].fill = PatternFill("solid", fgColor="4F81BD"); ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2'); ws['A2'] = f"Subject: {self.subject_input.text()} | Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.append(["Sl No.", "Material Code", "Description", "Qty", "Unit", "Rate", "Amount"])
        for col_num, cell in enumerate(ws[3], 1): cell.font = Font(bold=True)
        ws.column_dimensions['C'].width = 45; ws.column_dimensions['B'].width = 15
        
        total_lab = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Labor'])

        row = 4; ws.cell(row, 3, "A. MATERIALS").font = Font(bold=True); row += 1
        for i, item in enumerate([x for x in self.live_bom_data if x['type'] == 'Material'], 1):
            ws.append([i, item['code'], item['name'], round(item['qty'],3), item['unit'], item['rate'], round(item['amt'],2)]); row += 1
        
        mat_base = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Material'])
        ws.append(["", "", "Material Base Total", "", "", "", round(mat_base, 2)]); row += 1
        
        current_mat_val = mat_base
        for fy, esc_amt in self.escalations:
            ws.append(["", "", f"Add: Escalation @ 5% for FY {fy}", "", "", "", round(esc_amt, 2)]); row += 1
            current_mat_val += esc_amt
            
        sun = current_mat_val * 0.05
        mat_subtotal = current_mat_val + sun
        
        ws.append(["", "", "Add: Sundries @ 5%", "", "", "", round(sun, 2)]); row += 1
        ws.append(["", "", "TOTAL MATERIAL COST (A)", "", "", "", round(mat_subtotal, 2)])
        ws.cell(row, 3).font = Font(bold=True); ws.cell(row, 7).font = Font(bold=True); row += 2
        
        ws.cell(row, 3, "B. ERECTION / LABOR").font = Font(bold=True); row += 1
        for i, item in enumerate([x for x in self.live_bom_data if x['type'] == 'Labor'], 1):
            ws.append([i, "", item['name'], round(item['qty'],3), item['unit'], item['rate'], round(item['amt'],2)]); row += 1
            
        ws.append(["", "", "TOTAL LABOR COST (B)", "", "", "", round(total_lab, 2)])
        ws.cell(row, 3).font = Font(bold=True); ws.cell(row, 7).font = Font(bold=True); row += 2

        sup = (mat_subtotal + total_lab) * 0.10; gst = (total_lab + sup) * 0.18; sub_c = mat_subtotal + total_lab + sup + gst; g_tot = sub_c * 1.01
        ws.append(["", "", "C. OVERHEADS & TAXES"]); ws.cell(row, 3).font = Font(bold=True); row += 1
        ws.append(["", "", "Supervision @ 10% on (A+B)", "", "", "", round(sup, 2)]); row += 1
        ws.append(["", "", "GST @ 18% on (Labor + Sup)", "", "", "", round(gst, 2)]); row += 1
        ws.append(["", "", "Sub-Total", "", "", "", round(sub_c, 2)]); row += 1
        ws.append(["", "", "Add: Cess @ 1%", "", "", "", round(sub_c * 0.01, 2)]); row += 1
        ws.append(["", "", "GRAND TOTAL", "", "", "", round(g_tot, 2)])
        ws.cell(row, 3).font = Font(bold=True, size=12); ws.cell(row, 7).font = Font(bold=True, size=12, color="FF0000")
        
        wb.save(filename); QMessageBox.information(self, "Success", f"ERP Estimate Excel saved to:\n{filename}")

    def export_pdf(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Export PDF Drawing", "Project_Drawing.pdf", "PDF Files (*.pdf)")
        if not filename: return

        printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)

        source_rect = self.scene.itemsBoundingRect()
        if source_rect.isNull():
            QMessageBox.warning(self, "Empty", "Canvas is empty.")
            return

        if source_rect.width() > source_rect.height():
            printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        else:
            printer.setPageOrientation(QPageLayout.Orientation.Portrait)

        painter = QPainter(printer)
        page_rect_px = printer.pageRect(QPrinter.Unit.DevicePixel)
        margin_px = 10

        border_rect = page_rect_px.adjusted(margin_px, margin_px, -margin_px, -margin_px)

        painter.setPen(Qt.GlobalColor.black)
        title_font = QFont("Arial", 12, QFont.Weight.Bold); title_font.setUnderline(True)
        painter.setFont(title_font)
        title_text = self.subject_input.text() or 'ERP PROJECT DRAWING'
        text_flags = Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop | Qt.TextFlag.TextWordWrap
        title_rect_calc = QRectF(border_rect.x() + 5, border_rect.y(), border_rect.width() - 10, 9999)
        required_rect = painter.boundingRect(title_rect_calc, text_flags, title_text)
        title_height = required_rect.height()
        title_rect = QRectF(border_rect.x(), border_rect.y(), border_rect.width(), title_height)
        painter.drawText(title_rect, text_flags, title_text)
        
        scene_target_rect = QRectF(border_rect)
        scene_target_rect.setTop(border_rect.top() + title_height + 10)
        source_rect.adjust(-50, -50, 50, 50)
        self.scene.render(painter, scene_target_rect, source_rect, Qt.AspectRatioMode.KeepAspectRatio)
        
        painter.setFont(QFont("Arial", 8)); painter.setPen(Qt.GlobalColor.black)

        legend_data = {
            'New Poles': {'LT Pole': {'s': '🔵', 'q': 0}, 'HT Pole': {'s': '🔴', 'q': 0}, 'DP/DTR': {'s': '🟩', 'q': 0}},
            'Existing Poles': {'Ex. Pole': {'s': '⚪', 'q': 0}, 'Ex. DP/DTR': {'s': '⚪', 'q': 0}},
            'Ancillary': {'Consumer Home': {'s': '🏠', 'q': 0}, 'Earth': {'s': '⏚', 'q': 0}, 'Stay': {'s': 'S', 'q': 0}},
            'New Spans': {'New ACSR': {'s': '---', 'l': 0}, 'New ABC': {'s': '~--~', 'l': 0}, 'New PVC': {'s': '~~~', 'l': 0}},
            'Existing Spans': {'Ex. ACSR': {'s': '———', 'l': 0}, 'Ex. ABC': {'s': '~~~~~', 'l': 0}, 'Ex. PVC': {'s': '~~~~~', 'l': 0}}
        }
        
        for item in self.scene.items():
            if isinstance(item, SmartPole):
                legend_data['Ancillary']['Earth']['q'] += item.earth_count
                legend_data['Ancillary']['Stay']['q'] += item.stay_count
                if item.is_existing:
                    if item.pole_type == 'DTR': legend_data['Existing Poles']['Ex. DP/DTR']['q'] += 1
                    else: legend_data['Existing Poles']['Ex. Pole']['q'] += 1
                else:
                    if item.pole_type == 'LT': legend_data['New Poles']['LT Pole']['q'] += 1
                    elif item.pole_type == 'HT': legend_data['New Poles']['HT Pole']['q'] += 1
                    elif item.pole_type == 'DTR': legend_data['New Poles']['DP/DTR']['q'] += 1
            elif isinstance(item, SmartHome):
                legend_data['Ancillary']['Consumer Home']['q'] += 1
            elif isinstance(item, SmartSpan):
                span_len = item.length
                if item.is_existing_span:
                    if item.conductor == 'ACSR': legend_data['Existing Spans']['Ex. ACSR']['l'] += span_len
                    elif item.conductor == 'AB Cable': legend_data['Existing Spans']['Ex. ABC']['l'] += span_len
                    elif item.conductor == 'PVC Cable': legend_data['Existing Spans']['Ex. PVC']['l'] += span_len
                else:
                    if item.conductor == 'ACSR': legend_data['New Spans']['New ACSR']['l'] += span_len
                    elif item.conductor == 'AB Cable': legend_data['New Spans']['New ABC']['l'] += span_len
                    elif item.conductor == 'PVC Cable': legend_data['New Spans']['New PVC']['l'] += span_len

        used_items = []
        for category, items in legend_data.items():
            for desc, data in items.items():
                qty = data.get('q')
                length = data.get('l')
                if (qty and qty > 0) or (length and length > 0):
                    val = str(qty) if qty is not None else f"{length}m"
                    used_items.append({'desc': desc, 'sym': data['s'], 'val': val})

        if used_items:
            col_widths = {'sl': 30, 'sym': 50, 'desc': 130, 'qty': 50}
            table_width = sum(col_widths.values())
            row_height = 18
            
            table_height = (len(used_items) + 1) * row_height
            latlong_box_height = 25
            total_legend_height = table_height + latlong_box_height
            
            legend_block_rect = QRectF(border_rect.left() + 5, border_rect.bottom() - total_legend_height - 5, table_width, total_legend_height)
            
            painter.setBrush(QBrush(QColor(255, 255, 255, 220)))
            painter.setPen(QPen(QColor(200, 200, 200), 0.5))
            painter.drawRect(legend_block_rect)
            painter.setPen(QPen(Qt.GlobalColor.black))

            current_y = legend_block_rect.top()
            
            # Header
            painter.setFont(QFont("Arial", 8, QFont.Weight.Bold))
            current_x = legend_block_rect.left()
            headers = {'sl': ' Sl No. ', 'sym': ' Symbol ', 'desc': ' Description ', 'qty': ' Qty/Len '}
            for key, width in col_widths.items():
                painter.drawText(QRectF(current_x, current_y, width, row_height), Qt.AlignmentFlag.AlignCenter, headers[key])
                current_x += width
            
            current_y += row_height
            
            # Rows
            for i, item in enumerate(used_items):
                sl_no = str(i + 1)
                current_x = legend_block_rect.left()
                painter.setFont(QFont("Arial", 8))

                # Draw cell content
                painter.drawText(QRectF(current_x, current_y, col_widths['sl'], row_height), Qt.AlignmentFlag.AlignCenter, sl_no)
                current_x += col_widths['sl']
                painter.drawText(QRectF(current_x, current_y, col_widths['sym'], row_height), Qt.AlignmentFlag.AlignCenter, item['sym'])
                current_x += col_widths['sym']
                painter.drawText(QRectF(current_x + 5, current_y, col_widths['desc'] - 5, row_height), Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, item['desc'])
                current_x += col_widths['desc']
                painter.drawText(QRectF(current_x, current_y, col_widths['qty'], row_height), Qt.AlignmentFlag.AlignCenter, item['val'])
                
                current_y += row_height

            # Draw Grid Lines
            painter.setPen(QPen(QColor(220, 220, 220), 1))
            # -- Horizontal lines
            for i in range(len(used_items) + 2):
                 y = legend_block_rect.top() + (i * row_height)
                 painter.drawLine(int(legend_block_rect.left()), int(y), int(legend_block_rect.right()), int(y))
            # -- Vertical lines
            current_x = legend_block_rect.left()
            for width in col_widths.values():
                current_x += width
                painter.drawLine(int(current_x), int(legend_block_rect.top()), int(current_x), int(current_y - row_height))

            painter.setPen(QPen(Qt.GlobalColor.black))

            # Draw Lat/Long at the bottom
            painter.setFont(QFont("Arial", 7, QFont.Weight.Normal, italic=True))
            lat_long_text = f"Lat: {self.lat_input.text()}   Long: {self.long_input.text()}"
            lat_long_rect = QRectF(legend_block_rect.left(), current_y, table_width, latlong_box_height)
            painter.drawText(lat_long_rect, Qt.AlignmentFlag.AlignCenter, lat_long_text)
        
        painter.end()
        QMessageBox.information(self, "Success", f"PDF Drawing exported to:\n{filename}")

    def new_drawing(self):
        if QMessageBox.question(self, 'New Canvas', 'Clear canvas?', QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self.scene.clear(); self.subject_input.clear(); self.span_start_pole = None; self.uh_checkbox.setChecked(False); self.bom_overrides.clear()
            self.refresh_live_estimate()
            
    def compile_save_data(self):
        state = {'subject': self.subject_input.text(), 'lat': self.lat_input.text(), 'long': self.long_input.text(), 'uh_toggle': self.uh_checkbox.isChecked(), 'overrides': self.bom_overrides, 'nodes': [], 'spans': []}; node_map = {}
        for i, item in enumerate(self.scene.items()):
            if isinstance(item, (SmartPole, SmartHome)):
                item._temp_id = i; node_map[i] = item
                node_data = {'id': i, 'type': 'Pole' if isinstance(item, SmartPole) else 'Home', 'x': item.x(), 'y': item.y(), 'label_x': item.label.pos().x(), 'label_y': item.label.pos().y(), 'label_text': item.label.toPlainText()}
                if isinstance(item, SmartPole): node_data.update({'pole_type': item.pole_type, 'is_existing': item.is_existing, 'height': item.height, 'dtr_size': item.dtr_size, 'earth_count': item.earth_count, 'stay_count': item.stay_count, 'stay_type': getattr(item, 'stay_type', 'HT')})
                state['nodes'].append(node_data)
        for item in self.scene.items():
            if isinstance(item, SmartSpan): state['spans'].append({'p1_id': item.p1._temp_id, 'p2_id': item.p2._temp_id, 'length': item.length, 'conductor': item.conductor, 'has_cg': item.has_cg, 'aug_type': item.aug_type, 'wire_count': item.wire_count, 'wire_size': item.wire_size, 'cable_size': getattr(item, 'cable_size', '10 SQMM'), 'consider_cable': getattr(item, 'consider_cable', False), 'phase': getattr(item, 'phase', '3 Phase'), 'is_service_drop': getattr(item, 'is_service_drop', False), 'label_x': item.label.pos().x(), 'label_y': item.label.pos().y(), 'label_text': item.label.toPlainText()})
        return state

    def parse_load_data(self, state):
        self.scene.clear()
        self.subject_input.setText(state.get('subject', ''))
        self.lat_input.setText(state.get('lat', ''))
        self.long_input.setText(state.get('long', ''))
        self.uh_checkbox.setChecked(state.get('uh_toggle', False)); self.bom_overrides = state.get('overrides', {}); node_map = {}
        for n_data in state.get('nodes', []):
            if n_data['type'] == 'Pole':
                pole = SmartPole(n_data['x'], n_data['y'], n_data['pole_type'], n_data.get('is_existing', False)); pole.height = n_data['height']; pole.dtr_size = n_data['dtr_size']; pole.earth_count = n_data['earth_count']; pole.stay_count = n_data['stay_count']; pole.stay_type = n_data.get('stay_type', 'HT'); pole.update_visuals(); pole.label.setPos(n_data['label_x'], n_data['label_y']); pole.label.setPlainText(n_data['label_text']); self.scene.addItem(pole); node_map[n_data['id']] = pole
            else:
                home = SmartHome(n_data['x'], n_data['y']); home.label.setPos(n_data['label_x'], n_data['label_y']); home.label.setPlainText(n_data['label_text']); self.scene.addItem(home); node_map[n_data['id']] = home
        for s_data in state.get('spans', []):
            p1 = node_map.get(s_data['p1_id']); p2 = node_map.get(s_data['p2_id'])
            if p1 and p2:
                span = SmartSpan(p1, p2); span.length = s_data['length']; span.conductor = s_data['conductor']; span.has_cg = s_data.get('has_cg', False); span.aug_type = s_data.get('aug_type', 'New'); span.wire_count = s_data.get('wire_count', '3'); span.wire_size = s_data.get('wire_size', '50SQMM'); span.cable_size = s_data.get('cable_size', '10 SQMM'); span.consider_cable = s_data.get('consider_cable', False); span.phase = s_data.get('phase', '3 Phase'); span.is_service_drop = s_data.get('is_service_drop', False); span.update_visuals(); span.label.setPos(s_data['label_x'], s_data['label_y']); span.label.setPlainText(s_data['label_text']); p1.connected_spans.append(span); p2.connected_spans.append(span); self.scene.addItem(span); self.scene.addItem(span.label)
        self.refresh_live_estimate()

    def save_to_file(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Save Project", "", "JSON Files (*.json)")
        if filename:
            with open(filename, 'w') as f: json.dump(self.compile_save_data(), f)

    def load_from_file(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Open Project", "", "JSON Files (*.json)")
        if filename:
            with open(filename, 'r') as f: self.parse_load_data(json.load(f))

    def load_autosave(self):
        if os.path.exists(self.autosave_file):
            try:
                with open(self.autosave_file, 'r') as f:
                    if os.path.getsize(self.autosave_file) > 0:
                        self.parse_load_data(json.load(f))
            except json.JSONDecodeError:
                pass # Ignore corrupted autosave file

    def update_view_drag_mode(self):
        # Panning is enabled only when zoomed in (greater than 100%)
        is_zoomed_in = self.view.transform().m11() > 1.0
        if self.current_tool == "SELECT":
            if is_zoomed_in:
                self.view.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
            else:
                self.view.setDragMode(QGraphicsView.DragMode.RubberBandDrag)
        else:
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag)

    def closeEvent(self, event):
        with open(self.autosave_file, 'w') as f: json.dump(self.compile_save_data(), f)
        super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = EstimateAppV9()
    window.showMaximized()
    sys.exit(app.exec())