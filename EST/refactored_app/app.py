"""
Main application module for the ERP Estimate Generator.

This file contains the main window class `EstimateAppV9` and the application's
entry point.
"""

import sys
import math
import json
import os
import re
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QComboBox, QGraphicsScene, 
                             QFormLayout, QGroupBox, QSpinBox, QLineEdit, 
                             QFileDialog, QMessageBox, QCheckBox, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QSplitter, QGraphicsView)
from PyQt6.QtGui import QPen, QBrush, QColor, QPainter, QPageLayout, QFont
from PyQt6.QtCore import Qt, QTimer, QRectF, pyqtSignal
from PyQt6.QtPrintSupport import QPrinter

# --- Import from refactored modules ---
from constants import TOOLS
from database import setup_database
from rule_engine import DynamicRuleEngine
from ui_components import InteractiveView, DraggableLabel
from canvas_objects import SmartPole, SmartHome, SmartSpan
from ui_dialogs import SearchDialog, SettingsDialog, DatabaseManagerDialog, RulesetManagerDialog
from PyQt6.QtWidgets import QMessageBox


# --- 4. THE MASTER APPLICATION ---
class EstimateAppV9(QMainWindow):
    refresh_signal = pyqtSignal()

    def __init__(self):
        super().__init__()

        # --- Expiry Check ---
        if date.today() >= date(2027, 3, 31):
            QMessageBox.critical(self, "Application Expired", "This app was for testing, a new version has been released. Please check now.")
            sys.exit()

        setup_database()
        self.setWindowTitle("ERP Estimate Generator - Version 4.0 beta")
        self.setGeometry(50, 50, 1600, 900)
        self.current_tool = "SELECT"; self.span_start_pole = None; self.autosave_file = "autosave_erp.json"
        self.bom_overrides = {}
        self.live_bom_data = [] 
        self.escalations = [] 

        main_widget = QWidget(); main_layout = QHBoxLayout(main_widget); self.setCentralWidget(main_widget)
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self.splitter)

        left_panel = QWidget(); left_layout = QVBoxLayout(left_panel); self.splitter.addWidget(left_panel)

        file_toolbar = QHBoxLayout()
        for txt, cmd in [("📄 New", self.new_drawing), ("📂 Open", self.load_from_file), ("💾 Save", self.save_to_file)]:
            btn = QPushButton(txt); btn.clicked.connect(cmd); btn.setStyleSheet("padding: 5px; font-weight: bold;"); file_toolbar.addWidget(btn)
        
        credits_btn = QPushButton("🏆 Credits"); credits_btn.clicked.connect(self.show_credits)
        credits_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #f1c40f; color: black;")
        file_toolbar.addWidget(credits_btn)

        about_btn = QPushButton("ℹ️ About"); about_btn.clicked.connect(self.show_about_dialog)
        about_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #3498db; color: white;")
        file_toolbar.addWidget(about_btn)

        file_toolbar.addStretch()
        
        pdf_btn = QPushButton("🗺️ Export PDF Drawing"); pdf_btn.clicked.connect(self.export_pdf)
        pdf_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #d32f2f; color: white;")
        file_toolbar.addWidget(pdf_btn)

        xl_btn = QPushButton("📊 Generate ERP Excel"); xl_btn.clicked.connect(self.generate_excel)
        xl_btn.setStyleSheet("padding: 5px; font-weight: bold; background-color: #107C41; color: white;")
        file_toolbar.addWidget(xl_btn)

        settings_btn = QPushButton("⚙️")
        settings_btn.clicked.connect(self.open_settings_dialog)
        settings_btn.setFixedWidth(40)
        file_toolbar.addWidget(settings_btn)
        left_layout.addLayout(file_toolbar)

        draw_toolbar = QHBoxLayout(); self.tools_btns = {}
        self.tools = TOOLS
        for key, txt in self.tools.items():
            btn = QPushButton(txt); btn.clicked.connect(lambda checked, t=key: self.set_tool(t))
            btn.setStyleSheet("padding: 8px; font-weight: bold; background-color: lightgray;"); 
            draw_toolbar.addWidget(btn); self.tools_btns[key] = btn
        left_layout.addLayout(draw_toolbar)

        self.scene = QGraphicsScene(); self.scene.selectionChanged.connect(self.on_selection_changed)
        self.view = InteractiveView(self.scene, self); left_layout.addWidget(self.view)

        right_splitter = QSplitter(Qt.Orientation.Vertical)
        self.splitter.addWidget(right_splitter)
        self.splitter.setSizes([950, 650]) # Fixed width allocation
        
        editor_widget = QWidget(); editor_widget_layout = QVBoxLayout(editor_widget)
        self.subject_input = QLineEdit(); self.subject_input.setPlaceholderText("Enter Project Name / Subject...")
        editor_widget_layout.addWidget(QLabel("<b>Project Subject:</b>")); editor_widget_layout.addWidget(self.subject_input)
        
        lat_long_layout = QHBoxLayout()
        self.lat_input = QLineEdit(); self.lat_input.setPlaceholderText("Latitude...")
        self.long_input = QLineEdit(); self.long_input.setPlaceholderText("Longitude...")
        lat_long_layout.addWidget(QLabel("<b>Lat:</b>")); lat_long_layout.addWidget(self.lat_input)
        lat_long_layout.addWidget(QLabel("<b>Long:</b>")); lat_long_layout.addWidget(self.long_input)
        editor_widget_layout.addLayout(lat_long_layout)
        
        self.uh_checkbox = QCheckBox("Use UH (Readymade) Materials instead of Raw Steel"); self.uh_checkbox.setStyleSheet("font-weight: bold; color: #107C41;")
        self.uh_checkbox.stateChanged.connect(self.refresh_live_estimate); editor_widget_layout.addWidget(self.uh_checkbox)
        
        self.editor_group = QGroupBox("1. Object Properties"); self.editor_layout = QFormLayout(); self.editor_group.setLayout(self.editor_layout)
        editor_widget_layout.addWidget(self.editor_group)
        editor_widget_layout.addStretch()
        right_splitter.addWidget(editor_widget)

        table_widget = QWidget(); table_layout = QVBoxLayout(table_widget)
        table_layout.addWidget(QLabel("<b>2. Live Estimate (Double-Click Qty to Edit)</b>"))
        
        self.live_table = QTableWidget(0, 6)
        self.live_table.setHorizontalHeaderLabels(["Type", "Code", "Name", "Qty", "Unit", "Total (Rs)"])
        self.live_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.live_table.setColumnWidth(0, 60); self.live_table.setColumnWidth(1, 80); self.live_table.setColumnWidth(3, 60)
        self.live_table.itemChanged.connect(self.on_table_edit)
        table_layout.addWidget(self.live_table)
        
        search_layout = QHBoxLayout()
        add_mat_btn = QPushButton("+ Add Custom Material"); add_mat_btn.clicked.connect(lambda: self.open_search("Material"))
        add_lab_btn = QPushButton("+ Add Custom Labor"); add_lab_btn.clicked.connect(lambda: self.open_search("Labor"))
        add_mat_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; padding: 5px;")
        add_lab_btn.setStyleSheet("background-color: #e67e22; color: white; font-weight: bold; padding: 5px;")
        search_layout.addWidget(add_mat_btn); search_layout.addWidget(add_lab_btn)
        table_layout.addLayout(search_layout)

        self.grand_total_label = QLabel("<b>Grand Total: Rs. 0.00</b>")
        self.grand_total_label.setStyleSheet("font-size: 16px; color: #d32f2f; margin-top: 5px;")
        table_layout.addWidget(self.grand_total_label)
        right_splitter.addWidget(table_widget)
        right_splitter.setSizes([300, 700])

        self.set_tool("SELECT"); self.load_autosave()
        self.refresh_signal.connect(self.refresh_live_estimate)

    def open_db_manager(self):
        DatabaseManagerDialog(self).exec()
    
    def open_rule_manager(self):
        RulesetManagerDialog(self).exec() 

    def open_settings_dialog(self):
        dialog = SettingsDialog(self)
        dialog.exec()

    def open_search(self, item_type):
        """Opens a search dialog to add custom materials or labor."""
        dialog = SearchDialog(item_type, self)
        if dialog.exec():
            selected_item = dialog.get_selected()
            if selected_item:
                # Add the item to overrides with a default quantity of 1
                self.bom_overrides[selected_item['name']] = {
                    "qty": 1, 
                    "type": selected_item['type']
                }
                self.refresh_live_estimate()

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace): self.delete_selected_items()
        super().keyPressEvent(event)
        
    def delete_selected_items(self):
        items = self.scene.selectedItems()
        for item in items:
            if isinstance(item, SmartSpan): self.delete_item(item)
        for item in items:
            if isinstance(item, (SmartPole, SmartHome)): self.delete_item(item)

    def set_tool(self, tool_name):
        self.current_tool = tool_name
        if self.span_start_pole: self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1))
        self.span_start_pole = None
        for key, btn in self.tools_btns.items():
            btn.setStyleSheet("padding: 8px; font-weight: bold; background-color: " + ("lightblue" if key == tool_name else "lightgray"))
        self.update_view_drag_mode()

    def handle_canvas_click(self, event, view):
        if event.button() == Qt.MouseButton.RightButton: self.set_tool("SELECT"); return
        if self.current_tool == "SELECT": return

        pos = view.mapToScene(event.pos()); item_clicked = self.scene.itemAt(pos, view.transform())

        if self.current_tool in ["ADD_LT", "ADD_HT", "ADD_DTR", "ADD_EXISTING"]:
            p_type = "LT" if self.current_tool == "ADD_EXISTING" else self.current_tool.split("_")[1]
            pole = SmartPole(pos.x(), pos.y(), self.refresh_signal, p_type, self.current_tool == "ADD_EXISTING")
            self.scene.addItem(pole)
            self.refresh_live_estimate()
        elif self.current_tool == "ADD_HOME":
            home = SmartHome(pos.x(), pos.y(), self.refresh_signal)
            self.scene.addItem(home)
            self.refresh_live_estimate()
        elif self.current_tool == "ADD_SPAN" and isinstance(item_clicked, (SmartPole, SmartHome)):
            if not self.span_start_pole:
                self.span_start_pole = item_clicked; item_clicked.setPen(QPen(Qt.GlobalColor.yellow, 3)) 
            elif self.span_start_pole != item_clicked: 
                span = SmartSpan(self.span_start_pole, item_clicked)
                self.span_start_pole.connected_spans.append(span); item_clicked.connected_spans.append(span)
                if (isinstance(self.span_start_pole, SmartPole) and isinstance(item_clicked, SmartPole)) and (self.span_start_pole.pole_type == "HT" and item_clicked.pole_type == "LT" or self.span_start_pole.pole_type == "LT" and item_clicked.pole_type == "HT"):
                    choice = QMessageBox.question(self, 'Warning', "Are you sure you want to connect HT pole to LT pole?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                    if choice == QMessageBox.StandardButton.No: return
                self.scene.addItem(span); self.scene.addItem(span.label)
                self.span_start_pole.setPen(QPen(Qt.GlobalColor.black, 1)); self.span_start_pole = None
                self.refresh_live_estimate()

    def on_selection_changed(self):
        try:
            # Guard against the scene being deleted when this is called via a QTimer
            if not self.scene.views():
                return
        except RuntimeError:
            return

        while self.editor_layout.count():
            child = self.editor_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()

        sel = self.scene.selectedItems()
        if not sel: self.editor_group.setTitle("Select an item to edit"); return
        if len(sel) > 1: self.editor_group.setTitle(f"{len(sel)} Items Selected"); return
        item = sel[0]
        
        if isinstance(item, DraggableLabel): self.editor_group.setTitle("Text Label Selected"); return
        
        if isinstance(item, SmartHome):
            self.editor_group.setTitle("Editing Consumer Home")
            self.editor_layout.addRow(QLabel("<i>Service cable connection point. Select the actual Line (Span) connecting to this home to edit PVC sizes and Phase details.</i>"))
            
            note_edit = QLineEdit()
            note_edit.setText(getattr(item, 'custom_note', ''))
            note_edit.setPlaceholderText("Add a custom note...")
            note_edit.textChanged.connect(lambda t, i=item: self.update_item_note(i, t))
            self.editor_layout.addRow("Custom Note:", note_edit)

            del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
            del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)
            return

        if isinstance(item, SmartPole):
            if item.is_existing:
                self.editor_group.setTitle("Editing Existing Pole")
                pole_type_cb = QComboBox(); pole_type_cb.addItems(["LT", "HT", "DTR"]); pole_type_cb.setCurrentText(item.pole_type)
                pole_type_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "pole_type", t)); self.editor_layout.addRow("Pole Type:", pole_type_cb)
                stay_type_cb = QComboBox(); stay_type_cb.addItems(["HT", "LT"]); stay_type_cb.setCurrentText(getattr(item, 'stay_type', 'HT'))
                stay_type_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "stay_type", t)); self.editor_layout.addRow("Stay Type:", stay_type_cb)
                stay_spin = QSpinBox(); stay_spin.setRange(0, 10); stay_spin.setValue(item.stay_count)
                stay_spin.valueChanged.connect(lambda v: self.update_pole(item, "stay_count", v)); self.editor_layout.addRow("Stay Sets:", stay_spin)
                del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
                del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)
                return

            self.editor_group.setTitle(f"Editing {item.pole_type} Structure")
            height_cb = QComboBox(); height_cb.addItems(["8MTR", "9MTR"]); height_cb.setCurrentText(item.height)
            height_cb.currentTextChanged.connect(lambda t: self.update_pole(item, "height", t)); self.editor_layout.addRow("Height:", height_cb)

            if item.height == "8MTR" and item.pole_type in ["HT", "DTR"]:
                ext_check = QCheckBox("Add 8MTR Extension")
                ext_check.setChecked(item.has_extension)
                ext_check.stateChanged.connect(lambda v, i=item: self.update_pole_extension(i, v == 2))
                self.editor_layout.addRow(ext_check)
            if item.height == "9MTR" and item.pole_type in ["HT", "DTR"]:
                ext_check = QCheckBox("Add 9MTR Extension")
                ext_check.stateChanged.connect(lambda v, i=item: self.update_pole_extension(i, v == 2))
                self.editor_layout.addRow(ext_check)

            if item.pole_type == "DTR":
                dtr_cb = QComboBox(); dtr_cb.addItems(["None", "16 KVA", "25KVA", "63KVA", "100KVA", "160KVA"])
                dtr_cb.setCurrentText(item.dtr_size); dtr_cb.currentTextChanged.connect(lambda t: self.update_dtr_logic(item, t)); self.editor_layout.addRow("DTR Size:", dtr_cb)
            
            earth_spin = QSpinBox(); earth_spin.setRange(0, 10); earth_spin.setValue(item.earth_count)
            earth_spin.valueChanged.connect(lambda v: self.update_pole(item, "earth_count", v)); self.editor_layout.addRow("Earthing Sets:", earth_spin)
            stay_spin = QSpinBox(); stay_spin.setRange(0, 10); stay_spin.setValue(item.stay_count)
            stay_spin.valueChanged.connect(lambda v: self.update_pole(item, "stay_count", v)); self.editor_layout.addRow("Stay Sets:", stay_spin)

            note_edit = QLineEdit()
            note_edit.setText(getattr(item, 'custom_note', ''))
            note_edit.setPlaceholderText("Add a custom note...")
            note_edit.textChanged.connect(lambda t, i=item: self.update_item_note(i, t))
            self.editor_layout.addRow("Custom Note:", note_edit)
            if item.height == "8MTR" and item.pole_type in ["LT", "HT", "DTR"]:
                ext_check = QCheckBox("Add Extension");
                ext_check.setChecked(item.has_extension);
                ext_check.stateChanged.connect(lambda v, i=item:self.update_pole_extension(i, v==2)); self.editor_layout.addRow(ext_check)

        elif isinstance(item, SmartSpan):
            length_spin = QSpinBox(); length_spin.setRange(1, 150); length_spin.setValue(int(item.length))
            length_spin.valueChanged.connect(lambda v: self.update_span(item, "length", v))
            
            if item.is_service_drop:
                self.editor_group.setTitle("Editing Service Connection")
                self.editor_layout.addRow("Length (Meters):", length_spin)
                phase_cb = QComboBox(); phase_cb.addItems(["1 Phase", "3 Phase"]); phase_cb.setCurrentText(item.phase)
                phase_cb.currentTextChanged.connect(lambda t: self.update_span(item, "phase", t)); self.editor_layout.addRow("Phase:", phase_cb)
                cons_cb = QCheckBox("Consider Cable in Estimate?"); cons_cb.setChecked(item.consider_cable)
                cons_cb.stateChanged.connect(lambda v: self.update_span(item, "consider_cable", v == 2)); self.editor_layout.addRow(cons_cb)
                sz_cb = QComboBox(); sz_cb.addItems(["10 SQMM", "16 SQMM", "25 SQMM"]); sz_cb.setCurrentText(item.cable_size)
                sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "cable_size", t)); self.editor_layout.addRow("PVC Size:", sz_cb)
            else:
                self.editor_group.setTitle("Editing Span")
                self.editor_layout.addRow("Length (Meters):", length_spin)
                cg_chk = QCheckBox(); cg_chk.setChecked(item.has_cg)
                cg_chk.stateChanged.connect(lambda v: self.update_span(item, "has_cg", v == 2)); self.editor_layout.addRow("Cattle Guard:", cg_chk)
                
                cond_cb = QComboBox(); cond_cb.addItems(["ACSR", "AB Cable", "PVC Cable"]); cond_cb.setCurrentText(item.conductor)
                cond_cb.currentTextChanged.connect(lambda t: self.update_conductor_logic(item, t)); self.editor_layout.addRow("Conductor:", cond_cb)
                
                if item.conductor == "ACSR":
                    wire_cnt_cb = QComboBox(); wire_cnt_cb.addItems(["2", "3", "4"]); wire_cnt_cb.setCurrentText(item.wire_count)
                    wire_cnt_cb.currentTextChanged.connect(lambda t: self.update_span(item, "wire_count", t)); self.editor_layout.addRow("Wire Count:", wire_cnt_cb)
                    wire_sz_cb = QComboBox(); wire_sz_cb.addItems(["30SQMM", "50SQMM"]); wire_sz_cb.setCurrentText(item.wire_size)
                    wire_sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "wire_size", t)); self.editor_layout.addRow("Wire Size:", wire_sz_cb)
                elif item.conductor == "PVC Cable":
                    sz_cb = QComboBox(); sz_cb.addItems(["10 SQMM", "16 SQMM", "25 SQMM"]); sz_cb.setCurrentText(item.cable_size)
                    sz_cb.currentTextChanged.connect(lambda t: self.update_span(item, "cable_size", t)); self.editor_layout.addRow("PVC Size:", sz_cb)
                    
                aug_cb = QComboBox(); aug_cb.addItems(["New", "Replace 2W->4W", "Add-on 2W"]); aug_cb.setCurrentText(item.aug_type)
                aug_cb.currentTextChanged.connect(lambda t: self.update_span(item, "aug_type", t)); self.editor_layout.addRow("Work Nature:", aug_cb)

            note_edit = QLineEdit()
            note_edit.setText(getattr(item, 'custom_note', ''))
            note_edit.setPlaceholderText("Add a custom note...")
            note_edit.textChanged.connect(lambda t, i=item: self.update_item_note(i, t))
            self.editor_layout.addRow("Custom Note:", note_edit)

        del_btn = QPushButton("🗑 Delete Selected"); del_btn.setStyleSheet("background-color: #ff4c4c; color: white;")
        del_btn.clicked.connect(lambda: self.delete_item(item)); self.editor_layout.addRow(del_btn)

    def update_pole(self, item, prop, value):
        if prop == 'stay_count':
            item.override_auto_stay = True
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()
        QTimer.singleShot(10, self.on_selection_changed) 

    def update_pole_extension(self, item, value):
        item.has_extension = value
        item.update_visuals()
        self.refresh_live_estimate()
        
    def update_span(self, item, prop, value): 
        setattr(item, prop, value)
        item.update_visuals()
        self.refresh_live_estimate()

    def update_dtr_logic(self, item, size): 
        item.dtr_size = size
        item.earth_count = 5 if size != "None" else 2
        item.update_visuals()
        self.refresh_live_estimate()

    def update_conductor_logic(self, item, conductor): 
        item.conductor = conductor
        item.update_visuals()
        QTimer.singleShot(50, self.on_selection_changed)
        self.refresh_live_estimate()

    def update_item_note(self, item, text):
        item.custom_note = text
        item.update_visuals()

    def delete_item(self, item):
        if not item or not item.scene(): return
        if hasattr(item, 'connected_spans'):
            for span in list(item.connected_spans):
                if span.label.scene(): self.scene.removeItem(span.label)
                if span.scene(): self.scene.removeItem(span)
                if span in getattr(span.p1, 'connected_spans', []): span.p1.connected_spans.remove(span)
                if span in getattr(span.p2, 'connected_spans', []): span.p2.connected_spans.remove(span)
        if isinstance(item, SmartSpan) and item.label.scene(): self.scene.removeItem(item.label)
        if item.scene(): self.scene.removeItem(item)
        self.refresh_live_estimate()

    def on_table_edit(self, item):
        if item.column() == 3: 
            try:
                new_qty = float(item.text())
                name = self.live_table.item(item.row(), 2).text()
                row_type = self.live_table.item(item.row(), 0).text()
                self.bom_overrides[name] = {"qty": new_qty, "type": row_type}
                self.refresh_live_estimate()
            except (ValueError, RuntimeError): pass 

    def recalculate_all_span_types(self):
        all_poles = [item for item in self.scene.items() if isinstance(item, SmartPole)]
        effectively_existing_poles = set(p for p in all_poles if p.is_existing)
        
        while True:
            promoted_this_round = set()
            poles_to_check = [p for p in all_poles if p not in effectively_existing_poles]
            
            for pole in poles_to_check:
                connected_to_existing = 0
                for span in pole.connected_spans:
                    other_end = span.p1 if span.p2 == pole else span.p2
                    if other_end in effectively_existing_poles:
                        connected_to_existing += 1
                
                if connected_to_existing >= 2:
                    promoted_this_round.add(pole)
            
            if not promoted_this_round:
                break
            
            effectively_existing_poles.update(promoted_this_round)
            
        all_spans = [item for item in self.scene.items() if isinstance(item, SmartSpan)]
        for span in all_spans:
            is_now_existing = (span.p1 in effectively_existing_poles and span.p2 in effectively_existing_poles)
            
            if span.is_service_drop:
                is_now_existing = False
                
            if span.is_existing_span != is_now_existing:
                span.is_existing_span = is_now_existing
                span.update_visuals()

    def refresh_live_estimate(self):
        self.recalculate_all_span_types()

        all_poles = [item for item in self.scene.items() if isinstance(item, SmartPole)]
        poles_to_update = []

        for pole in all_poles:
            if pole.pole_type == "DTR" or getattr(pole, 'override_auto_stay', False):
                continue

            spans = [s for s in pole.connected_spans if not s.is_service_drop and not s.is_existing_span]
            num_spans = len(spans)
            
            should_have_stay = False
            if num_spans == 1:
                should_have_stay = True
            elif num_spans == 2:
                span1, span2 = spans[0], spans[1]
                p0 = pole
                p1 = span1.p1 if span1.p2 == p0 else span1.p2
                p2 = span2.p1 if span2.p2 == p0 else span2.p2
                v1 = (p1.x() - p0.x(), p1.y() - p0.y()); v2 = (p2.x() - p0.x(), p2.y() - p0.y())
                dot = v1[0] * v2[0] + v1[1] * v2[1]
                mag1 = math.sqrt(v1[0]**2 + v1[1]**2); mag2 = math.sqrt(v2[0]**2 + v2[1]**2)
                if mag1 > 0 and mag2 > 0:
                    angle_rad = math.acos(min(1.0, max(-1.0, dot / (mag1 * mag2))))
                    turning_angle = 180 - math.degrees(angle_rad)
                    if turning_angle > 20:
                        should_have_stay = True
            
            if pole.stay_count != int(should_have_stay):
                poles_to_update.append((pole, int(should_have_stay)))
        
        if poles_to_update:
            for pole, new_stay_count in poles_to_update:
                pole.stay_count = new_stay_count
                pole.update_visuals()

        raw_bom = {}; total_lab_tasks = {}; use_uh = self.uh_checkbox.isChecked()

        if not hasattr(self, 'rule_engine'):
            self.rule_engine = DynamicRuleEngine()
        
        rules = []
        try:
            with open('rules.json', 'r') as f:
                rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass 

        canvas_items = [item for item in self.scene.items() if isinstance(item, (SmartPole, SmartSpan, SmartHome))]
        raw_bom, total_lab_tasks = self.rule_engine.process(canvas_items, rules, use_uh)

        # --- Process the raw BOM into the final live_bom_data list ---
        self.live_bom_data = []
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()

        # Combine materials and labor into one list for processing
        combined_bom = [('Material', name, qty) for name, qty in raw_bom.items()] + \
                       [('Labor', name, qty) for name, qty in total_lab_tasks.items()]

        all_processed_names = set()

        for item_type, name, qty in combined_bom:
            # Apply overrides
            if name in self.bom_overrides and self.bom_overrides[name]['type'] == item_type:
                qty = self.bom_overrides[name]['qty']

            if item_type == 'Material':
                cursor.execute("SELECT item_code, rate, unit FROM materials WHERE item_name=?", (name,))
                res = cursor.fetchone()
                if res:
                    code, rate, unit = res
                    self.live_bom_data.append({'type': 'Material', 'code': code, 'name': name, 'qty': qty, 'unit': unit, 'rate': rate, 'amt': qty * rate})
            elif item_type == 'Labor':
                cursor.execute("SELECT labor_code, rate, unit FROM labor WHERE task_name=?", (name,))
                res = cursor.fetchone()
                if res:
                    code, rate, unit = res
                    self.live_bom_data.append({'type': 'Labor', 'code': code, 'name': name, 'qty': qty, 'unit': unit, 'rate': rate, 'amt': qty * rate})
            all_processed_names.add(name)

        # Add any custom items from overrides that were not in the auto-generated BOM
        for name, override in self.bom_overrides.items():
            if name not in all_processed_names:
                qty = override['qty']
                item_type = override['type']
                if item_type == 'Material':
                    cursor.execute("SELECT item_code, rate, unit FROM materials WHERE item_name=?", (name,))
                    res = cursor.fetchone()
                    if res:
                        code, rate, unit = res
                        self.live_bom_data.append({'type': 'Material', 'code': code, 'name': name, 'qty': qty, 'unit': unit, 'rate': rate, 'amt': qty * rate})
                elif item_type == 'Labor':
                    cursor.execute("SELECT labor_code, rate, unit FROM labor WHERE task_name=?", (name,))
                    res = cursor.fetchone()
                    if res:
                        code, rate, unit = res
                        self.live_bom_data.append({'type': 'Labor', 'code': code, 'name': name, 'qty': qty, 'unit': unit, 'rate': rate, 'amt': qty * rate})
        
        conn.close()

        # --- Update UI Table ---
        try:
            self.live_table.itemChanged.disconnect(self.on_table_edit)
        except TypeError:
            pass # Signal was not connected

        self.live_table.setRowCount(0)
        
        for i, item in enumerate(self.live_bom_data):
            self.live_table.insertRow(i)
            self.live_table.setItem(i, 0, QTableWidgetItem(item['type']))
            self.live_table.setItem(i, 1, QTableWidgetItem(item['code']))
            self.live_table.setItem(i, 2, QTableWidgetItem(item['name']))
            qty_item = QTableWidgetItem(f"{item['qty']:.3f}")
            qty_item.setBackground(QColor("#fff3cd"))
            self.live_table.setItem(i, 3, qty_item)
            self.live_table.setItem(i, 4, QTableWidgetItem(item['unit']))
            self.live_table.setItem(i, 5, QTableWidgetItem(f"{item['amt']:.2f}"))
            
            # Make all columns except Qty non-editable
            for col in [0, 1, 2, 4, 5]:
                t_item = self.live_table.item(i, col)
                if t_item:
                    t_item.setFlags(t_item.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self.live_table.itemChanged.connect(self.on_table_edit)
        
        # --- Recalculate Totals ---
        mat_base = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Material'])
        current_date = datetime.now()
        current_fy_start = current_date.year if current_date.month >= 4 else current_date.year - 1
        
        self.escalations = []
        current_mat_val = mat_base
        for year in range(2024, current_fy_start + 1):
            esc_amt = current_mat_val * 0.05
            self.escalations.append((f"{str(year)[-2:]}-{str(year+1)[-2:]}", esc_amt))
            current_mat_val += esc_amt
            
        sun = current_mat_val * 0.05
        mat_sub = current_mat_val + sun
        
        lab_sub = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Labor'])
        sup = (mat_sub + lab_sub) * 0.10
        gst = lab_sub * 0.18
        cess = (mat_sub + lab_sub + sup) * 0.01
        final_amt = mat_sub + lab_sub + sup + gst + cess
        self.grand_total_label.setText(f"<b>Estimated Cost (Inc Taxes): Rs. {final_amt:,.2f}</b>")

    def show_about_dialog(self):
        about_text = """
        <h2>ERP Estimate Generator v4.0 beta</h2>
        <p>This application is designed to help engineers create, visualize, and estimate costs for electrical network projects.</p>
        <p><b>Key Features:</b></p>
        <ul>
            <li>Interactive canvas for drawing poles, spans, and consumers.</li>
            <li>Live bill of materials (BOM) and labor cost estimation.</li>
            <li>Export drawings to PDF and estimates to ERP-compatible Excel formats.</li>
            <li>Customizable material and labor rates via an internal database.</li>
        </ul>
        <br>
        <p>Developed by: <b>Pramod Verma</b></p>
        """
        QMessageBox.information(self, "About ERP Estimate Generator", about_text)

    def show_credits(self):
        credits_text = """
        <h2 style='color:#3498db;'>Contributors & Special Thanks</h2>
        <p>This application has been improved with the help of the following individuals:</p>
        <ul style='list-style-type: none; padding-left: 0;'>
            <li style='margin-bottom: 10px;'>
                <b>Praful Singh:</b>
                <ul style='margin-top: 4px; list-style-type: disc; margin-left: 20px;'>
                    <li>Identified visual improvements for existing and new lines.</li>
                    <li>Helped in refining the PDF legend section.</li>
                </ul>
            </li>
            <li style='margin-bottom: 10px;'>
                <b>Rajsekhar Gorai:</b>
                <ul style='margin-top: 4px; list-style-type: disc; margin-left: 20px;'>
                    <li>Helped identify and fix the 8mtr HT pole extension logic.</li>
                </ul>
            </li>
            <li style='margin-bottom: 10px;'>
                <b>Amit Karmakar:</b>
                <ul style='margin-top: 4px; list-style-type: disc; margin-left: 20px;'>
                    <li>Suggested adding existing DTR properties.</li>
                    <li>Recommended adding Latitude/Longitude fields.</li>
                </ul>
            </li>
        </ul>
        <p style='margin-top: 15px; font-style: italic;'>And thanks to everyone who provided feedback!</p>
        """
        QMessageBox.information(self, "Credits", credits_text)

    def generate_excel(self):
        subject = self.subject_input.text()
        sanitized_subject = "".join(c for c in subject if c not in '\\/*?:"<>|')
        default_filename = f"{sanitized_subject}_Estimate.xlsx" if sanitized_subject else "ERP_Estimate.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self, "Export ERP Estimate", default_filename, "Excel Files (*.xlsx)")
        if not filename: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estimate"
        ws.merge_cells('A1:G1'); ws['A1'] = "AUTOMATED ERP ESTIMATE"; ws['A1'].font = Font(bold=True, size=14, color="FFFFFF"); ws['A1'].fill = PatternFill("solid", fgColor="4F81BD"); ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2'); ws['A2'] = f"Subject: {self.subject_input.text()} | Date: {datetime.now().strftime('%d-%m-%Y')}"
        ws.append(["Sl No.", "Material Code", "Description", "Qty", "Unit", "Rate", "Amount"])
        for col_num, cell in enumerate(ws[3], 1): cell.font = Font(bold=True)
        ws.column_dimensions['C'].width = 45; ws.column_dimensions['B'].width = 15
        
        total_lab = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Labor'])

        row = 4; ws.cell(row, 3, "A. MATERIALS").font = Font(bold=True); row += 1
        for i, item in enumerate([x for x in self.live_bom_data if x['type'] == 'Material'], 1):
            ws.append([i, item['code'], item['name'], round(item['qty'],3), item['unit'], item['rate'], round(item['amt'],2)]); row += 1
        
        mat_base = sum([x['amt'] for x in self.live_bom_data if x['type'] == 'Material'])
        ws.append(["", "", "Material Base Total", "", "", "", round(mat_base, 2)]); row += 1
        
        current_mat_val = mat_base
        for fy, esc_amt in self.escalations:
            ws.append(["", "", f"Add: Escalation @ 5% for FY {fy}", "", "", "", round(esc_amt, 2)]); row += 1
            current_mat_val += esc_amt
            
        sun = current_mat_val * 0.05
        mat_subtotal = current_mat_val + sun
        
        ws.append(["", "", "Add: Sundries @ 5%", "", "", "", round(sun, 2)]); row += 1
        ws.append(["", "", "TOTAL MATERIAL COST (A)", "", "", "", round(mat_subtotal, 2)])
        ws.cell(row, 3).font = Font(bold=True); ws.cell(row, 7).font = Font(bold=True); row += 2
        
        ws.cell(row, 3, "B. ERECTION / LABOR").font = Font(bold=True); row += 1
        for i, item in enumerate([x for x in self.live_bom_data if x['type'] == 'Labor'], 1):
            ws.append([i, "", item['name'], round(item['qty'],3), item['unit'], item['rate'], round(item['amt'],2)]); row += 1
            
        ws.append(["", "", "TOTAL LABOR COST (B)", "", "", "", round(total_lab, 2)])
        ws.cell(row, 3).font = Font(bold=True); ws.cell(row, 7).font = Font(bold=True); row += 2

        sup = (mat_subtotal + total_lab) * 0.10; gst = total_lab * 0.18; cess = (mat_subtotal + total_lab + sup) * 0.01; sub_c = mat_subtotal + total_lab + sup + gst; g_tot = sub_c + cess
        ws.append(["", "", "C. OVERHEADS & TAXES"]); ws.cell(row, 3).font = Font(bold=True); row += 1
        ws.append(["", "", "Supervision @ 10% on (A+B)", "", "", "", round(sup, 2)]); row += 1
        ws.append(["", "", "GST @ 18% on (Labor Only)", "", "", "", round(gst, 2)]); row += 1
        ws.append(["", "", "Sub-Total", "", "", "", round(sub_c, 2)]); row += 1
        ws.append(["", "", "Add: Cess @ 1% on (Mat+Lab+Sup)", "", "", "", round(cess, 2)]); row += 1
        ws.append(["", "", "GRAND TOTAL", "", "", "", round(g_tot, 2)])
        ws.cell(row, 3).font = Font(bold=True, size=12); ws.cell(row, 7).font = Font(bold=True, size=12, color="FF0000")
        
        wb.save(filename); QMessageBox.information(self, "Success", f"ERP Estimate Excel saved to:\n{filename}")

    def export_pdf(self):
        subject = self.subject_input.text()
        sanitized_subject = "".join(c for c in subject if c not in '\\/*?:"<>|')
        default_filename = f"{sanitized_subject}.pdf" if sanitized_subject else "Project_Drawing.pdf"
        filename, _ = QFileDialog.getSaveFileName(self, "Export PDF Drawing", default_filename, "PDF Files (*.pdf)")
        if not filename: return

        printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)

        source_rect = self.scene.itemsBoundingRect()
        if source_rect.isNull():
            QMessageBox.warning(self, "Empty", "Canvas is empty.")
            return

        center = source_rect.center()
        min_dim = 300
        new_width = max(source_rect.width(), min_dim)
        new_height = max(source_rect.height(), min_dim)
        source_rect = QRectF(0, 0, new_width, new_height)
        source_rect.moveCenter(center)

        if source_rect.width() > source_rect.height():
            printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        else:
            printer.setPageOrientation(QPageLayout.Orientation.Portrait)

        painter = QPainter(printer)
        page_rect_px = printer.pageRect(QPrinter.Unit.DevicePixel)
        margin_px = 10

        border_rect = page_rect_px.adjusted(margin_px, margin_px, -margin_px, -margin_px)

        painter.setPen(Qt.GlobalColor.black)
        title_font = QFont("Arial", 12, QFont.Weight.Bold); title_font.setUnderline(True)
        painter.setFont(title_font)
        title_text = self.subject_input.text() or 'ERP PROJECT DRAWING'
        text_flags = Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop | Qt.TextFlag.TextWordWrap
        title_rect_calc = QRectF(border_rect.x() + 5, border_rect.y(), border_rect.width() - 10, 9999)
        required_rect = painter.boundingRect(title_rect_calc, text_flags, title_text)
        title_height = required_rect.height()
        title_rect = QRectF(border_rect.x(), border_rect.y(), border_rect.width(), title_height)
        painter.drawText(title_rect, text_flags, title_text)
        
        scene_target_rect = QRectF(border_rect)
        scene_target_rect.setTop(border_rect.top() + title_height + 10)
        source_rect.adjust(-50, -50, 50, 50)
        self.scene.render(painter, scene_target_rect, source_rect, Qt.AspectRatioMode.KeepAspectRatio)
        
        painter.setFont(QFont("Arial", 8)); painter.setPen(Qt.GlobalColor.black)

        legend_data = {
            'New Poles': {'LT Pole': {'s': '🔵', 'q': 0}, 'HT Pole': {'s': '🔴', 'q': 0}, 'DP/DTR': {'s': '🟩', 'q': 0}},
            'Existing Poles': {'Ex. Pole': {'s': '⚪', 'q': 0}, 'Ex. DP/DTR': {'s': '⚪', 'q': 0}},
            'Ancillary': {'Consumer Home': {'s': '🏠', 'q': 0}, 'Earth': {'s': '⏚', 'q': 0}, 'Stay': {'s': 'S', 'q': 0}},
            'New Spans': {'New ACSR': {'s': '---', 'l': 0}, 'New ABC': {'s': '~--~', 'l': 0}, 'New PVC': {'s': '~~~', 'l': 0}},
            'Existing Spans': {'Ex. ACSR': {'s': '———', 'l': 0}, 'Ex. ABC': {'s': '~~~~~', 'l': 0}, 'Ex. PVC': {'s': '~~~~~', 'l': 0}}
        }
        
        for item in self.scene.items():
            if isinstance(item, SmartPole):
                legend_data['Ancillary']['Earth']['q'] += item.earth_count
                legend_data['Ancillary']['Stay']['q'] += item.stay_count
                if item.is_existing:
                    if item.pole_type == 'DTR': legend_data['Existing Poles']['Ex. DP/DTR']['q'] += 1
                    else: legend_data['Existing Poles']['Ex. Pole']['q'] += 1
                else:
                    if item.pole_type == 'LT': legend_data['New Poles']['LT Pole']['q'] += 1
                    elif item.pole_type == 'HT': legend_data['New Poles']['HT Pole']['q'] += 1
                    elif item.pole_type == 'DTR': legend_data['New Poles']['DP/DTR']['q'] += 1
            elif isinstance(item, SmartHome):
                legend_data['Ancillary']['Consumer Home']['q'] += 1
            elif isinstance(item, SmartSpan):
                span_len = item.length
                if item.is_existing_span:
                    if item.conductor == 'ACSR': legend_data['Existing Spans']['Ex. ACSR']['l'] += span_len
                    elif item.conductor == 'AB Cable': legend_data['Existing Spans']['Ex. ABC']['l'] += span_len
                    elif item.conductor == 'PVC Cable': legend_data['Existing Spans']['Ex. PVC']['l'] += span_len
                else:
                    if item.conductor == 'ACSR': legend_data['New Spans']['New ACSR']['l'] += span_len
                    elif item.conductor == 'AB Cable': legend_data['New Spans']['New ABC']['l'] += span_len
                    elif item.conductor == 'PVC Cable': legend_data['New Spans']['New PVC']['l'] += span_len

        used_items = []
        for category, items in legend_data.items():
            for desc, data in items.items():
                qty = data.get('q')
                length = data.get('l')
                if (qty and qty > 0) or (length and length > 0):
                    val = str(qty) if qty is not None else f"{length}m"
                    used_items.append({'desc': desc, 'sym': data['s'], 'val': val})

        if used_items:
            col_widths = {'sl': 30, 'sym': 50, 'desc': 130, 'qty': 50}
            table_width = sum(col_widths.values())
            row_height = 18
            
            table_height = (len(used_items) + 1) * row_height
            latlong_box_height = 25
            total_legend_height = table_height + latlong_box_height
            
            legend_block_rect = QRectF(border_rect.left() + 5, border_rect.bottom() - total_legend_height - 5, table_width, total_legend_height)
            
            painter.setBrush(QBrush(QColor(255, 255, 255, 220)))
            painter.setPen(QPen(QColor(200, 200, 200), 0.5))
            painter.drawRect(legend_block_rect)
            painter.setPen(QPen(Qt.GlobalColor.black))

            current_y = legend_block_rect.top()
            
            painter.setFont(QFont("Arial", 8, QFont.Weight.Bold))
            current_x = legend_block_rect.left()
            headers = {'sl': ' Sl No. ', 'sym': ' Symbol ', 'desc': ' Description ', 'qty': ' Qty/Len '}
            for key, width in col_widths.items():
                painter.drawText(QRectF(current_x, current_y, width, row_height), Qt.AlignmentFlag.AlignCenter, headers[key])
                current_x += width
            
            current_y += row_height
            
            for i, item in enumerate(used_items):
                sl_no = str(i + 1)
                current_x = legend_block_rect.left()
                painter.setFont(QFont("Arial", 8))

                painter.drawText(QRectF(current_x, current_y, col_widths['sl'], row_height), Qt.AlignmentFlag.AlignCenter, sl_no)
                current_x += col_widths['sl']
                painter.drawText(QRectF(current_x, current_y, col_widths['sym'], row_height), Qt.AlignmentFlag.AlignCenter, item['sym'])
                current_x += col_widths['sym']
                painter.drawText(QRectF(current_x + 5, current_y, col_widths['desc'] - 5, row_height), Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, item['desc'])
                current_x += col_widths['desc']
                painter.drawText(QRectF(current_x, current_y, col_widths['qty'], row_height), Qt.AlignmentFlag.AlignCenter, item['val'])
                
                current_y += row_height

            painter.setPen(QPen(QColor(220, 220, 220), 1))
            for i in range(len(used_items) + 2):
                 y = legend_block_rect.top() + (i * row_height)
                 painter.drawLine(int(legend_block_rect.left()), int(y), int(legend_block_rect.right()), int(y))
            current_x = legend_block_rect.left()
            for width in col_widths.values():
                current_x += width
                painter.drawLine(int(current_x), int(legend_block_rect.top()), int(current_x), int(current_y - row_height))

            painter.setPen(QPen(Qt.GlobalColor.black))

            painter.setFont(QFont("Arial", 7, QFont.Weight.Normal, italic=True))
            lat_long_text = f"Lat: {self.lat_input.text()}   Long: {self.long_input.text()}"
            lat_long_rect = QRectF(legend_block_rect.left(), current_y, table_width, latlong_box_height)
            painter.drawText(lat_long_rect, Qt.AlignmentFlag.AlignCenter, lat_long_text)
        
        painter.end()
        QMessageBox.information(self, "Success", f"PDF Drawing exported to:\n{filename}")

    def new_drawing(self):
        if QMessageBox.question(self, 'New Canvas', 'Clear canvas?', QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self.scene.clear(); self.subject_input.clear(); self.span_start_pole = None; self.uh_checkbox.setChecked(False); self.bom_overrides.clear()
            
    def update_view_drag_mode(self):
        is_zoomed_in = self.view.transform().m11() > 1.0
        if self.current_tool == "SELECT":
            if is_zoomed_in:
                self.view.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
            else:
                self.view.setDragMode(QGraphicsView.DragMode.RubberBandDrag)
        else:
            self.view.setDragMode(QGraphicsView.DragMode.NoDrag)
            
    def load_from_file(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Open Project", "", "JSON Files (*.json)")
        if filename:
            with open(filename, 'r') as f: self.parse_load_data(json.load(f))

    def save_to_file(self):
        filename, _ = QFileDialog.getSaveFileName(self, "Save Project", "", "JSON Files (*.json)")
        if filename:
            with open(filename, 'w') as f: json.dump(self.compile_save_data(), f)

    def load_autosave(self):
        if os.path.exists(self.autosave_file):
            try:
                with open(self.autosave_file, 'r') as f:
                    if os.path.getsize(self.autosave_file) > 0:
                        self.parse_load_data(json.load(f))
            except json.JSONDecodeError:
                pass

    def compile_save_data(self):
        state = {'subject': self.subject_input.text(), 'lat': self.lat_input.text(), 'long': self.long_input.text(), 'uh_toggle': self.uh_checkbox.isChecked(), 'overrides': self.bom_overrides, 'nodes': [], 'spans': []}; node_map = {}
        for i, item in enumerate(self.scene.items()):
            if isinstance(item, (SmartPole, SmartHome)):
                item._temp_id = i; node_map[i] = item
                node_data = {'id': i, 'type': 'Pole' if isinstance(item, SmartPole) else 'Home', 'x': item.x(), 'y': item.y(), 'label_x': item.label.pos().x(), 'label_y': item.label.pos().y(), 'label_text': item.label.toPlainText()}
                if isinstance(item, SmartPole): node_data.update({'pole_type': item.pole_type, 'is_existing': item.is_existing, 'height': item.height, 'dtr_size': item.dtr_size, 'earth_count': item.earth_count, 'stay_count': item.stay_count, 'stay_type': getattr(item, 'stay_type', 'HT'), 'override_auto_stay': getattr(item, 'override_auto_stay', False)})
                state['nodes'].append(node_data)
        for item in self.scene.items():
            if isinstance(item, SmartSpan): state['spans'].append({'p1_id': item.p1._temp_id, 'p2_id': item.p2._temp_id, 'length': item.length, 'conductor': item.conductor, 'has_cg': item.has_cg, 'aug_type': item.aug_type, 'wire_count': item.wire_count, 'wire_size': item.wire_size, 'cable_size': getattr(item, 'cable_size', '10 SQMM'), 'consider_cable': getattr(item, 'consider_cable', False), 'phase': getattr(item, 'phase', '3 Phase'), 'is_service_drop': getattr(item, 'is_service_drop', False), 'label_x': item.label.pos().x(), 'label_y': item.label.pos().y(), 'label_text': item.label.toPlainText()})
        return state

    def parse_load_data(self, state):
        self.scene.clear()
        self.subject_input.setText(state.get('subject', ''))
        self.lat_input.setText(state.get('lat', ''))
        self.long_input.setText(state.get('long', ''))
        self.uh_checkbox.setChecked(state.get('uh_toggle', False)); self.bom_overrides = state.get('overrides', {}); node_map = {}
        for n_data in state.get('nodes', []):
            if n_data['type'] == 'Pole':
                pole = SmartPole(n_data['x'], n_data['y'], self.refresh_signal, n_data['pole_type'], n_data.get('is_existing', False));
                pole.height = n_data['height']; pole.dtr_size = n_data['dtr_size']; pole.earth_count = n_data['earth_count']; pole.stay_count = n_data['stay_count']; pole.stay_type = n_data.get('stay_type', 'HT'); pole.override_auto_stay = n_data.get('override_auto_stay', False); pole.update_visuals(); pole.label.setPos(n_data['label_x'], n_data['label_y']); pole.label.setPlainText(n_data['label_text']); self.scene.addItem(pole); node_map[n_data['id']] = pole
            else:
                home = SmartHome(n_data['x'], n_data['y'], self.refresh_signal);
                home.label.setPos(n_data['label_x'], n_data['label_y']); home.label.setPlainText(n_data['label_text']); self.scene.addItem(home); node_map[n_data['id']] = home
        for s_data in state.get('spans', []):
            p1 = node_map.get(s_data['p1_id']); p2 = node_map.get(s_data['p2_id'])
            if p1 and p2:
                span = SmartSpan(p1, p2); span.length = s_data['length']; span.conductor = s_data['conductor']; span.has_cg = s_data.get('has_cg', False); span.aug_type = s_data.get('aug_type', 'New'); span.wire_count = s_data.get('wire_count', '3'); span.wire_size = s_data.get('wire_size', '50SQMM'); span.cable_size = s_data.get('cable_size', '10 SQMM'); span.consider_cable = s_data.get('consider_cable', False); span.phase = s_data.get('phase', '3 Phase'); span.is_service_drop = s_data.get('is_service_drop', False); span.update_visuals(); span.label.setPos(s_data['label_x'], s_data['label_y']); span.label.setPlainText(s_data['label_text']); p1.connected_spans.append(span); p2.connected_spans.append(span); self.scene.addItem(span); self.scene.addItem(span.label)
        self.refresh_live_estimate()
        
    def closeEvent(self, event):
        with open(self.autosave_file, 'w') as f: json.dump(self.compile_save_data(), f)
        super().closeEvent(event)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_win = EstimateAppV9()
    main_win.show()
    sys.exit(app.exec())
