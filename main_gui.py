from bill_calculator import BillCalculatorApp
from theft_calculator import TheftCalculatorApp
from tariff_editor import TariffEditor

import os
import shutil
import threading
import time
import csv
import json
import subprocess
import sys
import webbrowser
import ctypes
import re
from collections import defaultdict
from difflib import SequenceMatcher
from datetime import datetime
import openpyxl
from PIL import Image, ImageTk, ImageDraw

try:
    from rapidfuzz import fuzz as rapidfuzz_fuzz
except Exception:
    rapidfuzz_fuzz = None

import tkinter as tk
from tkinter import messagebox, filedialog, Menu, Toplevel, Listbox, ttk
from tkinter import CENTER, NE, NW, SW, SE, TOP, BOTTOM, LEFT, RIGHT, BOTH, X, Y, END
import ttkbootstrap as tb

import config
import database
import utils
from low_consumption import LowConsumptionVerifier
import documentation

# --- Global State Variables ---
img_tk = None
img = None
img_original = None
zoom_scale = 1.0
pan_x = 0
pan_y = 0
drag_start_x = 0
drag_start_y = 0
additional_folders = utils.load_additional_folders()
indexing_active = False
current_search_data = {}  
preview_references = [] 
preview_canvas_widgets = []

LIGHT_THEME = "cosmo"
DARK_THEME = "darkly"


def set_windows_app_id():
    """Set an explicit AppUserModelID so Windows taskbar uses this app identity."""
    if os.name != "nt":
        return

    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("SpotImageViewer.App")
    except Exception:
        pass


def ensure_app_icons():
    """Create app icon files once and return (ico_path, png_path)."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    assets_dir = os.path.join(base_dir, "assets")
    os.makedirs(assets_dir, exist_ok=True)
    png_path = os.path.join(assets_dir, "spot_icon.png")
    ico_path = os.path.join(assets_dir, "spot_icon.ico")

    if os.path.exists(png_path) and os.path.exists(ico_path):
        return ico_path, png_path

    size = 128
    img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)

    # Rounded-square tile and a simple camera glyph for clear readability.
    draw.rounded_rectangle((8, 8, size - 8, size - 8), radius=24, fill=(14, 86, 237, 255))
    draw.rounded_rectangle((28, 42, size - 28, size - 28), radius=14, fill=(255, 255, 255, 255))
    draw.ellipse((50, 50, 78, 78), fill=(14, 86, 237, 255))
    draw.rectangle((40, 34, 58, 44), fill=(255, 255, 255, 255))
    draw.ellipse((44, 36, 54, 46), fill=(14, 86, 237, 255))

    img.save(png_path, format="PNG")
    img.save(ico_path, format="ICO", sizes=[(16, 16), (24, 24), (32, 32), (48, 48), (64, 64), (128, 128)])
    return ico_path, png_path


def is_dark_mode_active():
    try:
        return root.style.theme_use() == DARK_THEME
    except Exception:
        return False


def refresh_non_ttk_widget_colors():
    """Update all plain-tk widget colours to match the current ttkbootstrap theme."""
    dark = is_dark_mode_active()

    if dark:
        listbox_bg   = "#2c3136"
        listbox_fg   = "#f8f9fa"
        select_bg    = "#4c6ef5"
        canvas_bg    = "#212529"
        scroller_bg  = "#2c3136"
        preview_bg   = "#2c3136"
        notes_bg     = "#3a3024"
        notes_fg     = "#f8f9fa"
        folders_bg   = "#1e3044"
        menu_bg      = "#2c3136"
        menu_fg      = "#f8f9fa"
        menu_act_bg  = "#4c6ef5"
    else:
        listbox_bg   = "#f8f9fa"
        listbox_fg   = "#212529"
        select_bg    = "#0d6efd"
        canvas_bg    = "#e9ecef"
        scroller_bg  = "#ffffff"
        preview_bg   = "#f0f0f0"
        notes_bg     = "#fff3cd"
        notes_fg     = "#212529"
        folders_bg   = "#e3f2fd"
        menu_bg      = "#f8f9fa"
        menu_fg      = "#212529"
        menu_act_bg  = "#0d6efd"

    select_fg   = "#ffffff"
    menu_act_fg = "#ffffff"

    # --- Listboxes ---
    for name, bg in [('listbox_dates', listbox_bg), ('folder_listbox', folders_bg)]:
        w = globals().get(name)
        if w:
            w.config(bg=bg, fg=listbox_fg, selectbackground=select_bg, selectforeground=select_fg)

    # --- Canvases ---
    w = globals().get('canvas')
    if w:
        w.config(bg=canvas_bg)

    w = globals().get('preview_canvas_scroller')
    if w:
        w.config(bg=scroller_bg)

    for cw in preview_canvas_widgets:
        try:
            if cw.winfo_exists():
                cw.config(bg=preview_bg)
        except Exception:
            continue

    # --- Text widget (notes) ---
    w = globals().get('txt_remarks')
    if w:
        w.config(bg=notes_bg, fg=notes_fg, insertbackground=notes_fg)

    # --- Frames that carry a fixed bootstyle ---
    frame_style = "dark" if dark else "light"
    for name in ('welcome_frame', 'top_f'):
        w = globals().get(name)
        if w:
            try:
                w.config(bootstyle=frame_style)
            except Exception:
                pass

    # --- Menu bar ---
    for name in ('mb', 'fm', 'bm', 'nm', 'vm', 'tm', 'hm'):
        w = globals().get(name)
        if w:
            try:
                w.config(bg=menu_bg, fg=menu_fg,
                         activebackground=menu_act_bg, activeforeground=menu_act_fg)
            except Exception:
                pass

    # --- Status-bar toggle buttons ---
    for btn_name, pane_name in [('btn_notes', 'notes_pane'), ('btn_folders', 'folder_pane')]:
        btn = globals().get(btn_name)
        pane = globals().get(pane_name)
        if btn and pane:
            btn.config(bootstyle=_toggle_btn_style(pane.winfo_ismapped()))


def update_theme_toggle_button_text():
    if 'btn_theme_toggle' not in globals():
        return

    if is_dark_mode_active():
        btn_theme_toggle.config(text="Light Mode", bootstyle="primary")
    else:
        btn_theme_toggle.config(text="Dark Mode", bootstyle="primary")


def toggle_theme():
    next_theme = DARK_THEME if not is_dark_mode_active() else LIGHT_THEME
    root.style.theme_use(next_theme)
    refresh_non_ttk_widget_colors()
    update_theme_toggle_button_text()


def _toggle_btn_style(active):
    """Return the right bootstyle for status-bar toggle buttons."""
    if is_dark_mode_active():
        return "light" if active else "light-outline"
    return "dark" if active else "dark-outline"


def launch_tool(script_name, tool_title):
    try:
        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)
        if not os.path.exists(script_path):
            messagebox.showerror("Tool Not Found", f"{tool_title} file was not found:\n{script_path}")
            return

        subprocess.Popen([sys.executable, script_path], cwd=os.path.dirname(script_path))
    except Exception as e:
        messagebox.showerror("Launch Error", f"Could not open {tool_title}.\n\n{e}")

def add_network_folder():
    try:
        folder = filedialog.askdirectory(title="Select Network Folder")
        if folder and folder not in additional_folders:
            additional_folders.append(folder)
            utils.save_additional_folders(additional_folders)
            update_folder_list_ui() 
            run_single_check()
            if messagebox.askyesno("Index Now?", "Folder added. Scan it for images now?"):
                start_indexing_process()
    except Exception as e:
        messagebox.showerror("Error", str(e))

def remove_network_folder():
    try:
        selection = folder_listbox.curselection()
        if selection:
            idx = selection[0]
            if idx == 0:
                messagebox.showwarning("Warning", "Cannot remove the default Image folder.")
                return
            
            folder_idx = idx - 1
            if 0 <= folder_idx < len(additional_folders):
                del additional_folders[folder_idx]
                utils.save_additional_folders(additional_folders)
                update_folder_list_ui()
                run_single_check() 
                messagebox.showinfo("Info", "Folder removed. Images remain until 'Reload Images' is clicked.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def index_images_thread(progress_callback, finish_callback):
    global indexing_active
    indexing_active = True
    
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL;")
        cursor.execute("PRAGMA synchronous=OFF;")

        folders = [config.IMAGE_FOLDER] + additional_folders
        batch_data = []
        BATCH_SIZE = 5000 
        total_inserted = 0

        cursor.execute("DELETE FROM images")
        cursor.execute("DELETE FROM directories") # Also clear directories
        conn.commit()
        
        start_time = time.time()

        for folder in folders:
            if not os.path.exists(folder):
                continue
            
            for root_dir, dirs, files in os.walk(folder):
                # Get the directory ID for the current root_dir
                cursor.execute("INSERT OR IGNORE INTO directories (dir_path) VALUES (?)", (root_dir,))
                cursor.execute("SELECT id FROM directories WHERE dir_path = ?", (root_dir,))
                dir_id = cursor.fetchone()[0]

                for filename in files:
                    try:
                        if len(filename) < 20: continue
                        if not filename[:8].isdigit(): continue
                        
                        date_orig = filename[:8]
                        try:
                            dt = datetime.strptime(date_orig, "%d%m%Y")
                            date_iso = dt.strftime("%Y-%m-%d")
                        except:
                            continue 

                        mru = filename[8:16]
                        cid = filename[16:25]

                        batch_data.append((cid, date_orig, date_iso, mru, filename, dir_id))

                        if len(batch_data) >= BATCH_SIZE:
                            cursor.executemany("INSERT OR IGNORE INTO images (consumer_id, date_original, date_iso, mru, filename, dir_id) VALUES (?,?,?,?,?,?)", batch_data)
                            conn.commit()
                            total_inserted += len(batch_data)
                            batch_data = []
                            elapsed = int(time.time() - start_time)
                            progress_callback(total_inserted, elapsed)
                            
                    except Exception:
                        continue

        if batch_data:
            cursor.executemany("INSERT OR IGNORE INTO images (consumer_id, date_original, date_iso, mru, filename, dir_id) VALUES (?,?,?,?,?,?)", batch_data)
            conn.commit()
            total_inserted += len(batch_data)
        
        cursor.execute("ANALYZE;") 
        
        # Optimize and compress database
        cursor.execute("PRAGMA wal_checkpoint(TRUNCATE);")
        cursor.execute("VACUUM;")
        
        conn.close()

    except Exception as e:
        print(f"Indexing error: {e}")
    finally:
        indexing_active = False
        finish_callback(total_inserted)

def add_new_note_option():
    def save_opt():
        new_opt = entry_opt.get().strip()
        if new_opt:
            utils.add_note_option(new_opt)
            # Refresh the note options in the UI
            global note_options
            note_options = utils.load_note_options()
            combo_notes['values'] = note_options
            note_var.set(new_opt)
            pop.destroy()

    pop = Toplevel()
    pop.title("Add Note Option")
    pop.geometry("300x150")
    
    tb.Label(pop, text="Enter New Option:").pack(pady=10)
    entry_opt = tb.Entry(pop)
    entry_opt.pack(pady=5, padx=10, fill=X)
    entry_opt.focus()
    entry_opt.bind("<Return>", lambda e: save_opt())
    tb.Button(pop, text="Add", command=save_opt, bootstyle="success").pack(pady=10)

def export_notes_csv():
    try:
        notes = utils.load_all_notes()
        if not notes:
            messagebox.showinfo("Info", "No notes to export.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".csv", 
            filetypes=[("CSV", "*.csv")],
            title="Export Notes"
        )
        if not path:
            return

        prog = Toplevel()
        prog.title("Exporting Notes")
        prog.geometry("350x100")
        tb.Label(prog, text="Exporting notes, please wait...").pack(padx=20, pady=10)
        prog.update()

        def worker(export_path, data):
            try:
                with open(export_path, "w", newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Consumer ID", "Note", "Remarks"])
                    for cid, d in data.items():
                        writer.writerow([cid, d.get('note',''), d.get('remarks','')])
                root.after(0, lambda: messagebox.showinfo("Success", f"Exported to {export_path}"))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror("Error", f"Export failed: {e}"))
            finally:
                try:
                    root.after(0, prog.destroy)
                except: pass

        threading.Thread(target=worker, args=(path, notes), daemon=True).start()

    except Exception as e:
        messagebox.showerror("Error", f"Export failed: {e}")

def perform_backup():
    def on_date_select():
        date_str = date_entry.get().strip()
        try:
            dt = datetime.strptime(date_str, "%d-%m-%Y")
            iso_limit = dt.strftime("%Y-%m-%d")
            top.withdraw() 
            backup_dir = filedialog.askdirectory(title="Select Backup Destination")
            top.destroy()
            
            if not backup_dir: return
            run_backup_thread(iso_limit, backup_dir)
        except ValueError:
            messagebox.showerror("Error", "Format: DD-MM-YYYY", parent=top)

    top = Toplevel()
    top.title("Backup Images")
    top.geometry("300x150")
    
    tb.Label(top, text="Backup images UP TO date (DD-MM-YYYY):").pack(padx=10, pady=10)
    date_entry = tb.Entry(top)
    date_entry.pack(padx=10)
    tb.Button(top, text="Start Backup", command=on_date_select).pack(pady=10)

def run_backup_thread(iso_date_limit, target_folder):
    prog_win = Toplevel()
    prog_win.title("Backing Up")
    prog_win.geometry("400x150")
    lbl = tb.Label(prog_win, text="Querying database...")
    lbl.pack(padx=20, pady=10)
    pb = ttk.Progressbar(prog_win, length=300, mode="determinate")
    pb.pack(padx=20, pady=10)
    
    def worker():
        try:
            conn = database.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT d.dir_path, i.filename 
                FROM images i
                JOIN directories d ON i.dir_id = d.id
                WHERE i.date_iso <= ?
            """, (iso_date_limit,))
            rows = cursor.fetchall()
            conn.close()
            
            total = len(rows)
            if total == 0:
                root.after(0, lambda: messagebox.showinfo("Info", "No images found up to that date."))
                root.after(0, prog_win.destroy)
                return

            root.after(0, lambda: lbl.config(text=f"Moving {total} images..."))
            root.after(0, lambda: pb.config(maximum=total))
            
            count = 0
            moved = 0
            
            for dir_path, filename in rows:
                if dir_path == config.IMAGE_FOLDER:
                    try:
                        full_path = os.path.join(dir_path, filename)
                        if os.path.exists(full_path):
                            dst = os.path.join(target_folder, filename)
                            shutil.move(full_path, dst)
                            moved += 1
                    except: pass
                count += 1
                if count % 100 == 0:
                    root.after(0, lambda c=count: pb.config(value=c))
            
            root.after(0, prog_win.destroy)
            root.after(0, lambda: messagebox.showinfo("Backup Complete", f"Moved {moved} images."))
            root.after(0, lambda: ask_add_backup(target_folder))
        except Exception as e:
             root.after(0, lambda: messagebox.showerror("Backup Error", str(e)))
             try: root.after(0, prog_win.destroy)
             except: pass

    threading.Thread(target=worker, daemon=True).start()

def ask_add_backup(target_folder):
    if messagebox.askyesno("Update Index", "Do you want to add this backup folder to the network list and re-index?"):
        if target_folder not in additional_folders:
            additional_folders.append(target_folder)
            utils.save_additional_folders(additional_folders)
            update_folder_list_ui()
        start_indexing_process()

def start_indexing_process():
    if indexing_active: return
    btn_reload.config(state="disabled")
    status_label.config(text="Indexing...", bootstyle="warning")
    progress_bar.pack(side=RIGHT, padx=10)
    progress_bar.start(10)
    
    def on_prog(c, t):
        root.after(0, lambda: status_label.config(text=f"Indexed: {c} ({t}s)"))
    
    def on_done(total):
        root.after(0, lambda: finish_indexing(total))
        
    threading.Thread(target=index_images_thread, args=(on_prog, on_done), daemon=True).start()

def finish_indexing(total):
    progress_bar.stop()
    progress_bar.pack_forget()
    status_label.config(text=f"Total Images: {total}", bootstyle="success")
    btn_reload.config(state="normal")
    messagebox.showinfo("Done", f"Indexing Complete.\nTotal Images: {total}")


def _safe_text(value):
    if value is None:
        return "-"
    text = str(value).strip()
    return text if text else "-"


def _format_consumer_details(cid, mru):
    profile = utils.get_consumer_profile(cid) or {}
    lines = [
        f"Consumer ID: {_safe_text(cid)}",
        f"Meter No: {_safe_text(profile.get('meter_no'))}",
        f"Name: {_safe_text(profile.get('name'))}",
        f"Address: {_safe_text(profile.get('address'))}",
        f"Mobile Number: {_safe_text(profile.get('mobile_number'))}",
        f"Contractual Load: {_safe_text(profile.get('contractual_load'))}",
        f"Class: {_safe_text(profile.get('class'))}",
        f"MRU: {_safe_text(mru)}",
    ]
    return "\n".join(lines)


def _open_consumer_from_selection(cid):
    entry_consumer_id.delete(0, tk.END)
    entry_consumer_id.insert(0, cid)
    search_consumer()


def show_search_results_selector(results, title):
    if not results:
        return

    picker = Toplevel(root)
    picker.title(title)
    picker.geometry("1000x420")
    picker.transient(root)
    picker.grab_set()

    tb.Label(
        picker,
        text=f"{len(results)} result(s). Select one consumer:",
        font=("Segoe UI", 10, "bold")
    ).pack(anchor=NW, padx=10, pady=(10, 4))

    grid_frame = tb.Frame(picker, padding=8)
    grid_frame.pack(fill=BOTH, expand=True)

    cols = ("consumer_id", "meter_no", "name", "mobile_number", "address")
    tree = ttk.Treeview(grid_frame, columns=cols, show="headings", height=12)
    tree.heading("consumer_id", text="Consumer ID")
    tree.heading("meter_no", text="Meter No")
    tree.heading("name", text="Name")
    tree.heading("mobile_number", text="Mobile")
    tree.heading("address", text="Address")

    tree.column("consumer_id", width=120, anchor="w")
    tree.column("meter_no", width=120, anchor="w")
    tree.column("name", width=210, anchor="w")
    tree.column("mobile_number", width=120, anchor="w")
    tree.column("address", width=390, anchor="w")

    yscroll = tb.Scrollbar(grid_frame, orient="vertical", command=tree.yview)
    xscroll = tb.Scrollbar(grid_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

    tree.grid(row=0, column=0, sticky="nsew")
    yscroll.grid(row=0, column=1, sticky="ns")
    xscroll.grid(row=1, column=0, sticky="ew")
    grid_frame.grid_rowconfigure(0, weight=1)
    grid_frame.grid_columnconfigure(0, weight=1)

    for item in results:
        tree.insert(
            "",
            "end",
            values=(
                _safe_text(item.get("consumer_id")),
                _safe_text(item.get("meter_no")),
                _safe_text(item.get("name")),
                _safe_text(item.get("mobile_number")),
                _safe_text(item.get("address")),
            )
        )

    btns = tb.Frame(picker, padding=8)
    btns.pack(fill=X)

    def choose_and_open():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select one row.", parent=picker)
            return
        values = tree.item(selected[0], "values")
        chosen_cid = values[0] if values else ""
        if not chosen_cid:
            return
        picker.destroy()
        _open_consumer_from_selection(chosen_cid)

    tb.Button(btns, text="Open Selected", bootstyle="primary", command=choose_and_open).pack(side=LEFT)
    tb.Button(btns, text="Cancel", bootstyle="secondary", command=picker.destroy).pack(side=LEFT, padx=6)

    tree.bind("<Double-1>", lambda e: choose_and_open())

def search_consumer():
    cid = entry_consumer_id.get().strip()
    if not cid.isdigit() or len(cid) != 9:
        messagebox.showwarning("Error", "Invalid Consumer ID (9 digits required)")
        return
    
    welcome_frame.pack_forget()
    listbox_dates.delete(0, tk.END)
    raise_preview_pane()
    hide_buttons() 
    
    try:
        conn = database.get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT i.date_original, i.mru, d.dir_path, i.filename 
            FROM images i 
            JOIN directories d ON i.dir_id = d.id 
            WHERE i.consumer_id = ? 
            ORDER BY i.date_iso DESC
        """, (cid,))
        rows = cur.fetchall()
        conn.close()
        
        if not rows:
            label_consumer_details.config(text=f"Consumer ID: {cid}\nNo images found.", bootstyle="danger")
            label_latest_date.config(text="")
            clear_previews()
            return

        mru = rows[0][1]
        
        global current_search_data
        current_search_data = {}
        dates_ordered = []
        
        for date, _, dir_path, filename in rows:
            path = os.path.join(dir_path, filename)
            if date not in current_search_data:
                current_search_data[date] = []
                dates_ordered.append(date)
            current_search_data[date].append(path)
            
        count = len(rows)
        label_consumer_details.config(text=_format_consumer_details(cid, mru), bootstyle="primary")
        label_total_images.config(text=f"Images Found: {count}")
        
        latest = dates_ordered[0]
        pretty_latest = f"{latest[:2]}-{latest[2:4]}-{latest[4:]}"
        label_latest_date.config(text=f"Latest Image: {pretty_latest}", bootstyle="success")

        for d in dates_ordered:
            pretty = f"{d[:2]}-{d[2:4]}-{d[4:]}"
            listbox_dates.insert(tk.END, pretty)
            
        show_dynamic_previews(dates_ordered)
        load_consumer_note(cid)
        utils.save_search_history("consumer_ids", cid)

    except Exception as e:
        messagebox.showerror("DB Error", f"Search failed: {e}")

def clear_previews():
    global preview_canvas_widgets
    for widget in scrollable_frame.winfo_children():
        widget.destroy()
    global preview_references
    preview_references = []
    preview_canvas_widgets = []

def show_dynamic_previews(dates):
    preview_bg = "#2c3136" if is_dark_mode_active() else "#f0f0f0"
    clear_previews()
    for i, date in enumerate(dates):
        path = current_search_data[date][0]
        if os.path.exists(path):
            try:
                im = Image.open(path)
                im.thumbnail((180, 180)) 
                ph = ImageTk.PhotoImage(im)
                preview_references.append(ph)
                
                cols = 4 
                pf = tb.Labelframe(scrollable_frame, text=f"{date[:2]}-{date[2:4]}-{date[4:]}", padding=5, bootstyle="info")
                pf.grid(row=i//cols, column=i%cols, padx=10, pady=10, sticky="nsew")
                
                pc = tk.Canvas(pf, width=180, height=180, bg=preview_bg, highlightthickness=0)
                pc.pack()
                pc.create_image(90, 90, image=ph)
                preview_canvas_widgets.append(pc)
                
                def mk_click(p): return lambda e: load_image_to_canvas(p)
                pc.bind("<Button-1>", mk_click(path))
                pf.bind("<Button-1>", mk_click(path)) 
            except: pass

def on_date_select(event):
    idx = listbox_dates.curselection()
    if not idx: return
    txt = listbox_dates.get(idx[0]) 
    raw = txt.replace("-", "") 
    
    if raw in current_search_data:
        paths = current_search_data[raw]
        best_path = paths[0]
        for p in paths:
            if config.IMAGE_FOLDER in p:
                best_path = p
                break
        load_image_to_canvas(best_path)

def load_image_to_canvas(path):
    raise_canvas_pane()
    global img_original, zoom_scale, img, img_tk, pan_x, pan_y
    if not os.path.exists(path):
        messagebox.showerror("Error", "Image file not accessible (Offline?)")
        return
    try:
        img_original = Image.open(path)
        zoom_scale = 1.0
        pan_x = 0
        pan_y = 0
        render_image()
        show_buttons() 
    except Exception as e:
        messagebox.showerror("Error", str(e))

def raise_preview_pane():
    preview_container.tkraise()
    btn_show_previews.pack_forget()

def raise_canvas_pane():
    canvas_container.tkraise()
    btn_show_previews.pack(side=TOP, anchor=NE, padx=5, pady=2) 

def render_image():
    global img, img_tk, pan_x, pan_y
    if not img_original: return
    
    cw = canvas.winfo_width()
    ch = canvas.winfo_height()
    if cw < 50: cw = 800
    if ch < 50: ch = 600
    
    w, h = img_original.size
    
    if zoom_scale == 1.0:
        ratio = min(cw/w, ch/h)
        nw, nh = int(w*ratio), int(h*ratio)
        pan_x = 0
        pan_y = 0
    else:
        nw, nh = int(w*zoom_scale), int(h*zoom_scale)
        
    img = img_original.resize((nw, nh), Image.Resampling.LANCZOS)
    img_tk = ImageTk.PhotoImage(img)
    canvas.delete("all")
    canvas.create_image((cw//2) + pan_x, (ch//2) + pan_y, anchor=CENTER, image=img_tk)

def zoom(factor):
    global zoom_scale
    zoom_scale *= factor
    render_image()

def start_pan(event):
    global drag_start_x, drag_start_y
    if zoom_scale == 1.0: return
    drag_start_x = event.x
    drag_start_y = event.y
    canvas.config(cursor="fleur")

def do_pan(event):
    global pan_x, pan_y, drag_start_x, drag_start_y
    if zoom_scale == 1.0: return
    dx = event.x - drag_start_x
    dy = event.y - drag_start_y
    pan_x += dx
    pan_y += dy
    drag_start_x = event.x
    drag_start_y = event.y
    canvas.move("all", dx, dy)

def end_pan(event):
    if zoom_scale == 1.0: return
    canvas.config(cursor="")

def on_image_mousewheel(event):
    if not img_original: return
    if event.delta > 0:
        zoom(1.2)
    elif event.delta < 0:
        zoom(0.8)

def _bind_image_mousewheel(event):
    canvas.bind_all("<MouseWheel>", on_image_mousewheel)

def _unbind_image_mousewheel(event):
    canvas.unbind_all("<MouseWheel>")

def print_image():
    if not img_original: return
    try:
        tmp = "temp_print.png"
        img_original.save(tmp)
        os.startfile(tmp, "print")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def save_image():
    if not img_original: return
    
    original_filename = os.path.basename(img_original.filename)
    
    path = filedialog.asksaveasfilename(
        initialfile=original_filename,
        defaultextension=".png",
        filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg"), ("All files", "*.*")]
    )
    if not path: return

    try:
        img_copy = img_original.copy()
    except:
        img_copy = img_original

    prog = Toplevel()
    prog.title("Saving Image")
    prog.geometry("320x90")
    tb.Label(prog, text="Saving image, please wait...").pack(padx=20, pady=10)
    prog.update()

    def worker(im, out_path):
        try:
            im.save(out_path)
            root.after(0, lambda: messagebox.showinfo("Saved", "Image Saved."))
        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error", f"Save failed: {e}"))
        finally:
            try:
                root.after(0, prog.destroy)
            except:
                pass

    threading.Thread(target=worker, args=(img_copy, path), daemon=True).start()

def save_all_images():
    cid = entry_consumer_id.get().strip()
    if not cid or not current_search_data: return 
    
    dest = filedialog.askdirectory(title="Select Destination Folder")
    if not dest: return
    
    target_dir = os.path.join(dest, cid)
    os.makedirs(target_dir, exist_ok=True)
    
    prog = Toplevel()
    prog.title("Saving Images")
    prog.geometry("360x110")
    lbl = tb.Label(prog, text="Saving images, please wait...")
    lbl.pack(padx=20, pady=10)
    pbar = ttk.Progressbar(prog, length=300, mode="indeterminate")
    pbar.pack(padx=10, pady=5)
    pbar.start(10)
    prog.update()
    
    def worker(data, out_dir):
        count = 0
        try:
            for date, paths in data.items():
                for p in paths:
                    if os.path.exists(p):
                        original_filename = os.path.basename(p)
                        try:
                            shutil.copy2(p, os.path.join(out_dir, original_filename))
                            count += 1
                        except Exception as copy_e:
                            print(f"Could not copy {p}: {copy_e}")
            root.after(0, lambda: messagebox.showinfo("Success", f"Saved {count} images to {out_dir}"))
        except Exception as e:
            root.after(0, lambda: messagebox.showerror("Error", f"Save all failed: {e}"))
        finally:
            try:
                root.after(0, pbar.stop)
                root.after(0, prog.destroy)
            except:
                pass

    threading.Thread(target=worker, args=(current_search_data, target_dir), daemon=True).start()

def toggle_notes():
    if notes_pane.winfo_ismapped():
        notes_pane.pack_forget()
        btn_notes.config(bootstyle=_toggle_btn_style(False))
    else:
        notes_pane.pack(side=RIGHT, fill=Y, padx=5, pady=5)
        btn_notes.config(bootstyle=_toggle_btn_style(True))
        folder_pane.pack_forget()
        btn_folders.config(bootstyle=_toggle_btn_style(False))

def toggle_folders():
    if folder_pane.winfo_ismapped():
        folder_pane.pack_forget()
        btn_folders.config(bootstyle=_toggle_btn_style(False))
    else:
        folder_pane.pack(side=RIGHT, fill=Y, padx=5, pady=5)
        btn_folders.config(bootstyle=_toggle_btn_style(True))
        notes_pane.pack_forget()
        btn_notes.config(bootstyle=_toggle_btn_style(False))

def load_consumer_note(cid):
    try:
        all_n = utils.load_all_notes()
        data = all_n.get(cid, {})
        n_txt = data.get('note', '')
        r_txt = data.get('remarks', '')
        
        if n_txt:
            label_prev_note.config(text=f"Note: {n_txt}\n{r_txt}")
            note_var.set(n_txt)
            txt_remarks.delete("1.0", tk.END)
            txt_remarks.insert("1.0", r_txt)
            if not notes_pane.winfo_ismapped(): toggle_notes()
        else:
            label_prev_note.config(text="No prior notes.")
            note_var.set(note_options[0] if note_options else "")
            txt_remarks.delete("1.0", tk.END)
    except: pass

def save_current_note():
    cid = entry_consumer_id.get().strip()
    if not cid: return
    note = note_var.get()
    remarks = txt_remarks.get("1.0", tk.END).strip()
    utils.save_note(cid, note, remarks)
    messagebox.showinfo("Saved", "Note updated.")
    load_consumer_note(cid)

def delete_current_note():
    cid = entry_consumer_id.get().strip()
    if not cid: return
    if messagebox.askyesno("Confirm", "Are you sure you want to delete the note for this consumer?"):
        utils.delete_note(cid)
        load_consumer_note(cid)
        messagebox.showinfo("Deleted", "Note removed.")

def update_folder_list_ui():
    folder_listbox.delete(0, tk.END)
    folder_listbox.insert(tk.END, config.IMAGE_FOLDER)
    for f in additional_folders:
        folder_listbox.insert(tk.END, f)
    run_single_check()

def run_single_check():
    threading.Thread(target=_check_paths_thread, daemon=True).start()

def _check_paths_thread():
    folders_to_check = [config.IMAGE_FOLDER] + additional_folders
    results = []
    for f in folders_to_check:
        results.append((f, os.path.exists(f)))
    if 'root' in globals():
        root.after(0, lambda: apply_network_colors(results))

def apply_network_colors(results):
    try:
        dark = is_dark_mode_active()
        for idx, (path, is_online) in enumerate(results):
            if is_online:
                color = "#00e676" if dark else "green"
            else:
                color = "#ff5252" if dark else "red"
            folder_listbox.itemconfig(idx, foreground=color)
    except: pass



def update_meter_search_state():
    if database.has_meter_data():
        entry_meter_number.config(state="normal")
        meter_button.config(state="normal", bootstyle="primary")
        entry_name.config(state="normal")
        entry_mobile.config(state="normal")
        btn_search_name.config(state="normal", bootstyle="primary")
        btn_search_mobile.config(state="normal", bootstyle="primary")
        lbl_consumer_hint.grid_remove()
        try:
            tm.entryconfig("Fuzzy Lookup", state="normal")
        except Exception:
            pass
    else:
        entry_meter_number.config(state="disabled")
        meter_button.config(state="disabled", bootstyle="secondary")
        entry_name.config(state="disabled")
        entry_mobile.config(state="disabled")
        btn_search_name.config(state="disabled", bootstyle="secondary")
        btn_search_mobile.config(state="disabled", bootstyle="secondary")
        lbl_consumer_hint.grid()
        try:
            tm.entryconfig("Fuzzy Lookup", state="disabled")
        except Exception:
            pass



def show_history(event, key, widget):
    try:
        items = utils.load_search_history(key)[::-1] # Reverse for recent first
        if not items: return
        
        if hasattr(widget, 'history_popup') and widget.history_popup.winfo_exists():
            return

        top = Toplevel()
        widget.history_popup = top
        top.overrideredirect(True)
        top.geometry(f"+{widget.winfo_rootx()}+{widget.winfo_rooty()+widget.winfo_height()}")
        _dark = is_dark_mode_active()
        _lb_bg  = "#2c3136" if _dark else "#ffffff"
        _lb_fg  = "#f8f9fa" if _dark else "#212529"
        _lb_sel = "#4c6ef5" if _dark else "#0d6efd"
        if _dark:
            top.config(bg=_lb_bg)
        lb = Listbox(top, height=10, font=("Arial", 11),
                     bg=_lb_bg, fg=_lb_fg,
                     selectbackground=_lb_sel, selectforeground="#ffffff")
        lb.pack(fill=BOTH, expand=True)
        for i in items: lb.insert(tk.END, i)
        
        def pick(e):
            if lb.curselection():
                val = lb.get(lb.curselection())
                widget.delete(0, tk.END)
                widget.insert(0, val)
                top.destroy()
                if key == "consumer_ids":
                    search_consumer()
                elif key == "meter_numbers":
                    search_meter()
                elif key == "consumer_names":
                    search_name()
                else:
                    search_mobile()
                
        lb.bind("<Button-1>", pick)
        lb.bind("<FocusOut>", lambda e: top.destroy())
        lb.focus_set()
    except: pass

def search_meter():
    m = entry_meter_number.get().strip()
    if not m: return

    cid = utils.get_consumer_by_meter(m)
    if cid:
        entry_consumer_id.delete(0, tk.END)
        entry_consumer_id.insert(0, cid)
        search_consumer()
        utils.save_search_history("meter_numbers", m)
    else:
        messagebox.showinfo("Not Found", "Meter Number not found.")


def search_name():
    name_query = entry_name.get().strip()
    if len(name_query) < 3:
        messagebox.showwarning("Error", "Enter at least 3 characters for Name search.")
        return

    results = utils.search_consumers_by_name(name_query, limit=300)
    if not results:
        messagebox.showinfo("Not Found", "No consumer found for this name.")
        return

    utils.save_search_history("consumer_names", name_query)
    if len(results) == 1:
        _open_consumer_from_selection(results[0].get("consumer_id", ""))
        return

    show_search_results_selector(results, f"Name Search: {name_query}")


def search_mobile():
    raw_mobile = entry_mobile.get().strip()
    normalized = re.sub(r"\D", "", raw_mobile)
    if len(normalized) == 12 and normalized.startswith("91"):
        normalized = normalized[2:]

    if len(normalized) != 10:
        messagebox.showwarning("Error", "Mobile number must be 10 digits.")
        return

    entry_mobile.delete(0, tk.END)
    entry_mobile.insert(0, normalized)

    results = utils.search_consumers_by_mobile(normalized, limit=300)
    if not results:
        messagebox.showinfo("Not Found", "No consumer found for this mobile number.")
        return

    utils.save_search_history("mobile_numbers", normalized)
    if len(results) == 1:
        _open_consumer_from_selection(results[0].get("consumer_id", ""))
        return

    show_search_results_selector(results, f"Mobile Search: {normalized}")

def generate_consumer_data_template():
    try:
        save_path = filedialog.asksaveasfilename(
            title="Save Consumer Data Template",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="consumer_data_template.xlsx"
        )
        if not save_path:
            return

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "ConsumerData"

        headers = [
            "CONSUMER ID",
            "METER NO",
            "NAME",
            "ADDRESS",
            "MOBILE NUMBER",
            "CONTRACTUAL LOAD",
            "CLASS",
        ]
        sheet.append(headers)

        # Keep one sample row so users can directly copy/paste from row 2.
        sheet.append(["", "", "", "", "", "", ""])
        sheet.freeze_panes = "A2"

        widths = [18, 16, 28, 36, 18, 18, 14]
        for idx, width in enumerate(widths, start=1):
            col = openpyxl.utils.get_column_letter(idx)
            sheet.column_dimensions[col].width = width

        wb.save(save_path)
        messagebox.showinfo("Template Generated", f"Template created successfully:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate template.\n{e}")


def import_consumer_data_threaded():
    utils.console_log("FUNCTION STARTED: import_consumer_data_threaded")

    def open_file_picker():
        utils.console_log("Step 1: Opening File Dialog (askopenfilename)...")
        try:
            path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
            utils.console_log(f"Step 1: Dialog Result -> {path}")
        except Exception as e:
            utils.console_log(f"!!! ERROR in File Dialog: {e}")
            return

        if not path:
            utils.console_log("Action Cancelled by user.")
            return

        utils.console_log("Step 2: Preparing UI for update...")
        btn_update_consumer.config(state="disabled", text="Updating...")
        progress_bar.pack(side=RIGHT, padx=10)
        progress_bar.start(10)

        threading.Thread(target=worker, args=(path,), daemon=True).start()

    def worker(file_path):
        utils.console_log(f"WORKER: Reading file {file_path} with openpyxl")
        try:
            d = {}
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            header_map = {
                "consumer id": "consumer_id",
                "consumerid": "consumer_id",
                "meter no": "meter_no",
                "meterno": "meter_no",
                "name": "name",
                "address": "address",
                "mobile number": "mobile_number",
                "mobilenumber": "mobile_number",
                "mobile": "mobile_number",
                "contractual load": "contractual_load",
                "contractualload": "contractual_load",
                "class": "class",
            }

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Excel sheet is empty.")

            first_row = rows[0]
            index_map = {}
            for idx, val in enumerate(first_row):
                if val is None:
                    continue
                key = str(val).strip().lower().replace("_", " ")
                key = " ".join(key.split())
                key_compact = key.replace(" ", "")
                mapped = header_map.get(key) or header_map.get(key_compact)
                if mapped:
                    index_map[mapped] = idx

            has_headers = "consumer_id" in index_map and "meter_no" in index_map
            data_rows = rows[1:] if has_headers else rows

            for row in data_rows:
                if not row:
                    continue

                def get_value(field, fallback_idx):
                    idx = index_map.get(field, fallback_idx)
                    if idx is None or idx >= len(row):
                        return ""
                    val = row[idx]
                    if val is None:
                        return ""
                    if isinstance(val, float) and val.is_integer():
                        return str(int(val)).strip()
                    return str(val).strip()

                cid = get_value("consumer_id", 0)
                meter_no = get_value("meter_no", 1)
                if not cid or not meter_no:
                    continue

                mobile = re.sub(r"\D", "", get_value("mobile_number", 4))
                if len(mobile) == 12 and mobile.startswith("91"):
                    mobile = mobile[2:]

                d[cid] = {
                    "meter_no": meter_no,
                    "name": get_value("name", 2),
                    "address": get_value("address", 3),
                    "mobile_number": mobile,
                    "contractual_load": get_value("contractual_load", 5),
                    "class": get_value("class", 6),
                }

            utils.console_log(f"WORKER: Processed {len(d)} rows.")

            utils.update_meter_mapping(d)
            utils.console_log("WORKER: Database updated.")

            root.after(0, lambda: messagebox.showinfo("Success", f"Consumer data updated.\nRecords imported: {len(d)}"))
        except Exception as e:
            utils.console_log(f"!!! WORKER ERROR: {e}")
            root.after(0, lambda e=e: messagebox.showerror("Error", f"Failed to read Excel.\n{str(e)}\n\nEnsure 'openpyxl' is installed."))
        finally:
            utils.console_log("WORKER: Cleaning up UI.")
            root.after(0, reset_ui)

    def reset_ui():
        btn_update_consumer.config(state="normal", text="Update Consumer Data")
        update_meter_search_state()
        progress_bar.stop()
        progress_bar.pack_forget()

    root.after(100, open_file_picker)


def _normalize_fuzzy_text(value):
    text = str(value or "").upper()
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_mobile_10(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    return digits if len(digits) == 10 else ""


def _tokenize_for_lookup(text):
    tokens = [t for t in _normalize_fuzzy_text(text).split(" ") if len(t) >= 2]
    return tokens


def _fast_text_similarity(a, b):
    """Return text similarity percentage using RapidFuzz when available."""
    if not a or not b:
        return 0.0

    tokens_a = set(_tokenize_for_lookup(a))
    tokens_b = set(_tokenize_for_lookup(b))

    def token_sim(x, y):
        if rapidfuzz_fuzz is not None:
            return float(rapidfuzz_fuzz.ratio(x, y)) / 100.0
        return SequenceMatcher(None, x, y).ratio()

    # Coverage from input side: if most input tokens are found in DB text
    # (even with minor spelling differences), score should remain high.
    if tokens_a:
        covered = 0
        for ta in tokens_a:
            best = 0.0
            for tb in tokens_b:
                s = token_sim(ta, tb)
                if s > best:
                    best = s
            if best >= 0.80:
                covered += 1
        coverage_input = covered / len(tokens_a)
    else:
        coverage_input = 0.0

    if rapidfuzz_fuzz is not None:
        ratio_score = float(rapidfuzz_fuzz.ratio(a, b))
        sort_score = float(rapidfuzz_fuzz.token_sort_ratio(a, b))
        partial_score = float(rapidfuzz_fuzz.partial_ratio(a, b))
        set_score = float(rapidfuzz_fuzz.token_set_ratio(a, b))

        # Base score plus input-coverage boost for containment-style matches.
        blended = (
            (0.20 * ratio_score)
            + (0.20 * sort_score)
            + (0.25 * partial_score)
            + (0.35 * set_score)
        )
        adjusted = blended * (0.55 + (0.90 * coverage_input))
        return min(100.0, adjusted)
    return SequenceMatcher(None, a, b).ratio() * 100.0


def generate_fuzzy_lookup_template():
    try:
        save_path = filedialog.asksaveasfilename(
            title="Save Fuzzy Lookup Template",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="fuzzy_lookup_input_template.xlsx"
        )
        if not save_path:
            return

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "FuzzyLookupInput"
        sheet.append(["NAME", "C/O", "ADDRESS", "MOBILE NUMBER"])
        sheet.append(["", "", "", ""])
        sheet.freeze_panes = "A2"

        widths = [28, 28, 42, 18]
        for idx, width in enumerate(widths, start=1):
            col = openpyxl.utils.get_column_letter(idx)
            sheet.column_dimensions[col].width = width

        wb.save(save_path)
        messagebox.showinfo("Template Generated", f"Fuzzy lookup template created:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate fuzzy template.\n{e}")


def run_fuzzy_lookup_threaded(threshold=0.85, top_n=5):
    try:
        threshold = float(threshold)
    except Exception:
        threshold = 0.85
    threshold = max(0.0, min(1.0, threshold))

    try:
        top_n = int(top_n)
    except Exception:
        top_n = 20
    top_n = max(1, min(200, top_n))

    def open_input_file():
        input_path = filedialog.askopenfilename(
            title="Select Fuzzy Lookup Input File",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if not input_path:
            return

        input_dir = os.path.dirname(os.path.abspath(input_path))
        output_path = os.path.join(input_dir, "fuzzy_lookup_results.xlsx")
        if os.path.exists(output_path):
            ts = time.strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(input_dir, f"fuzzy_lookup_results_{ts}.xlsx")

        progress_bar.config(mode="determinate", maximum=100)
        progress_bar["value"] = 0
        progress_bar.pack(side=RIGHT, padx=10)
        status_label.config(text="Fuzzy Lookup: starting…")
        threading.Thread(target=worker, args=(input_path, output_path), daemon=True).start()

    def worker(input_path, output_path):
        start_time = time.time()
        _processed = [0]
        _total = [0]
        _last_ui = [0.0]

        def _update_progress():
            pct = (_processed[0] / _total[0] * 100.0) if _total[0] else 0.0
            elapsed_s = int(time.time() - start_time)
            progress_bar["value"] = pct
            status_label.config(
                text=f"Fuzzy Lookup: {pct:.0f}% ({_processed[0]}/{_total[0]}) | Elapsed: {elapsed_s}s"
            )

        try:
            db_profiles = utils.get_all_consumer_profiles()
            if not db_profiles:
                root.after(0, lambda: messagebox.showwarning("No Consumer Data", "No consumer data found. Please update consumer data first."))
                return

            prepped_db = []
            mobile_index = defaultdict(set)
            prefix_index = defaultdict(set)
            token_index = defaultdict(set)

            for idx, row in enumerate(db_profiles):
                db_name = str(row.get("name", ""))
                db_address = str(row.get("address", ""))
                db_combined_raw = f"{db_name} {db_address}".strip()
                combined_norm = _normalize_fuzzy_text(db_combined_raw)
                combined_tokens = set(_tokenize_for_lookup(combined_norm))
                mobile_norm = _normalize_mobile_10(row.get("mobile_number", ""))

                prepped_db.append({
                    "consumer_id": str(row.get("consumer_id", "")),
                    "name": db_name,
                    "address": db_address,
                    "mobile_number": str(row.get("mobile_number", "")),
                    "mobile_norm": mobile_norm,
                    "combined_norm": combined_norm,
                    "tokens": combined_tokens,
                })

                if mobile_norm:
                    mobile_index[mobile_norm].add(idx)

                for tok in combined_tokens:
                    if len(tok) >= 3:
                        prefix_index[tok[:3]].add(idx)
                    if len(tok) >= 4:
                        token_index[tok].add(idx)

            wb_in = openpyxl.load_workbook(input_path)
            sh_in = wb_in.active
            rows = list(sh_in.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Input Excel is empty.")

            header_map = {
                "name": "name",
                "co": "co",
                "c/o": "co",
                "careof": "co",
                "address": "address",
                "mobile": "mobile",
                "mobilenumber": "mobile",
                "mobile number": "mobile",
            }

            first_row = rows[0]
            idx_map = {}
            for i, v in enumerate(first_row):
                if v is None:
                    continue
                key = str(v).strip().lower().replace("_", " ")
                key = " ".join(key.split())
                compact = key.replace(" ", "")
                mapped = header_map.get(key) or header_map.get(compact)
                if mapped:
                    idx_map[mapped] = i

            has_headers = "name" in idx_map and "address" in idx_map
            data_rows = rows[1:] if has_headers else rows
            _total[0] = len(data_rows)

            def _get_input_value(row_vals, field, fallback_idx):
                idx = idx_map.get(field, fallback_idx)
                if idx is None or idx >= len(row_vals):
                    return ""
                val = row_vals[idx]
                return "" if val is None else str(val).strip()

            out_wb = openpyxl.Workbook()
            out_sh = out_wb.active
            out_sh.title = "FuzzyLookupResults"
            out_headers = [
                "Input Name",
                "Input C/O",
                "Input Address",
                "Input Mobile",
                "Matched Consumer ID",
                "Matched Name",
                "Matched Address",
                "Matched Mobile",
                "Combined Text Match %",
                "Mobile Exact Match %",
                "Final Score %",
                "Match Type",
                "Rank",
            ]
            out_sh.append(out_headers)

            for in_row in data_rows:
                _processed[0] += 1
                _now = time.time()
                if _now - _last_ui[0] >= 0.15:
                    _last_ui[0] = _now
                    root.after(0, _update_progress)

                if not in_row:
                    continue

                input_name = _get_input_value(in_row, "name", 0)
                input_co = _get_input_value(in_row, "co", 1)
                input_address = _get_input_value(in_row, "address", 2)
                input_mobile = _get_input_value(in_row, "mobile", 3)

                if not (input_name or input_co or input_address or input_mobile):
                    continue

                input_combined = _normalize_fuzzy_text(f"{input_name} {input_co} {input_address}")
                input_mobile_norm = _normalize_mobile_10(input_mobile)
                input_tokens = [t for t in _tokenize_for_lookup(input_combined) if len(t) >= 3]

                candidate_ids = set()

                # Exact mobile is the strongest and fastest blocker.
                if input_mobile_norm:
                    candidate_ids.update(mobile_index.get(input_mobile_norm, set()))

                # Prefix blocking from first few input tokens.
                for tok in input_tokens[:6]:
                    candidate_ids.update(prefix_index.get(tok[:3], set()))

                # Add candidates by longest tokens to improve precision.
                longest_tokens = sorted({t for t in input_tokens if len(t) >= 4}, key=len, reverse=True)[:4]
                for tok in longest_tokens:
                    candidate_ids.update(token_index.get(tok, set()))

                # Fallback: if blocking is too narrow, relax using only prefixes.
                if not candidate_ids and input_tokens:
                    for tok in input_tokens:
                        candidate_ids.update(prefix_index.get(tok[:3], set()))

                # Last-resort fallback keeps behavior robust for unusual/short inputs.
                if not candidate_ids:
                    candidate_ids = set(range(len(prepped_db)))

                candidates = []
                for cand_idx in candidate_ids:
                    db_row = prepped_db[cand_idx]
                    text_score = 0.0
                    if input_combined and db_row["combined_norm"]:
                        if input_tokens and db_row["tokens"]:
                            overlap = len(set(input_tokens).intersection(db_row["tokens"]))
                            if overlap == 0 and (not input_mobile_norm or db_row["mobile_norm"] != input_mobile_norm):
                                continue
                        text_score = _fast_text_similarity(input_combined, db_row["combined_norm"])

                    mobile_score = 100.0 if (input_mobile_norm and db_row["mobile_norm"] == input_mobile_norm) else 0.0
                    final_score = max(text_score, mobile_score)

                    if text_score >= (threshold * 100.0) or mobile_score == 100.0:
                        candidates.append({
                            "db": db_row,
                            "text_score": text_score,
                            "mobile_score": mobile_score,
                            "final_score": final_score,
                        })

                candidates.sort(key=lambda x: (x["final_score"], x["text_score"], x["mobile_score"]), reverse=True)
                candidates = candidates[:top_n]

                if not candidates:
                    out_sh.append([
                        input_name,
                        input_co,
                        input_address,
                        input_mobile,
                        "",
                        "",
                        "",
                        "",
                        "0.00",
                        "0.00",
                        "0.00",
                        "No Match",
                        "",
                    ])
                    continue

                rank = 1
                for c in candidates:
                    db_row = c["db"]
                    if c["mobile_score"] == 100.0 and c["text_score"] >= (threshold * 100.0):
                        match_type = "Both"
                    elif c["mobile_score"] == 100.0:
                        match_type = "Mobile Exact"
                    else:
                        match_type = "Fuzzy Text"

                    out_sh.append([
                        input_name,
                        input_co,
                        input_address,
                        input_mobile,
                        db_row["consumer_id"],
                        db_row["name"],
                        db_row["address"],
                        db_row["mobile_number"],
                        f"{c['text_score']:.2f}",
                        f"{c['mobile_score']:.2f}",
                        f"{c['final_score']:.2f}",
                        match_type,
                        rank,
                    ])
                    rank += 1

            widths = [24, 24, 36, 16, 18, 26, 36, 16, 20, 18, 14, 14, 10]
            for idx, width in enumerate(widths, start=1):
                col = openpyxl.utils.get_column_letter(idx)
                out_sh.column_dimensions[col].width = width
            out_sh.freeze_panes = "A2"

            out_wb.save(output_path)
            root.after(0, lambda: messagebox.showinfo("Fuzzy Lookup Complete", f"Results exported to:\n{output_path}"))
        except Exception as e:
            root.after(0, lambda e=e: messagebox.showerror("Fuzzy Lookup Error", str(e)))
        finally:
            root.after(0, finish_ui)

    def finish_ui():
        progress_bar.stop()
        progress_bar.pack_forget()
        progress_bar.config(mode="indeterminate")
        progress_bar["value"] = 0
        status_label.config(text="Ready")

    root.after(100, open_input_file)


def open_fuzzy_lookup_tool_dialog():
    dialog = Toplevel(root)
    dialog.title("Fuzzy Lookup Tool")
    dialog.geometry("500x260")
    dialog.transient(root)
    dialog.grab_set()

    frm = tb.Frame(dialog, padding=16)
    frm.pack(fill=BOTH, expand=True)

    tb.Label(frm, text="Fuzzy Lookup Tool", font=("Segoe UI", 11, "bold"), bootstyle="primary").pack(anchor=NW)
    tb.Label(
        frm,
        text=(
            "Input columns: NAME, C/O, ADDRESS, MOBILE NUMBER.\n"
            "Text matching uses combined NAME + C/O + ADDRESS against database NAME + ADDRESS.\n"
            "Mobile is matched exactly and included in scoring/output."
        ),
        justify=LEFT,
        wraplength=460,
        bootstyle="secondary"
    ).pack(anchor=NW, pady=(4, 10))

    opts = tb.Frame(frm)
    opts.pack(fill=X, pady=(0, 12))
    tb.Label(opts, text="Threshold (0 to 1):", font=("Segoe UI", 9)).pack(side=LEFT)
    threshold_var = tk.StringVar(value="0.85")
    tb.Entry(opts, textvariable=threshold_var, width=8, justify="center").pack(side=LEFT, padx=(6, 16))

    tb.Label(opts, text="Result Count:", font=("Segoe UI", 9)).pack(side=LEFT)
    topn_var = tk.StringVar(value="5")
    tb.Entry(opts, textvariable=topn_var, width=8, justify="center").pack(side=LEFT, padx=(6, 0))

    btn_row = tb.Frame(frm)
    btn_row.pack(fill=X)

    tb.Button(
        btn_row,
        text="Generate Fuzzy Template",
        width=24,
        bootstyle="primary",
        command=generate_fuzzy_lookup_template
    ).pack(side=LEFT, padx=(0, 10))

    def _run_and_close():
        dialog.destroy()
        run_fuzzy_lookup_threaded(threshold=threshold_var.get(), top_n=topn_var.get())

    tb.Button(
        btn_row,
        text="Run Fuzzy Lookup",
        width=20,
        bootstyle="primary",
        command=_run_and_close
    ).pack(side=LEFT)

    tb.Button(frm, text="Close", bootstyle="secondary", command=dialog.destroy).pack(side=BOTTOM, pady=(14, 0))


def update_meter_list_threaded():
    dialog = Toplevel(root)
    dialog.title("Update Consumer Data")
    dialog.transient(root)
    dialog.grab_set()

    frm = tb.Frame(dialog, padding=16)
    frm.pack(fill=BOTH, expand=True)

    tb.Label(
        frm,
        text="Choose an action for Consumer Data:",
        font=("Segoe UI", 10, "bold"),
        bootstyle="primary"
    ).pack(anchor=NW, pady=(0, 8))

    tb.Label(
        frm,
        text=(
            "1. Generate Template: creates an Excel template with the required columns.\n"
            "2. Import Consumer Data: import your filled Excel file."
        ),
        justify=LEFT,
        wraplength=390
    ).pack(anchor=NW, pady=(0, 12))

    tb.Label(
        frm,
        text=(
            "Required columns (any order):\n"
            "CONSUMER ID, METER NO, NAME, ADDRESS,\n"
            "MOBILE NUMBER, CONTRACTUAL LOAD, CLASS"
        ),
        justify=LEFT,
        wraplength=390,
        bootstyle="secondary"
    ).pack(anchor=NW, pady=(0, 12))

    btn_row = tb.Frame(frm)
    btn_row.pack(fill=X)

    def _generate_then_close():
        generate_consumer_data_template()

    def _import_then_close():
        dialog.destroy()
        import_consumer_data_threaded()

    tb.Button(
        btn_row,
        text="Generate Template",
        width=20,
        bootstyle="primary",
        command=_generate_then_close
    ).pack(side=LEFT, padx=(0, 10))

    tb.Button(
        btn_row,
        text="Import Consumer Data",
        width=20,
        bootstyle="primary",
        command=_import_then_close
    ).pack(side=LEFT)

    tb.Button(frm, text="Close", bootstyle="secondary", command=dialog.destroy).pack(side=BOTTOM, pady=(14, 0))

    # Auto-fit dialog to content and keep it within visible screen bounds.
    dialog.update_idletasks()
    required_width = dialog.winfo_reqwidth()
    required_height = dialog.winfo_reqheight()
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    screen_margin = 24

    width = min(required_width, screen_width - (screen_margin * 2))
    height = min(required_height, screen_height - (screen_margin * 2))
    x = max(screen_margin, (screen_width - width) // 2)
    y = max(screen_margin, (screen_height - height) // 2)

    dialog.geometry(f"{width}x{height}+{x}+{y}")

def show_about():
    about_win = Toplevel(root)
    about_win.title("About Spot Image Viewer")
    about_win.geometry("450x350")
    about_win.transient(root)
    about_win.grab_set()

    header_font = ("Segoe UI", 14, "bold")
    title_font = ("Segoe UI", 11, "bold")
    link_font = ("Segoe UI", 10, "underline")

    main_frame = tb.Frame(about_win, padding=20)
    main_frame.pack(fill=BOTH, expand=True)

    tb.Label(main_frame, text=f"Spot Image Viewer V{config.CURRENT_VERSION}", font=header_font, bootstyle="primary").pack(pady=(0, 15))
    
    tb.Label(main_frame, text="Developer:", font=title_font).pack(anchor=NW)
    tb.Label(main_frame, text="Pramod Verma", font=("Segoe UI", 10)).pack(anchor=NW, padx=(10, 0), pady=(0, 10))

    tb.Label(main_frame, text="Contact & Support:", font=title_font).pack(anchor=NW, pady=(10, 0))
    
    # Email
    email_label = tb.Label(main_frame, text="je.kushidaccc@gmail.com", font=link_font, cursor="hand2", bootstyle="info")
    email_label.pack(anchor=NW, padx=(10, 0))
    email_label.bind("<Button-1>", lambda e: webbrowser.open("mailto:je.kushidaccc@gmail.com"))

    # WhatsApp
    whatsapp_label = tb.Label(main_frame, text="WhatsApp Community", font=link_font, cursor="hand2", bootstyle="info")
    whatsapp_label.pack(anchor=NW, padx=(10, 0), pady=(5, 0))
    whatsapp_label.bind("<Button-1>", lambda e: webbrowser.open("https://chat.whatsapp.com/LZKLg40n8FxCLdnAIO9HGE"))
    
    tb.Button(main_frame, text="Close", command=about_win.destroy, bootstyle="secondary").pack(side=BOTTOM, pady=(20, 0))

def open_help():
    documentation.show_documentation(root)

def prompt_update(version, notes, link):
    title = f"New Version Available: v{version}"
    message = f"A new version of the application is available.\n\nRelease Notes:\n{notes}\n\nDo you want to download it now?"
    if messagebox.askyesno(title, message):
        webbrowser.open(link)

def on_update_found(version, notes, link):
    root.after(0, lambda: prompt_update(version, notes, link))

def on_update_check_finished_auto(status, data):
    if status == "update_found":
        on_update_found(
            data.get("version"),
            data.get("release_notes"),
            data.get("download_url")
        )

def on_update_check_finished_manual(status, data):
    if status == "update_found":
        version = data.get("version")
        notes = data.get("release_notes")
        link = data.get("download_url")
        root.after(0, lambda: prompt_update(version, notes, link))
    elif status == "no_update":
        latest_version = data.get("version") if data else config.CURRENT_VERSION
        release_notes = data.get("release_notes", "No feature notes available.") if data else "No feature notes available."
        message = (
            "You are already using the latest version.\n\n"
            f"Current Version: v{config.CURRENT_VERSION}\n"
            f"Latest Version: v{latest_version}\n\n"
            "Latest Version Features:\n"
            f"{release_notes}"
        )
        root.after(0, lambda: messagebox.showinfo("No Updates", message))
    elif status == "error":
        error_msg = data.get('error', 'An unknown error occurred.')
        root.after(0, lambda: messagebox.showerror("Update Error", f"Failed to check for updates:\n{error_msg}"))

def manual_update_check():
    utils.check_for_updates_background(
        config.CURRENT_VERSION, 
        config.UPDATE_URL, 
        on_update_check_finished_manual
    )

def on_startup_check():
    try:
        utils.check_for_updates_background(config.CURRENT_VERSION, config.UPDATE_URL, on_update_check_finished_auto)
        update_folder_list_ui()
        update_meter_search_state()
        cnt = database.get_total_image_count()
        status_label.config(text=f"Total Indexed Images: {cnt}")
        
        if cnt == 0:
            if messagebox.askyesno("Startup", "Database is empty. Index images now?"):
                start_indexing_process()
    except Exception as e:
        messagebox.showerror("Startup Error", str(e))

# ==============================================================================
# GUI SETUP 
# ==============================================================================
set_windows_app_id()

root = tb.Window(themename="cosmo") 
root.title(f"Spot Image Viewer V{config.CURRENT_VERSION}")
root.geometry("1300x850")
root.state("zoomed")

try:
    icon_ico, icon_png = ensure_app_icons()
    if os.name == "nt":
        root.iconbitmap(icon_ico)
        root.wm_iconbitmap(icon_ico)

    # Keep a persistent reference so Tk does not garbage-collect the icon image.
    app_icon = ImageTk.PhotoImage(Image.open(icon_png))
    root.iconphoto(True, app_icon)

    if os.name == "nt":
        root.after(100, lambda: root.iconbitmap(icon_ico))
except Exception as e:
    print(f"Icon setup warning: {e}")
    app_icon = None

def on_close():
    root.destroy()
    os._exit(0) 

root.protocol("WM_DELETE_WINDOW", on_close)

mb = Menu(root)
root.config(menu=mb)

fm = Menu(mb, tearoff=0)
mb.add_cascade(label="File", menu=fm)
fm.add_command(label="Reload Images", command=start_indexing_process)
fm.add_command(label="Update Consumer Data", command=update_meter_list_threaded)
fm.add_separator()
fm.add_command(label="Exit", command=root.quit)

bm = Menu(mb, tearoff=0)
mb.add_cascade(label="Backup", menu=bm)
bm.add_command(label="Backup Images", command=perform_backup)

nm = Menu(mb, tearoff=0)
mb.add_cascade(label="Notes", menu=nm)
nm.add_command(label="Export Notes", command=export_notes_csv)

vm = Menu(mb, tearoff=0)
mb.add_cascade(label="Verification", menu=vm)
vm.add_command(label="Low Consumption Check", command=lambda: LowConsumptionVerifier(root))

# Remove the old launch_tool references and use the classes instead
tm = Menu(mb, tearoff=0)
mb.add_cascade(label="Tools", menu=tm)
tm.add_command(label="Bill Calculator", command=lambda: BillCalculatorApp(root))
tm.add_command(label="Theft Bill Calculator", command=lambda: TheftCalculatorApp(root))
tm.add_command(label="Tariff Editor", command=lambda: TariffEditor(root))
tm.add_command(label="Fuzzy Lookup", command=open_fuzzy_lookup_tool_dialog)
# Disabled by default until consumer data confirms presence (update_meter_search_state re-enables it)
tm.entryconfig("Fuzzy Lookup", state="disabled")

hm = Menu(mb, tearoff=0)
mb.add_cascade(label="Help", menu=hm)
hm.add_command(label="Documentation", command=open_help)
hm.add_command(label="Check for Updates", command=manual_update_check)
hm.add_command(label="About", command=show_about)

top_f = tb.Frame(root, padding=8, bootstyle="light") 
top_f.pack(fill=X)

search_card = tb.Labelframe(top_f, text="Search Controls", padding=6, bootstyle="default")
search_card.pack(fill=X, expand=True)

# Col 6 is a spacer so Update Consumer Data and utility buttons stay right-aligned
search_card.columnconfigure(6, weight=1)

# ── Row 0: Consumer ID (always enabled) + Meter No (enabled only with data) ──
tb.Label(search_card, text="Consumer ID", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=(4, 4), pady=2, sticky="e")
entry_consumer_id = tb.Entry(search_card, width=22, font=("Segoe UI", 10))
entry_consumer_id.grid(row=0, column=1, padx=(0, 6), pady=2, sticky="w")
entry_consumer_id.bind("<Return>", lambda e: search_consumer())
entry_consumer_id.bind("<space>", lambda e: show_history(e, "consumer_ids", entry_consumer_id))

tb.Button(search_card, text="Search", width=12, bootstyle="primary", command=search_consumer).grid(row=0, column=2, padx=(0, 10), pady=2)

tb.Label(search_card, text="Meter No", font=("Segoe UI", 9, "bold")).grid(row=0, column=3, padx=(0, 4), pady=2, sticky="e")
entry_meter_number = tb.Entry(search_card, width=22, font=("Segoe UI", 10))
entry_meter_number.grid(row=0, column=4, padx=(0, 6), pady=2, sticky="w")
entry_meter_number.bind("<Return>", lambda e: search_meter())
entry_meter_number.bind("<space>", lambda e: show_history(e, "meter_numbers", entry_meter_number))

meter_button = tb.Button(search_card, text="Search Meter", width=14, bootstyle="primary", command=search_meter)
meter_button.grid(row=0, column=5, padx=(0, 4), pady=2, sticky="w")

btn_update_consumer = tb.Button(search_card, text="Update Consumer Data", width=23, bootstyle="danger", command=update_meter_list_threaded)
btn_update_consumer.grid(row=0, column=6, padx=(8, 8), pady=2)

# ── Row 1: Name + Mobile (greyed when no consumer data loaded) ──
tb.Label(search_card, text="Name", font=("Segoe UI", 9, "bold")).grid(row=1, column=0, padx=(4, 4), pady=2, sticky="e")
entry_name = tb.Entry(search_card, width=22, font=("Segoe UI", 10))
entry_name.grid(row=1, column=1, padx=(0, 6), pady=2, sticky="w")
entry_name.bind("<Return>", lambda e: search_name())
entry_name.bind("<space>", lambda e: show_history(e, "consumer_names", entry_name))
btn_search_name = tb.Button(search_card, text="Search Name", width=12, bootstyle="primary", command=search_name)
btn_search_name.grid(row=1, column=2, padx=(0, 10), pady=2)

tb.Label(search_card, text="Mobile", font=("Segoe UI", 9, "bold")).grid(row=1, column=3, padx=(0, 4), pady=2, sticky="e")
entry_mobile = tb.Entry(search_card, width=22, font=("Segoe UI", 10))
entry_mobile.grid(row=1, column=4, padx=(0, 6), pady=2, sticky="w")
entry_mobile.bind("<Return>", lambda e: search_mobile())
entry_mobile.bind("<space>", lambda e: show_history(e, "mobile_numbers", entry_mobile))
btn_search_mobile = tb.Button(search_card, text="Search Mobile", width=14, bootstyle="primary", command=search_mobile)
btn_search_mobile.grid(row=1, column=5, padx=(0, 4), pady=2, sticky="w")

# ── Row 1 col 6: spacer-aligned with Update Consumer Data above ──
# (empty cell keeps columns aligned)

# ── Row 2: hint label (shown only when consumer data is absent) ──
lbl_consumer_hint = tb.Label(
    search_card,
    text="\u26a0  Update Consumer Data to enable Meter No, Name & Mobile search",
    font=("Segoe UI", 8),
    bootstyle="warning"
)
lbl_consumer_hint.grid(row=2, column=0, columnspan=7, padx=(4, 4), pady=(0, 2), sticky="w")

# ── Utility buttons — far right (spacer col 6 keeps visual separation) ──
btn_theme_toggle = tb.Button(search_card, text="Dark Mode", width=14, command=toggle_theme, bootstyle="primary")
btn_theme_toggle.grid(row=0, column=8, padx=(4, 6), pady=2, sticky="e")

btn_reload = tb.Button(search_card, text="Reload Images", width=14, bootstyle="primary", command=start_indexing_process)
btn_reload.grid(row=1, column=8, padx=(4, 6), pady=2, sticky="e")

stat_f = tb.Frame(root, padding=5, bootstyle="secondary")
stat_f.pack(side=BOTTOM, fill=X)
status_label = tb.Label(stat_f, text="Ready", font=("Segoe UI", 10), bootstyle="inverse-secondary")
status_label.pack(side=LEFT, padx=5)
progress_bar = ttk.Progressbar(stat_f, mode="indeterminate", length=200)

toggle_f = tb.Frame(stat_f, bootstyle="secondary")
toggle_f.pack(side=RIGHT)
btn_folders = tb.Button(toggle_f, text="Networks", command=toggle_folders, bootstyle="dark-outline")
btn_folders.pack(side=RIGHT, padx=5)
btn_notes = tb.Button(toggle_f, text="Notes", command=toggle_notes, bootstyle="dark-outline")
btn_notes.pack(side=RIGHT, padx=5)

container = tb.Frame(root, padding=10)
container.pack(fill=BOTH, expand=True)

left_p = tb.Labelframe(container, text="Details", width=250, padding=10, bootstyle="default")
left_p.pack(side=LEFT, fill=Y, padx=5)

label_consumer_details = tb.Label(left_p, text="Welcome", font=("Segoe UI", 9), wraplength=220, justify=LEFT, bootstyle="primary")
label_consumer_details.pack(pady=10)
label_latest_date = tb.Label(left_p, text="", font=("Segoe UI", 10, "bold"), bootstyle="success")
label_latest_date.pack(pady=5)
label_total_images = tb.Label(left_p, text="", font=("Segoe UI", 10))
label_total_images.pack(pady=5)

tb.Label(left_p, text="Available Dates:", font=("Segoe UI", 10, "bold")).pack(fill=X, pady=(15,0))
listbox_dates = Listbox(left_p, font=("Segoe UI", 11), relief="flat")
listbox_dates.pack(fill=BOTH, expand=True, pady=5)
listbox_dates.bind("<<ListboxSelect>>", on_date_select)

right_outer_frame = tb.Frame(container)
right_outer_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=5)

right_inner_frame = tb.Frame(right_outer_frame)
right_inner_frame.pack(fill=BOTH, expand=True)

right_inner_frame.grid_rowconfigure(0, weight=1)
right_inner_frame.grid_columnconfigure(0, weight=1)

welcome_frame = tb.Frame(right_inner_frame, bootstyle="light")
welcome_frame.grid(row=0, column=0, sticky="nsew")
welcome_msg = tb.Label(welcome_frame, text="Welcome to Spot Image Viewer", font=("Segoe UI", 24, "bold"), bootstyle="secondary")
welcome_msg.place(relx=0.5, rely=0.4, anchor=CENTER)
welcome_sub = tb.Label(welcome_frame, text="Enter a Consumer ID or Meter Number to begin search.", font=("Segoe UI", 14), bootstyle="secondary")
welcome_sub.place(relx=0.5, rely=0.5, anchor=CENTER)

preview_container = tb.Frame(right_inner_frame)
preview_container.grid(row=0, column=0, sticky="nsew")

preview_canvas_scroller = tk.Canvas(preview_container, highlightthickness=0)
preview_scrollbar = tb.Scrollbar(preview_container, orient="vertical", command=preview_canvas_scroller.yview)
scrollable_frame = tb.Frame(preview_canvas_scroller)

scrollable_frame.bind("<Configure>", lambda e: preview_canvas_scroller.configure(scrollregion=preview_canvas_scroller.bbox("all")))
preview_canvas_scroller.create_window((0, 0), window=scrollable_frame, anchor="nw")
preview_canvas_scroller.configure(yscrollcommand=preview_scrollbar.set)

preview_canvas_scroller.pack(side="left", fill="both", expand=True)
preview_scrollbar.pack(side="right", fill="y")

def _on_preview_mousewheel(event):
    preview_canvas_scroller.yview_scroll(int(-1 * (event.delta / 120)), "units")

def _bind_preview_mousewheel(event):
    preview_canvas_scroller.bind_all("<MouseWheel>", _on_preview_mousewheel)

def _unbind_preview_mousewheel(event):
    preview_canvas_scroller.unbind_all("<MouseWheel>")

preview_canvas_scroller.bind('<Enter>', _bind_preview_mousewheel)
preview_canvas_scroller.bind('<Leave>', _unbind_preview_mousewheel)

canvas_container = tb.Frame(right_inner_frame)
canvas_container.grid(row=0, column=0, sticky="nsew")

btn_show_previews = tb.Button(right_outer_frame, text="Show Previews", command=raise_preview_pane, bootstyle="info-outline")

canvas = tk.Canvas(canvas_container, highlightthickness=0)
canvas.pack(fill=BOTH, expand=True)

canvas.bind("<ButtonPress-1>", start_pan)
canvas.bind("<B1-Motion>", do_pan)
canvas.bind("<ButtonRelease-1>", end_pan)

canvas.bind('<Enter>', _bind_image_mousewheel)
canvas.bind('<Leave>', _unbind_image_mousewheel)

btns_f = tb.Frame(canvas_container)
btns_f.pack(pady=5)
btn_zoomin = tb.Button(btns_f, text="+", width=4, command=lambda: zoom(1.2), bootstyle="secondary")
btn_zoomout = tb.Button(btns_f, text="-", width=4, command=lambda: zoom(0.8), bootstyle="secondary")
btn_prnt = tb.Button(btns_f, text="Print", command=print_image, bootstyle="info")
btn_sv = tb.Button(btns_f, text="Save", command=save_image, bootstyle="success")
btn_svall = tb.Button(btns_f, text="Save All", command=save_all_images, bootstyle="success")

def show_buttons():
    btn_zoomout.pack(side=LEFT, padx=3)
    btn_zoomin.pack(side=LEFT, padx=3)
    btn_prnt.pack(side=LEFT, padx=3)
    btn_sv.pack(side=LEFT, padx=3)
    btn_svall.pack(side=LEFT, padx=3)

def hide_buttons():
    for w in btns_f.winfo_children(): w.pack_forget()

welcome_frame.tkraise()
hide_buttons()

notes_pane = tb.Labelframe(container, text="Notes", width=280, padding=10, bootstyle="warning")
label_prev_note = tb.Label(notes_pane, text="No notes", wraplength=250, justify=LEFT, bootstyle="secondary")
label_prev_note.pack(fill=X, pady=5)

note_opts_f = tb.Frame(notes_pane)
note_opts_f.pack(fill=X, pady=5)
note_options = utils.load_note_options()
note_var = tk.StringVar(value=note_options[0] if note_options else "")
combo_notes = ttk.Combobox(note_opts_f, textvariable=note_var, values=note_options)
combo_notes.pack(side=LEFT, fill=X, expand=True)
tb.Button(note_opts_f, text="+", width=3, command=add_new_note_option, bootstyle="success-outline").pack(side=LEFT, padx=(5,0))

txt_remarks = tk.Text(notes_pane, height=6, width=25, relief="flat")
txt_remarks.pack(fill=X, pady=5)
tb.Button(notes_pane, text="Save Note", command=save_current_note, bootstyle="success").pack(fill=X, pady=2)
tb.Button(notes_pane, text="Delete Note", command=delete_current_note, bootstyle="danger-outline").pack(fill=X, pady=2)

folder_pane = tb.Labelframe(container, text="Networks", width=280, padding=10, bootstyle="info")
folder_listbox = Listbox(folder_pane, height=15, relief="flat")
folder_listbox.pack(fill=BOTH, expand=True, pady=5)
tb.Button(folder_pane, text="Add Folder", command=add_network_folder, bootstyle="success-outline").pack(fill=X, pady=2)
tb.Button(folder_pane, text="Remove Folder", command=remove_network_folder, bootstyle="danger-outline").pack(fill=X, pady=2)

def run_app():
    success, msg = database.init_db()
    if not success:
        print(f"Database init failed: {msg}")

    refresh_non_ttk_widget_colors()
    update_theme_toggle_button_text()
    
    root.after(500, on_startup_check)
    root.mainloop()
