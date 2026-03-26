"""
This module contains all QDialog-based windows for the application,
such as the settings, search, database manager, and rule manager dialogs.
"""

import sqlite3
import json
import re
import openpyxl

from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLineEdit, QListWidget, QPushButton,
                             QCheckBox, QTabWidget, QTableWidget, QTableWidgetItem, QHBoxLayout, 
                             QFileDialog, QMessageBox, QGroupBox, QFormLayout, QComboBox, 
                             QSpinBox, QHeaderView, QInputDialog, QWidget, QSplitter, 
                             QTreeWidget, QTreeWidgetItem, QLabel)
from PyQt6.QtCore import Qt

# Import shared constants
from constants import PROPERTY_DATA, FORMULA_VARS

class SearchDialog(QDialog):
    """
    A dialog for searching the materials or labor database and adding
    an item to the estimate.
    """
    def __init__(self, db_type, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Search & Add {db_type}")
        self.setFixedSize(600, 400)
        self.layout = QVBoxLayout(self)
        
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Type to search official items...")
        self.layout.addWidget(self.search_box)
        
        self.list_widget = QListWidget()
        self.layout.addWidget(self.list_widget)
        self.search_box.textChanged.connect(self.filter_list)
        
        self.add_btn = QPushButton("Add Selected to Estimate")
        self.add_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        self.add_btn.clicked.connect(self.accept)
        self.layout.addWidget(self.add_btn)

        self.items_data = {}
        self.load_data(db_type)

    def load_data(self, db_type):
        """Loads data from the specified database table."""
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        if db_type == "Material":
            cursor.execute("SELECT item_code, item_name, unit, rate FROM materials")
        else:
            cursor.execute("SELECT labor_code, task_name, unit, rate FROM labor")
        
        for row in cursor.fetchall():
            display_text = f"{row[1]} ({row[2]}) - Rs. {row[3]}"
            self.items_data[display_text] = {"code": row[0], "name": row[1], "unit": row[2], "rate": row[3], "type": db_type}
            self.list_widget.addItem(display_text)
        conn.close()

    def filter_list(self, text):
        """Filters the list widget based on the search box text."""
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def get_selected(self):
        """Returns the data for the currently selected item."""
        selected = self.list_widget.currentItem()
        if selected:
            return self.items_data[selected.text()]
        return None


class SettingsDialog(QDialog):
    """
    A dialog for accessing advanced application settings like the rule
    engine toggle and the database/ruleset managers.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("Advanced Settings")
        self.setFixedSize(300, 150)
        
        layout = QVBoxLayout(self)
        
        db_sync_btn = QPushButton("🗃️ Master DB (Excel Sync)")
        db_sync_btn.clicked.connect(self.parent_app.open_db_manager)
        layout.addWidget(db_sync_btn)

        rules_btn = QPushButton("🧠 Ruleset Manager")
        rules_btn.clicked.connect(self.parent_app.open_rule_manager)
        layout.addWidget(rules_btn)

        layout.addStretch()


class DatabaseManagerDialog(QDialog):
    """
    A dialog for managing the master database, allowing import from and
    export to Excel files.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Master DB Sync (Excel)")
        self.setGeometry(100, 100, 800, 600)
        
        layout = QVBoxLayout(self)
        
        button_layout = QHBoxLayout()
        import_btn = QPushButton("📥 Import from Excel")
        import_btn.clicked.connect(self.import_from_excel)
        export_btn = QPushButton("📤 Export to Excel")
        export_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(import_btn)
        button_layout.addWidget(export_btn)
        layout.addLayout(button_layout)
        
        self.tabs = QTabWidget()
        self.materials_table = QTableWidget()
        self.labor_table = QTableWidget()
        self.tabs.addTab(self.materials_table, "Materials")
        self.tabs.addTab(self.labor_table, "Labor")
        layout.addWidget(self.tabs)
        
        self.load_table_data()

    def populate_table(self, table_widget, data, headers):
        """Fills a table widget with data."""
        table_widget.setRowCount(0)
        table_widget.setColumnCount(len(headers))
        table_widget.setHorizontalHeaderLabels(headers)
        for row_num, row_data in enumerate(data):
            table_widget.insertRow(row_num)
            for col_num, col_data in enumerate(row_data):
                table_widget.setItem(row_num, col_num, QTableWidgetItem(str(col_data)))
        table_widget.resizeColumnsToContents()

    def load_table_data(self):
        """Loads data from the database and populates the tables."""
        self.materials_table.clear()
        self.labor_table.clear()
        
        conn = sqlite3.connect('erp_master.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM materials")
        self.populate_table(self.materials_table, cursor.fetchall(), ["Item Code", "Item Name", "Rate", "Unit"])
        
        cursor.execute("SELECT * FROM labor")
        self.populate_table(self.labor_table, cursor.fetchall(), ["Labor Code", "Task Name", "Rate", "Unit"])
        
        conn.close()

    def export_to_excel(self):
        """Exports the database tables to an Excel file."""
        filename, _ = QFileDialog.getSaveFileName(self, "Export DB to Excel", "master_database.xlsx", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                ws = wb.create_sheet(title=table_name)
                cursor.execute(f"PRAGMA table_info({table_name})")
                headers = [info[1] for info in cursor.fetchall()]
                ws.append(headers)

                cursor.execute(f"SELECT * FROM {table_name}")
                for row in cursor.fetchall():
                    ws.append(row)
            
            conn.close()
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Database exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export database: {e}")

    def import_from_excel(self):
        """Imports data from an Excel file into the database."""
        filename, _ = QFileDialog.getOpenFileName(self, "Import DB from Excel", "", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = openpyxl.load_workbook(filename)
            conn = sqlite3.connect('erp_master.db')
            cursor = conn.cursor()

            for table_name in ["materials", "labor"]:
                if table_name in wb.sheetnames:
                    ws = wb[table_name]
                    cursor.execute(f"DELETE FROM {table_name}")
                    headers = [cell.value for cell in ws[1]]
                    placeholders = ', '.join(['?'] * len(headers))
                    query = f"INSERT OR REPLACE INTO {table_name} ({', '.join(headers)}) VALUES ({placeholders})"
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            cursor.execute(query, row)
            conn.commit()
            conn.close()
            self.load_table_data()
            QMessageBox.information(self, "Success", "Database imported successfully. Please restart or refresh to see updates in 'Add Custom' lists.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import database: {e}")


class RulesetManagerDialog(QDialog):
    """
    The main UI for creating, viewing, and deleting rules for the
    DynamicRuleEngine.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Advanced Ruleset Manager")
        self.setGeometry(150, 150, 1200, 700)
        
        # Main layout
        self.main_layout = QHBoxLayout(self)
        
        # Data
        self.rules = []
        self.selected_rule_index = -1
        self.selected_result_item = None # To hold item from DB search
        self.condition_widgets = []


        # Use the imported constant data
        self.property_data = PROPERTY_DATA
        self.formula_vars = FORMULA_VARS

        # --- UI Components ---
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left Panel: Rule Explorer
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        
        explorer_tools = QHBoxLayout()
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("🔍 Search rules...")
        self.search_bar.textChanged.connect(self.filter_tree)
        add_rule_btn = QPushButton("➕ New")
        add_rule_btn.clicked.connect(self.create_new_rule)
        del_rule_btn = QPushButton("🗑️ Delete")
        del_rule_btn.clicked.connect(self.delete_selected_rule)
        
        explorer_tools.addWidget(self.search_bar)
        explorer_tools.addWidget(add_rule_btn)
        explorer_tools.addWidget(del_rule_btn)
        
        self.rule_tree = QTreeWidget()
        self.rule_tree.setHeaderLabels(["Rules"])
        self.rule_tree.itemClicked.connect(self.on_rule_selected)

        left_layout.addLayout(explorer_tools)
        left_layout.addWidget(self.rule_tree)
        
        # Right Panel: Rule Editor
        self.editor_panel = QGroupBox("Select a rule to edit")
        self.editor_panel.setDisabled(True) # Disabled until a rule is selected
        
        # Add panels to splitter
        self.splitter.addWidget(left_panel)
        self.splitter.addWidget(self.editor_panel)
        self.splitter.setSizes([350, 850])

        self.main_layout.addWidget(self.splitter)
        
        self.load_rules()

    def on_rule_selected(self, item, column):
        """Handles the selection of a rule in the tree."""
        rule_index = item.data(0, Qt.ItemDataRole.UserRole)
        
        if rule_index is not None:
            self.selected_rule_index = rule_index
            rule = self.rules[self.selected_rule_index]
            self.selected_result_item = rule 
            
            self.editor_panel.setDisabled(False)
            self.editor_panel.setTitle(f"Editing Rule: {rule.get('item_name', 'Unnamed Rule')}")
            
            self.build_editor_ui(rule)
        else:
            self.selected_rule_index = -1
            self.editor_panel.setTitle("Select a rule to edit")
            self.editor_panel.setDisabled(True)
            if self.editor_panel.layout() is not None:
                while self.editor_panel.layout().count():
                    child = self.editor_panel.layout().takeAt(0)
                    if child.widget():
                        child.widget().deleteLater()
                self.editor_panel.layout().deleteLater()
                
    def build_editor_ui(self, rule):
        """Dynamically builds the UI for editing the given rule."""
        if self.editor_panel.layout() is not None:
            while self.editor_panel.layout().count():
                child = self.editor_panel.layout().takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
            self.editor_panel.layout().deleteLater()

        editor_layout = QVBoxLayout(self.editor_panel)
        
        details_group = QGroupBox("Rule Details")
        details_layout = QFormLayout(details_group)
        
        self.rule_name_input = QLineEdit(rule.get("item_name", ""))
        obj_type_label = QLabel(f"<b>{rule.get('object', 'N/A')}</b>")
        
        details_layout.addRow("Rule Display Name:", self.rule_name_input)
        details_layout.addRow("Applies To:", obj_type_label)
        editor_layout.addWidget(details_group)

        cond_group = QGroupBox("Conditions (IF all are true)")
        self.cond_layout = QVBoxLayout(cond_group)
        
        self.conditions_container = QWidget()
        self.cond_rows_layout = QVBoxLayout(self.conditions_container)
        self.cond_rows_layout.setContentsMargins(0, 0, 0, 0)
        self.cond_layout.addWidget(self.conditions_container)
        
        add_cond_btn = QPushButton("➕ Add Condition")
        add_cond_btn.clicked.connect(self.add_condition_row)
        self.cond_layout.addWidget(add_cond_btn)
        self.cond_layout.addStretch()
        editor_layout.addWidget(cond_group)
        
        actions_group = QGroupBox("Action (THEN)")
        actions_layout = QFormLayout(actions_group)
        
        self.item_label = QLineEdit(f"({rule.get('type')}) {rule.get('item_name')}")
        self.item_label.setReadOnly(True)
        search_btn = QPushButton("🔍 Search DB")
        search_btn.clicked.connect(self.search_database)
        
        item_select_layout = QHBoxLayout()
        item_select_layout.addWidget(self.item_label)
        item_select_layout.addWidget(search_btn)
        
        self.formula_input = QLineEdit(rule.get("formula", "1"))

        actions_layout.addRow("Add Item:", item_select_layout)
        actions_layout.addRow("Quantity Formula:", self.formula_input)
        editor_layout.addWidget(actions_group)

        save_btn = QPushButton("💾 Save Changes")
        save_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 8px;")
        save_btn.clicked.connect(self.save_current_rule)
        editor_layout.addWidget(save_btn, 0, Qt.AlignmentFlag.AlignRight)
        
        editor_layout.addStretch()

        self.condition_widgets = []
        self.parse_and_display_conditions(rule)
        
    def parse_and_display_conditions(self, rule):
        """Parses a rule's condition string and populates the UI."""
        for widget_map in self.condition_widgets:
            widget_map['widget'].deleteLater()
        self.condition_widgets.clear()
        
        condition_str = rule.get('condition', 'True')
        if not condition_str or condition_str == 'True':
            self.add_condition_row()
            return

        tokens = re.split(r'\s+(and|or)\s+', condition_str, flags=re.IGNORECASE)
        
        self.add_condition_row(expression=tokens[0])
        
        if len(tokens) > 1:
            for i in range(1, len(tokens), 2):
                self.add_condition_row(logical_op=tokens[i].upper(), expression=tokens[i+1])

    def add_condition_row(self, logical_op=None, expression=None):
        """Adds a new row of widgets for building a rule condition."""
        cond_row_widget = QWidget()
        row_layout = QHBoxLayout(cond_row_widget)
        
        logical_op_combo = QComboBox()
        logical_op_combo.addItems(["AND", "OR"])
        logical_op_combo.setVisible(len(self.condition_widgets) > 0)
        
        rem_button = QPushButton("➖")
        rem_button.setFixedWidth(30)
        
        rule = self.rules[self.selected_rule_index]
        obj_name = rule.get('object')
        
        prop_combo = QComboBox()
        if obj_name and obj_name in self.property_data:
            prop_combo.addItems(list(self.property_data[obj_name].keys()))
        
        op_combo = QComboBox()
        op_combo.addItems(['==', '!=', '>', '<', '>=', '<=', 'in', 'not in'])

        value_widget = QLineEdit()
        
        row_layout.addWidget(logical_op_combo, 1)
        row_layout.addWidget(prop_combo, 3)
        row_layout.addWidget(op_combo, 2)
        row_layout.addWidget(value_widget, 4)
        row_layout.addWidget(rem_button, 0)

        widget_map = {
            'widget': cond_row_widget, 'logical_op': logical_op_combo,
            'prop': prop_combo, 'op': op_combo, 'value': value_widget,
            'rem_btn': rem_button
        }
        self.condition_widgets.append(widget_map)
        self.cond_rows_layout.addWidget(cond_row_widget)
        
        prop_combo.currentTextChanged.connect(lambda text, w=widget_map: self.on_property_change(text, w))
        rem_button.clicked.connect(lambda ch, w=cond_row_widget: self.remove_condition_row(w))
        
        if logical_op:
            logical_op_combo.setCurrentText(logical_op)
            
        if expression:
            match = re.match(r"(\S+)\s*([<>=!in\s]+)\s*(.*)", expression.strip())
            if match:
                prop, op, val = match.groups()
                prop = prop.strip(); op = op.strip(); val = val.strip().strip("'\"")
                
                prop_combo.setCurrentText(prop)
                self.on_property_change(prop, widget_map)
                
                value_widget_after = widget_map['value']
                op_combo.setCurrentText(op)

                if isinstance(value_widget_after, QComboBox): value_widget_after.setCurrentText(val)
                elif isinstance(value_widget_after, QSpinBox): value_widget_after.setValue(int(float(val)))
                else: value_widget_after.setText(val)
        else:
            self.on_property_change(prop_combo.currentText(), widget_map)

    def remove_condition_row(self, widget_to_remove):
        """Removes a condition row from the UI."""
        if len(self.condition_widgets) <= 1: 
            QMessageBox.warning(self, "Cannot Remove", "A rule must have at least one condition.")
            return

        for i, w_map in enumerate(self.condition_widgets):
            if w_map['widget'] == widget_to_remove:
                self.condition_widgets.pop(i)
                break
        widget_to_remove.deleteLater()

        if len(self.condition_widgets) > 0:
            self.condition_widgets[0]['logical_op'].setVisible(False)
            
    def on_property_change(self, prop_name, widget_map):
        """Updates the value input widget based on the selected property."""
        if self.selected_rule_index == -1: return
        rule = self.rules[self.selected_rule_index]
        obj_name = rule.get('object')
        
        if not obj_name or not prop_name or obj_name not in self.property_data or prop_name not in self.property_data[obj_name]:
            return

        prop_info = self.property_data[obj_name].get(prop_name)
        current_widget = widget_map['value']
        
        target_class = QLineEdit
        if isinstance(prop_info, list): target_class = QComboBox
        elif prop_info == 'int': target_class = QSpinBox

        if not isinstance(current_widget, target_class):
            new_widget = target_class()
            if isinstance(new_widget, QSpinBox): new_widget.setRange(-10000, 10000)
            
            layout = widget_map['widget'].layout()
            layout.replaceWidget(current_widget, new_widget)
            current_widget.deleteLater()
            widget_map['value'] = new_widget
            current_widget = new_widget

        if isinstance(current_widget, QComboBox):
            current_widget.clear()
            if isinstance(prop_info, list): current_widget.addItems([str(p) for p in prop_info])
    
    def search_database(self):
        """Opens the search dialog to select a result item for the rule."""
        db_type, ok = QInputDialog.getItem(self, "Select Database", "Source:", ["Material", "Labor"], 0, False)
        if ok and db_type:
            dialog = SearchDialog(db_type, self)
            if dialog.exec():
                self.selected_result_item = dialog.get_selected()
                if self.selected_result_item:
                    self.item_label.setText(f"({self.selected_result_item['type']}) {self.selected_result_item['name']}")

    def save_current_rule(self):
        """Compiles the UI fields into a rule dictionary and saves it."""
        if self.selected_rule_index == -1: return
        
        if not self.selected_result_item:
            QMessageBox.warning(self, "Incomplete Rule", "Please select a result item from the database first.")
            return

        condition_parts = []
        for i, w_map in enumerate(self.condition_widgets):
            prop = w_map['prop'].currentText()
            if not prop: continue

            if i > 0:
                condition_parts.append(w_map['logical_op'].currentText().lower())
            
            op = w_map['op'].currentText()
            
            value_widget = w_map['value']
            if isinstance(value_widget, QComboBox): val = value_widget.currentText()
            elif isinstance(value_widget, QSpinBox): val = value_widget.value()
            else: val = value_widget.text()
            
            val_str = str(val)
            if val_str.lower() == 'true': val = True
            elif val_str.lower() == 'false': val = False
            
            if isinstance(val, str):
                try:
                    float(val) # Check if it can be a number
                except ValueError:
                    val = f"'{val}'" # Add quotes if it's a non-numeric string

            condition_parts.append(f"{prop} {op} {val}")
        
        condition_str = " ".join(condition_parts) if any(p.strip() for p in condition_parts) else "True"

        rule = self.rules[self.selected_rule_index]
        rule['item_name'] = self.rule_name_input.text()
        rule['condition'] = condition_str
        rule['formula'] = self.formula_input.text()
        rule['type'] = self.selected_result_item['type']
        rule['item_code'] = self.selected_result_item['code']
        # Also update the name from the selected item, as it's the canonical one
        rule['item_name'] = self.selected_result_item.get('name') or self.selected_result_item.get('item_name')
        self.rule_name_input.setText(rule['item_name']) # Sync UI

        self.save_rules()
        self.populate_tree()
        QMessageBox.information(self, "Success", "Rule saved successfully.")

    def create_new_rule(self):
        """Creates a new, blank rule."""
        obj_types = list(self.property_data.keys())
        obj_type, ok = QInputDialog.getItem(self, "Create New Rule", "Select object type:", obj_types, 0, False)
        
        if ok and obj_type:
            new_rule = {
                "object": obj_type,
                "item_name": "New Unnamed Rule",
                "condition": "",
                "type": "Material",
                "item_code": "NEW-ITEM",
                "formula": "1"
            }
            self.rules.append(new_rule)
            self.populate_tree()
            self.save_rules()
            
            for i in range(self.rule_tree.topLevelItemCount()):
                parent = self.rule_tree.topLevelItem(i)
                if parent.text(0) == obj_type:
                    child = parent.child(parent.childCount() - 1)
                    if child:
                        self.rule_tree.setCurrentItem(child)
                        self.on_rule_selected(child, 0)
                    break
    
    def delete_selected_rule(self):
        """Deletes the selected rule."""
        if self.selected_rule_index == -1:
            QMessageBox.warning(self, "No Selection", "Please select a rule to delete.")
            return
        
        rule = self.rules[self.selected_rule_index]
        reply = QMessageBox.question(self, 'Delete Rule', f"Are you sure you want to delete the rule:\n'{rule.get('item_name')}'?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            del self.rules[self.selected_rule_index]
            self.selected_rule_index = -1
            self.save_rules()
            self.populate_tree()
            self.editor_panel.setDisabled(True)
            self.editor_panel.setTitle("Select a rule to edit")
    
    def filter_tree(self, text):
        """Filters the rule tree based on the search bar text."""
        for i in range(self.rule_tree.topLevelItemCount()):
            parent = self.rule_tree.topLevelItem(i)
            has_visible_child = False
            for j in range(parent.childCount()):
                child = parent.child(j)
                is_match = text.lower() in child.text(0).lower()
                child.setHidden(not is_match)
                if is_match:
                    has_visible_child = True
            parent.setHidden(not has_visible_child)

    def populate_tree(self):
        """Populates the rule tree from the loaded self.rules list."""
        current_selection = self.selected_rule_index
        self.rule_tree.clear()
        
        parent_items = {}
        item_to_select = None
        for i, rule in enumerate(self.rules):
            obj_type = rule.get("object", "Uncategorized")
            
            if obj_type not in parent_items:
                parent = QTreeWidgetItem(self.rule_tree, [obj_type])
                parent_items[obj_type] = parent
            else:
                parent = parent_items[obj_type]
            
            rule_name = rule.get("item_name", "Unnamed Rule")
            display_text = f"{rule_name}"
            child = QTreeWidgetItem(parent, [display_text])
            child.setData(0, Qt.ItemDataRole.UserRole, i)
            
            if i == current_selection:
                item_to_select = child

        self.rule_tree.expandAll()
        if item_to_select:
            self.rule_tree.setCurrentItem(item_to_select)

    def load_rules(self):
        """Loads rules from `rules.json` into memory and populates the tree."""
        try:
            with open('rules.json', 'r') as f:
                self.rules = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.rules = []
        self.populate_tree()

    def save_rules(self):
        """Saves the current set of rules to `rules.json`."""
        try:
            with open('rules.json', 'w') as f:
                json.dump(self.rules, f, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save rules.json: {e}")
